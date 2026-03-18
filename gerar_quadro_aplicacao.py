#!/usr/bin/env python3
"""
Quadro de Aplicação do Regulamento (UE) 2023/0447
Determinações individuais por artigo vs. legislação nacional portuguesa

Fontes:
  - Regulamento (UE) 2023/0447 — texto EN (ficheiro @regulamento)
  - @traducao 20230447PT_PE-CONS_PE_START.docx — texto PT oficial PE-CONS
  - @legislacao DL n.º 276/2001 e DL n.º 82/2019
  - @codigo — Código do Animal (DL n.º 214/2013)
  - @rgbeac — Regime Geral do Bem-Estar dos Animais de Companhia (jun. 2025)
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =====================================================================
# CORES
# =====================================================================
C_TOTAL   = "C8E6C9"   # verde muito claro — cobertura total
C_PARCIAL = "FFF9C4"   # amarelo muito claro — cobertura parcial
C_SEM     = "FFCDD2"   # rosa muito claro — sem correspondência
C_HEADER  = "1A237E"   # azul escuro — cabeçalho principal
C_ART     = "E8EAF6"   # azul muito claro — separador de artigo
C_SUBGRP  = "F5F5F5"   # cinza muito claro — sub-grupo (alínea)

def fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def hair_border():
    s = Side(style='hair')
    return Border(left=s, right=s, top=s, bottom=s)

# =====================================================================
# COLUNAS
# =====================================================================
COLS = [
    ("ID",                          11),
    ("Art.º",                        7),
    ("Ref. Regulamento",            22),
    ("Determinação",                30),
    ("Texto Regulamento EN\n(ipsis verbis)", 62),
    ("Texto Regulamento PT\n(ipsis verbis)", 62),
    ("Legislação Vigente\n— Texto", 54),
    ("Leg. Vig.\n— Referência",     24),
    ("Leg. Vig.\n— Cobertura",      14),
    ("@codigo\n— Texto",            50),
    ("@codigo\n— Referência",       22),
    ("@codigo\n— Cobertura",        14),
    ("@rgbeac\n— Texto",            54),
    ("@rgbeac\n— Referência",       22),
    ("@rgbeac\n— Cobertura",        14),
]

# =====================================================================
# TEXTOS VERBATIM — REGULAMENTO (UE) 2023/0447
# =====================================================================

# ---- ARTIGO 6 ----
EN_6_INTRO = ("Operators shall apply the following general welfare principles "
              "with respect to dogs or cats bred or kept in their establishment:")
PT_6_INTRO = ("Os operadores devem aplicar os seguintes princípios gerais de "
              "bem-estar no que diz respeito aos cães ou gatos criados ou detidos "
              "no seu estabelecimento:")

EN_6a = ("a) dogs and cats are provided with water and feed of a quality and "
         "quantity that affords them appropriate nutrition and hydration;")
PT_6a = ("a) Os cães e os gatos recebem água e alimentos de qualidade e numa "
         "quantidade que lhes proporciona uma nutrição e hidratação adequadas;")

EN_6b = ("b) dogs and cats are kept in a physical environment that is appropriate "
         "and regularly cleaned, that is secure and comfortable, especially in terms "
         "of space, air quality, temperature, light, protection against adverse "
         "climatic conditions and that is big enough to prevent overcrowding and "
         "to afford them ease of movement;")
PT_6b = ("b) Os cães e os gatos são detidos num ambiente físico que é adequado e "
         "limpo com regularidade, que é seguro e confortável, especialmente em "
         "termos de espaço, qualidade do ar, temperatura, luz, proteção contra "
         "condições climáticas adversas e que é suficientemente grande para prevenir "
         "a sobrelotação e para lhes proporcionar facilidade de circulação;")

EN_6c = ("c) dogs and cats are kept safe, clean and in good health, and diseases, "
         "injuries, and pain due, in particular, to management, handling practices "
         "and breeding practices, are prevented;")
PT_6c = ("c) Os cães e os gatos são detidos em segurança, limpos e em boa saúde, "
         "devendo as doenças, lesões e dores que decorrem em especial da gestão, "
         "das práticas de manuseamento e das práticas de criação, ser prevenidas;")

EN_6d = ("d) dogs and cats are kept in an environment that enables them to exhibit "
         "species-specific and social non-harmful behaviour, and to establish a "
         "positive relationship with human beings;")
PT_6d = ("d) Os cães e os gatos são detidos num ambiente que lhes permite exibir "
         "comportamentos específicos da espécie e sociais não nocivos e estabelecer "
         "uma relação positiva com os seres humanos;")

EN_6e = ("e) dogs and cats are kept in such a way as to optimise their mental state "
         "by preventing or reducing negative stimuli in duration and intensity, as "
         "well as by maximising opportunities for positive stimuli in duration and "
         "intensity, preventing the development of abnormal repetitive or other "
         "behaviours indicative of negative animal welfare, and taking into "
         "consideration the individual animal's needs in the domains referred to "
         "in points (a) to (d).")
PT_6e = ("e) Os cães e os gatos são detidos de forma a otimizar o seu estado mental, "
         "prevenindo ou reduzindo os estímulos negativos em termos de duração e "
         "intensidade, bem como maximizando as oportunidades de estímulos positivos "
         "em termos de duração e intensidade, de forma que previna o desenvolvimento "
         "de comportamentos repetitivos anormais ou de outros comportamentos "
         "indicadores de um deficiente bem-estar animal, e que tenha em conta as "
         "necessidades de cada animal nos domínios referidos nas alíneas a) a d).")

# ---- ARTIGO 7 ----
EN_7_1 = ("1. Operators shall be responsible for the welfare of dogs and cats kept "
          "in establishments under their responsibility and control, and shall be "
          "responsible for minimising any risks to the welfare of those animals.")
PT_7_1 = ("1. Os operadores são responsáveis pelo bem-estar dos cães e gatos detidos "
          "em estabelecimentos sob a sua responsabilidade e controlo e pela "
          "minimização de quaisquer riscos para o bem-estar desses animais.")

EN_7_2 = ("2. In the case of foster homes, the responsibility shall lie with the "
          "operator on whose behalf dogs or cats are kept. Such operators shall not "
          "place more than a combined total of five dogs or cats or one litter with "
          "or without mother in a foster home at any given time and shall provide "
          "the foster family with adequate information on the animal welfare "
          "obligations as well as the individual needs of the dogs or cats, and "
          "shall ensure that the relevant obligations laid down in this Regulation "
          "are complied with in foster homes.")
PT_7_2 = ("2. No caso das casas de acolhimento, a responsabilidade incumbe aos "
          "operadores em cujo nome os cães ou gatos são detidos. Tais operadores "
          "não devem, em nenhum momento, colocar mais do que um total combinado "
          "de cinco cães ou gatos, ou mais do que uma ninhada, com ou sem mãe, "
          "numa casa de acolhimento, e devem fornecer à família de acolhimento "
          "as informações adequadas sobre as obrigações em matéria de bem-estar "
          "dos animais, assim como sobre as necessidades individuais dos cães ou "
          "gatos, devendo ainda assegurar que nessas casas de acolhimento sejam "
          "respeitadas as obrigações pertinentes previstas no presente regulamento.")

EN_7_3 = ("3. The Member State in which the foster home is located may allow a "
          "greater number of dogs, cats or litters to be placed in the foster home, "
          "provided that the premises of the foster home have sufficient space, "
          "including outdoor space, and that the number of animal carers in the "
          "foster home is sufficient, to ensure the welfare of the dogs or cats.")
PT_7_3 = ("O Estado-Membro em que se situa a casa de acolhimento pode autorizar "
          "que um maior número de cães, gatos ou ninhadas seja colocado na casa "
          "de acolhimento, desde que as instalações desta disponham de espaço "
          "suficiente, incluindo espaço exterior, e que o número de cuidadores "
          "de animais na referida casa seja suficiente para assegurar o bem-estar "
          "dos cães ou gatos.")

EN_7_4 = ("4. Operators shall not subject any dog or cat to cruelty, abuse or "
          "mistreatment, including by making them participate in activities likely "
          "to result in cruelty to or abuse or mistreatment of the dogs or cats "
          "bred or kept by the operator.")
PT_7_4 = ("3. Os operadores não podem sujeitar nenhum cão ou gato a crueldade, "
          "abuso ou maus-tratos, nomeadamente fazendo-os participar em atividades "
          "suscetíveis de resultar em atos de crueldade, abuso ou maus-tratos "
          "dos cães ou gatos criados ou detidos pelo operador.")

EN_7_5 = "5. Operators shall not abandon the dogs or cats bred or kept by them."
PT_7_5 = "4. Os operadores não devem abandonar os cães ou gatos por eles criados ou detidos."

EN_7_6 = ("6. Before operators cease activities at an establishment, they shall "
          "ensure that the dogs or cats kept there are rehomed, either by taking "
          "up the pet ownership themselves or by transferring the responsibility "
          "for, or the ownership of, the dogs and cats to other operators or "
          "acquirers.")
PT_7_6 = ("5. Antes de cessarem as suas atividades num estabelecimento, os "
          "operadores devem garantir que os cães ou gatos aí detidos sejam "
          "realojados, quer assumindo eles próprios a propriedade dos animais "
          "de companhia, quer transferindo a responsabilidade pelos cães e gatos, "
          "ou a propriedade dos mesmos, para outros operadores ou adquirentes.")

EN_7_7 = ("7. Operators shall ensure that dogs and cats are handled by a number "
          "of animal carers sufficient to meet the welfare needs of dogs or cats "
          "kept in their establishments, and that those carers have the competences "
          "required under Article 12.")
PT_7_7 = ("6. Os operadores devem assegurar que os cães e gatos sejam manuseados "
          "por um número de cuidadores de animais suficiente para satisfazer as "
          "necessidades de bem-estar dos cães ou dos gatos que são detidos nos "
          "seus estabelecimentos, e que esses cuidadores possuam as competências "
          "exigidas nos termos do artigo 12.º")

EN_7_8 = ("8. Operators shall ensure the welfare of the dogs or cats for which "
          "they are responsible by monitoring animal-based indicators concerning "
          "behaviour and physical appearance, and by taking actions based on the "
          "results of such monitoring.")
PT_7_8 = ("7. Os operadores devem assegurar o bem-estar dos cães ou dos gatos "
          "pelos quais são responsáveis, fazendo uso dos indicadores relativos "
          "ao comportamento e aspeto físico baseados nos animais e tomando "
          "medidas em função do resultado dessa monitorização.")

EN_7_9 = ("9. The Commission is empowered to adopt delegated acts in accordance "
          "with Article 28 supplementing this Regulation by laying down the "
          "animal-based indicators concerning behaviour and physical appearance "
          "that operators are to use for monitoring, in accordance with paragraph "
          "8 of this Article, and the methods by which operators are to measure them.")
PT_7_9 = ("8. A Comissão fica habilitada a adotar atos delegados em conformidade "
          "com o artigo 28.º para complementar o presente regulamento, prevendo "
          "indicadores relativos ao comportamento e aspeto físico baseados nos "
          "animais, que devem ser usados pelos operadores para efeitos da "
          "monitorização a realizar em conformidade com o n.º 7 do presente artigo, "
          "assim como os métodos a usar pelos operadores para a sua medição.")

# =====================================================================
# TEXTOS VERBATIM — LEGISLAÇÃO VIGENTE (DL 276/2001)
# =====================================================================
DL276_7_1 = ("1 — As condições de detenção e de alojamento para reprodução, "
             "criação, manutenção e acomodação dos animais de companhia devem "
             "salvaguardar os seus parâmetros de bem-estar animal, nomeadamente "
             "nos termos dos artigos seguintes.")
DL276_7_2 = ("2 — Nenhum animal deve ser detido como animal de companhia se não "
             "estiverem asseguradas as condições referidas no número anterior ou "
             "se não se adaptar ao cativeiro.")
DL276_7_3 = ("3 — São proibidas todas as violências contra animais, considerando-se "
             "como tais os atos consistentes em, sem necessidade, se infligir a "
             "morte, o sofrimento ou lesões a um animal.")
DL276_7_4 = ("4 — É proibido utilizar animais para fins didáticos e lúdicos, de "
             "treino, filmagens, exibições, publicidade ou atividades semelhantes, "
             "na medida em que daí resultem para eles dor ou sofrimentos "
             "consideráveis, salvo experiência científica de comprovada necessidade "
             "e justificada nos termos da lei.")

# =====================================================================
# TEXTOS VERBATIM — @codigo (DL n.º 214/2013)
# =====================================================================
COD_4 = ("Cabe aos detentores de animais de companhia assegurar as necessidades "
         "básicas de bem-estar dos mesmos, garantir o controlo da sua reprodução, "
         "salvaguardar a sua saúde e prevenir os riscos inerentes à transmissão "
         "de doenças a pessoas e a outros animais e, ainda, a segurança das "
         "populações, garantindo a salubridade dos locais e a tranquilidade "
         "das pessoas.")
COD_5_1 = ("1 — As condições de detenção e de alojamento para reprodução, criação, "
           "manutenção e acomodação dos animais de companhia devem salvaguardar "
           "os seus parâmetros de bem-estar animal.")
COD_5_2 = ("2 — Nenhum animal deve ser detido como animal de companhia se não "
           "estiverem asseguradas as condições referidas no número anterior ou "
           "se não se adaptar ao cativeiro.")
COD_5_3 = ("3 — É proibida a violência contra animais, considerando-se como tal "
           "todos os atos que, sem necessidade, infligem a morte, o sofrimento, "
           "abuso ou lesões a um animal, designadamente: agredir animais; restringir "
           "a liberdade de movimentos de tal forma que lhes seja impedido levantar-se, "
           "deitar-se ou virar-se; usar equipamentos que causem sofrimento "
           "desnecessário; incitar, realizar ou promover lutas entre animais; criar "
           "cães e gatos para consumo; promover ou concretizar o sacrifício ritual.")
COD_5_4 = ("4 — É proibido utilizar animais para fins didáticos e lúdicos, de treino, "
           "filmagens, exibições, publicidade ou atividades semelhantes, sempre que "
           "daí resultem para aqueles dor ou sofrimentos consideráveis, salvo "
           "experiência científica de comprovada necessidade e justificada nos "
           "termos da lei.")
COD_9 = ("Considera-se abandono de animais de companhia a não prestação de cuidados "
         "no alojamento, bem como a remoção dos animais, efetuada pelos seus "
         "detentores, para fora do domicílio ou dos locais onde costumam estar "
         "mantidos, com vista a pôr termo à sua detenção, sem que procedam à sua "
         "transmissão para a guarda e responsabilidade de outras pessoas, das "
         "autarquias locais ou de associações de proteção dos animais legalmente "
         "constituídas.")

# =====================================================================
# TEXTOS VERBATIM — @rgbeac (jun. 2025)
# =====================================================================
RGBEAC_7_1 = ("1 — Todos têm o dever de garantir condições de bem-estar aos animais "
              "de companhia, em especial que:\n"
              "a) não passem fome ou sede, nem sejam sujeitos a malnutrição;\n"
              "b) não sejam expostos a situações de desconforto físico ou térmico;\n"
              "c) não sofram dor, lesão física ou doença;\n"
              "d) possam expressar padrões normais de comportamento; e\n"
              "e) não sejam colocados em situações que lhes provoquem stresse, "
              "medo ou ansiedade injustificados.")
RGBEAC_7_1a = "a) não passem fome ou sede, nem sejam sujeitos a malnutrição;"
RGBEAC_7_1b = "b) não sejam expostos a situações de desconforto físico ou térmico;"
RGBEAC_7_1c = "c) não sofram dor, lesão física ou doença;"
RGBEAC_7_1d = "d) possam expressar padrões normais de comportamento;"
RGBEAC_7_1e = ("e) não sejam colocados em situações que lhes provoquem stresse, "
               "medo ou ansiedade injustificados.")
RGBEAC_7_2 = ("2 — As condições de detenção e de alojamento dos animais de companhia "
              "devem salvaguardar os seus parâmetros de bem-estar animal e a expressão "
              "dos comportamentos naturais da espécie.")
RGBEAC_7_3 = ("3 — Nenhum animal deve ser detido como animal de companhia se não "
              "estiverem asseguradas as condições referidas no presente artigo.")

RGBEAC_10_1_INTRO = ("1 — O detentor do animal de companhia deve:")
RGBEAC_10_1a = ("a) Assegurar o bem-estar do animal, de acordo com sua espécie, "
                "raça, idade e necessidades físicas e etológicas, proporcionando-lhe:\n"
                "— Atenção, supervisão, controlo, exercício físico e estímulo mental;\n"
                "— Alimentos saudáveis, adequados e convenientes ao seu normal "
                "desenvolvimento e acesso permanente a água potável;\n"
                "— Condições higiossanitárias que atendam, no mínimo, ao estabelecido "
                "no presente decreto-lei e na demais legislação aplicável;\n"
                "— Liberdade de movimento, sendo proibidos todos os sistemas de "
                "contenção permanentes e, no caso de animais que, pelas características "
                "da espécie ou comportamento, tenham de ser mantidos em canis, espaços "
                "vedados ou outros, sendo obrigatório que os mesmos disponham de espaço "
                "e enriquecimento ambiental adequados para garantir o seu bem-estar e a "
                "expressão dos seus comportamentos naturais;\n"
                "— Abrigo adequado, em termos de tamanho e qualidade, com vista a "
                "proteger de condições atmosféricas adversas, incluindo frio, chuva, "
                "sol ou calor excessivos, com cama seca, limpa e confortável;\n"
                "— Contacto social adequado a cada espécie, de acordo com a sua idade "
                "e atividade.")
RGBEAC_10_1a_agua = ("a) [...] — Alimentos saudáveis, adequados e convenientes ao seu "
                     "normal desenvolvimento e acesso permanente a água potável;")
RGBEAC_10_1a_lib = ("a) [...] — Liberdade de movimento, sendo proibidos todos os "
                    "sistemas de contenção permanentes [...] sendo obrigatório que os "
                    "mesmos disponham de espaço e enriquecimento ambiental adequados "
                    "para garantir o seu bem-estar e a expressão dos seus "
                    "comportamentos naturais [...]")
RGBEAC_10_1a_abr = ("a) [...] — Abrigo adequado, em termos de tamanho e qualidade, "
                    "com vista a proteger de condições atmosféricas adversas, incluindo "
                    "frio, chuva, sol ou calor excessivos, com cama seca, limpa e "
                    "confortável;")
RGBEAC_10_1a_hig = ("a) [...] — Condições higiossanitárias que atendam, no mínimo, "
                    "ao estabelecido no presente decreto-lei e na demais legislação "
                    "aplicável;")
RGBEAC_10_1c = ("c) Educar o animal com recurso a métodos de reforço positivo visando "
               "a sua vinculação e integração positivas no espaço familiar e no meio "
               "ambiente;")

RGBEAC_12_1 = "1 — São interditas quaisquer práticas que violem o bem-estar dos animais de companhia, em especial as seguintes:"
RGBEAC_12_1b = ("b) Sujeitá-los a qualquer prática que lhes possa causar sofrimento, "
               "dano, stresse ou angústia, por ação ou omissão, salvo nas situações "
               "legalmente previstas;")
RGBEAC_12_1e = "e) Abandoná-los;"
RGBEAC_12_1u = ("u) Utilizá-los em eventos, confrontos entre animais, ou outras "
               "atividades que envolvam crueldade ou tratamento humilhante, "
               "maus-tratos, sofrimento ou morte, ou coloquem em risco a sua "
               "saúde e bem-estar;")

RGBEAC_FAM_DEF = ("«Família de acolhimento temporário» o detentor no âmbito de uma "
                  "medida transitória de proteção e salvaguarda do bem-estar de animais "
                  "de companhia, que recebe temporariamente um animal de companhia em "
                  "nome de um operador. [definição — art.º 4.º]")

# =====================================================================
# MARCADORES DE SEM CORRESPONDÊNCIA
# =====================================================================
SEM_DL276 = ("Sem correspondência — DL n.º 276/2001 não contempla este "
             "requisito específico.")
SEM_DL276_OP = ("Sem correspondência — o conceito de «operador» (estabelecimento "
                "profissional) é distinto do «detentor» regulado pelo DL n.º 276/2001; "
                "sem previsão específica de responsabilidade de operador.")
SEM_COD  = "Sem correspondência — @codigo não contempla este requisito."
SEM_RG   = "Sem correspondência — @rgbeac não contempla este requisito."
SEM_ALL  = "Sem correspondência — conceito inexistente na legislação nacional."
SEM_ACOLH_LV = ("Sem correspondência — conceito de «casa de acolhimento» com limites "
                "numéricos e responsabilidade atribuída ao operador é inexistente "
                "na legislação vigente.")
SEM_ACOLH_COD = ("Sem correspondência — @codigo não prevê o regime de famílias de "
                 "acolhimento nem a atribuição de responsabilidade ao operador.")
SEM_META  = ("Sem correspondência — disposição de habilitação regulatória europeia "
             "sem equivalente na legislação nacional.")

REF_NONE = "—"

# =====================================================================
# LINHAS DE DADOS
# =====================================================================
# Cada tuplo: (id, artigo, ref_reg, determinacao,
#              en_text, pt_text,
#              lv_texto, lv_ref, lv_cob,
#              cod_texto, cod_ref, cod_cob,
#              rg_texto,  rg_ref,  rg_cob)
#
# lv_cob / cod_cob / rg_cob ∈ {"Total", "Parcial", "Sem correspondência"}

T = "Total"
P = "Parcial"
S = "Sem correspondência"

ROWS = [

    # ================================================================
    # ARTIGO 6 — Princípios gerais de bem-estar
    # ================================================================

    # -- alínea a) --
    ("ART-06-a-1", "Art.º 6.º", "al. a), art.º 6.º",
     "Água de qualidade",
     EN_6a, PT_6a,
     DL276_7_1,
     "n.º 1, art.º 7.º do DL n.º 276/2001, de 17 de outubro",
     P,
     COD_5_1,
     "n.º 1, art.º 5.º do @codigo",
     P,
     RGBEAC_10_1a_agua,
     "al. a), n.º 1, art.º 10.º do @rgbeac",
     T),

    ("ART-06-a-2", "Art.º 6.º", "al. a), art.º 6.º",
     "Alimentos de qualidade",
     EN_6a, PT_6a,
     DL276_7_1,
     "n.º 1, art.º 7.º do DL n.º 276/2001, de 17 de outubro",
     P,
     COD_5_1,
     "n.º 1, art.º 5.º do @codigo",
     P,
     RGBEAC_10_1a_agua,
     "al. a), n.º 1, art.º 10.º do @rgbeac",
     T),

    ("ART-06-a-3", "Art.º 6.º", "al. a), art.º 6.º",
     "Quantidade adequada — nutrição",
     EN_6a, PT_6a,
     DL276_7_1 + "\n\n" + COD_4[:120] + " [...]",
     "n.º 1, art.º 7.º do DL n.º 276/2001\nart.º 4.º do @codigo (por remissão geral)",
     P,
     RGBEAC_7_1a + "\n\n" + COD_4[:120] + " [...]",
     "art.º 4.º e n.º 1, art.º 5.º do @codigo",
     P,
     RGBEAC_7_1a + "\n\n" + RGBEAC_10_1a_agua,
     "al. a), n.º 1, art.º 7.º e al. a), n.º 1, art.º 10.º do @rgbeac",
     P),

    ("ART-06-a-4", "Art.º 6.º", "al. a), art.º 6.º",
     "Quantidade adequada — hidratação",
     EN_6a, PT_6a,
     DL276_7_1,
     "n.º 1, art.º 7.º do DL n.º 276/2001, de 17 de outubro",
     P,
     COD_5_1,
     "n.º 1, art.º 5.º do @codigo",
     P,
     RGBEAC_7_1a + "\n\n" + RGBEAC_10_1a_agua,
     "al. a), n.º 1, art.º 7.º e al. a), n.º 1, art.º 10.º do @rgbeac",
     P),

    # -- alínea b) --
    ("ART-06-b-1", "Art.º 6.º", "al. b), art.º 6.º",
     "Ambiente físico adequado",
     EN_6b, PT_6b,
     DL276_7_1,
     "n.º 1, art.º 7.º do DL n.º 276/2001, de 17 de outubro",
     P,
     COD_5_1,
     "n.º 1, art.º 5.º do @codigo",
     P,
     RGBEAC_7_2,
     "n.º 2, art.º 7.º do @rgbeac",
     P),

    ("ART-06-b-2", "Art.º 6.º", "al. b), art.º 6.º",
     "Regularmente limpo",
     EN_6b, PT_6b,
     DL276_7_1,
     "n.º 1, art.º 7.º do DL n.º 276/2001, de 17 de outubro",
     P,
     COD_5_1,
     "n.º 1, art.º 5.º do @codigo",
     P,
     RGBEAC_10_1a_hig,
     "al. a), n.º 1, art.º 10.º do @rgbeac",
     P),

    ("ART-06-b-3", "Art.º 6.º", "al. b), art.º 6.º",
     "Seguro",
     EN_6b, PT_6b,
     DL276_7_1,
     "n.º 1, art.º 7.º do DL n.º 276/2001, de 17 de outubro",
     P,
     COD_5_1,
     "n.º 1, art.º 5.º do @codigo",
     P,
     RGBEAC_7_2,
     "n.º 2, art.º 7.º do @rgbeac",
     P),

    ("ART-06-b-4", "Art.º 6.º", "al. b), art.º 6.º",
     "Confortável: espaço, qualidade do ar, temperatura, luz, proteção climática",
     EN_6b, PT_6b,
     DL276_7_1,
     "n.º 1, art.º 7.º do DL n.º 276/2001, de 17 de outubro",
     P,
     COD_5_1,
     "n.º 1, art.º 5.º do @codigo",
     P,
     RGBEAC_10_1a_abr,
     "al. a), n.º 1, art.º 10.º do @rgbeac",
     P),

    ("ART-06-b-5", "Art.º 6.º", "al. b), art.º 6.º",
     "Facilidade de movimentação / circulação",
     EN_6b, PT_6b,
     SEM_DL276 + " O DL n.º 276/2001 não prevê explicitamente a facilidade de"
     " movimentação como requisito autónomo de bem-estar.",
     REF_NONE,
     S,
     SEM_COD,
     REF_NONE,
     S,
     RGBEAC_10_1a_lib,
     "al. a), n.º 1, art.º 10.º do @rgbeac",
     P),

    ("ART-06-b-6", "Art.º 6.º", "al. b), art.º 6.º",
     "Prevenção de sobrelotação",
     EN_6b, PT_6b,
     "Sem correspondência — DL n.º 276/2001 prevê regras de número máximo de"
     " animais por habitação (art.º 11.º), não sobrelotação como requisito de"
     " bem-estar do espaço físico.",
     REF_NONE,
     S,
     SEM_COD,
     REF_NONE,
     S,
     "Sem correspondência — @rgbeac define capacidade máxima de alojamentos"
     " (art.º 11.º) mas não especifica sobrelotação como requisito autónomo"
     " de bem-estar físico.",
     REF_NONE,
     S),

    # -- alínea c) --
    ("ART-06-c-1", "Art.º 6.º", "al. c), art.º 6.º",
     "Mantidos seguros",
     EN_6c, PT_6c,
     DL276_7_1,
     "n.º 1, art.º 7.º do DL n.º 276/2001, de 17 de outubro",
     P,
     COD_5_1,
     "n.º 1, art.º 5.º do @codigo",
     P,
     RGBEAC_7_2,
     "n.º 2, art.º 7.º do @rgbeac",
     P),

    ("ART-06-c-2", "Art.º 6.º", "al. c), art.º 6.º",
     "Mantidos limpos",
     EN_6c, PT_6c,
     DL276_7_1,
     "n.º 1, art.º 7.º do DL n.º 276/2001, de 17 de outubro",
     P,
     COD_5_1,
     "n.º 1, art.º 5.º do @codigo",
     P,
     RGBEAC_10_1a_hig,
     "al. a), n.º 1, art.º 10.º do @rgbeac",
     P),

    ("ART-06-c-3", "Art.º 6.º", "al. c), art.º 6.º",
     "Mantidos com boa saúde",
     EN_6c, PT_6c,
     DL276_7_1 + "\n\n" + COD_4[:80] + " [...] salvaguardar a sua saúde [...]",
     "n.º 1, art.º 7.º do DL n.º 276/2001\nart.º 4.º do @codigo (por remissão)",
     P,
     COD_4[:80] + " [...] salvaguardar a sua saúde [...]",
     "art.º 4.º do @codigo",
     P,
     RGBEAC_7_1c,
     "al. c), n.º 1, art.º 7.º do @rgbeac",
     T),

    ("ART-06-c-4", "Art.º 6.º", "al. c), art.º 6.º",
     "Prevenção de doenças (maneio, manuseamento e reprodução)",
     EN_6c, PT_6c,
     DL276_7_1,
     "n.º 1, art.º 7.º do DL n.º 276/2001, de 17 de outubro",
     P,
     COD_4[:80] + " [...] prevenir os riscos inerentes à transmissão de doenças [...]",
     "art.º 4.º do @codigo",
     P,
     RGBEAC_7_1c,
     "al. c), n.º 1, art.º 7.º do @rgbeac",
     T),

    ("ART-06-c-5", "Art.º 6.º", "al. c), art.º 6.º",
     "Prevenção de lesões",
     EN_6c, PT_6c,
     DL276_7_3,
     "n.º 3, art.º 7.º do DL n.º 276/2001, de 17 de outubro",
     P,
     COD_5_3[:120] + " [...]",
     "n.º 3, art.º 5.º do @codigo",
     P,
     RGBEAC_7_1c,
     "al. c), n.º 1, art.º 7.º do @rgbeac",
     T),

    ("ART-06-c-6", "Art.º 6.º", "al. c), art.º 6.º",
     "Prevenção de dor",
     EN_6c, PT_6c,
     DL276_7_3,
     "n.º 3, art.º 7.º do DL n.º 276/2001, de 17 de outubro",
     P,
     COD_5_3[:120] + " [...]",
     "n.º 3, art.º 5.º do @codigo",
     P,
     RGBEAC_7_1c,
     "al. c), n.º 1, art.º 7.º do @rgbeac",
     T),

    # -- alínea d) --
    ("ART-06-d-1", "Art.º 6.º", "al. d), art.º 6.º",
     "Comportamentos específicos da espécie",
     EN_6d, PT_6d,
     DL276_7_1 + " (parâmetros etológicos próprios da espécie)",
     "n.º 1, art.º 7.º do DL n.º 276/2001, de 17 de outubro",
     P,
     COD_5_1,
     "n.º 1, art.º 5.º do @codigo",
     P,
     RGBEAC_7_1d + "\n\n" + RGBEAC_7_2,
     "al. d), n.º 1, art.º 7.º e n.º 2, art.º 7.º do @rgbeac",
     T),

    ("ART-06-d-2", "Art.º 6.º", "al. d), art.º 6.º",
     "Comportamentos sociais não nocivos",
     EN_6d, PT_6d,
     DL276_7_1,
     "n.º 1, art.º 7.º do DL n.º 276/2001, de 17 de outubro",
     P,
     COD_5_1,
     "n.º 1, art.º 5.º do @codigo",
     P,
     "a) [...] — Contacto social adequado a cada espécie, de acordo com a sua"
     " idade e atividade.",
     "al. a), n.º 1, art.º 10.º do @rgbeac",
     P),

    ("ART-06-d-3", "Art.º 6.º", "al. d), art.º 6.º",
     "Relação positiva com seres humanos [NOVO — sem correspondência]",
     EN_6d, PT_6d,
     "Sem correspondência — o requisito de «estabelecer uma relação positiva"
     " com os seres humanos» é um conceito novo que não encontra paralelo no"
     " DL n.º 276/2001 nem na restante legislação vigente.",
     REF_NONE,
     S,
     "Sem correspondência — @codigo não contempla o requisito de estabelecer"
     " uma relação positiva com seres humanos.",
     REF_NONE,
     S,
     RGBEAC_10_1c,
     "al. c), n.º 1, art.º 10.º do @rgbeac",
     P),

    # -- alínea e) --
    ("ART-06-e-1", "Art.º 6.º", "al. e), art.º 6.º",
     "Otimizar estado mental — prevenir/reduzir estímulos negativos",
     EN_6e, PT_6e,
     "Sem correspondência — DL n.º 276/2001 refere-se a parâmetros fisiológicos"
     " e etológicos mas não contempla a gestão do estado mental nem a redução"
     " de estímulos negativos.",
     REF_NONE,
     S,
     SEM_COD,
     REF_NONE,
     S,
     RGBEAC_7_1e,
     "al. e), n.º 1, art.º 7.º do @rgbeac",
     P),

    ("ART-06-e-2", "Art.º 6.º", "al. e), art.º 6.º",
     "Maximizar oportunidades de estímulos positivos",
     EN_6e, PT_6e,
     "Sem correspondência — conceito de maximização de estímulos positivos"
     " inexistente na legislação vigente.",
     REF_NONE,
     S,
     SEM_COD,
     REF_NONE,
     S,
     "a) [...] — Atenção, supervisão, controlo, exercício físico e estímulo"
     " mental;",
     "al. a), n.º 1, art.º 10.º do @rgbeac",
     P),

    ("ART-06-e-3", "Art.º 6.º", "al. e), art.º 6.º",
     "Prevenir comportamentos repetitivos anormais [NOVO]",
     EN_6e, PT_6e,
     "Sem correspondência — conceito de prevenção de comportamentos repetitivos"
     " anormais como indicador de bem-estar é inexistente na legislação vigente.",
     REF_NONE,
     S,
     SEM_COD,
     REF_NONE,
     S,
     "Sem correspondência — @rgbeac não contempla especificamente a prevenção"
     " de comportamentos repetitivos anormais como indicador de bem-estar.",
     REF_NONE,
     S),

    ("ART-06-e-4", "Art.º 6.º", "al. e), art.º 6.º",
     "Ter em conta necessidades individuais do animal (nos domínios a) a d))",
     EN_6e, PT_6e,
     "Sem correspondência — legislação vigente regula condições gerais de"
     " alojamento; não prevê a individualização das necessidades por animal.",
     REF_NONE,
     S,
     SEM_COD,
     REF_NONE,
     S,
     "1 — O detentor do animal de companhia deve:\na) Assegurar o bem-estar"
     " do animal, de acordo com sua espécie, raça, idade e necessidades"
     " físicas e etológicas [...]",
     "n.º 1 e al. a), n.º 1, art.º 10.º do @rgbeac",
     P),

    # ================================================================
    # ARTIGO 7 — Obrigações gerais em matéria de bem-estar
    # ================================================================

    # -- n.º 1 --
    ("ART-07-1-1", "Art.º 7.º", "n.º 1, art.º 7.º",
     "Responsabilidade dos operadores pelo bem-estar dos animais nos estabelecimentos",
     EN_7_1, PT_7_1,
     DL276_7_1 + "\n\nNota: DL n.º 276/2001 regula o «detentor» (conceito mais"
     " amplo que «operador»). Não existe disposição específica de responsabilidade"
     " do operador profissional.",
     "n.º 1, art.º 7.º do DL n.º 276/2001, de 17 de outubro",
     P,
     COD_4[:120] + " [...]",
     "art.º 4.º do @codigo",
     P,
     RGBEAC_10_1_INTRO + "\na) Assegurar o bem-estar do animal, de acordo com"
     " sua espécie, raça, idade e necessidades físicas e etológicas [...]",
     "n.º 1 e al. a), n.º 1, art.º 10.º do @rgbeac",
     P),

    ("ART-07-1-2", "Art.º 7.º", "n.º 1, art.º 7.º",
     "Minimizar riscos para o bem-estar",
     EN_7_1, PT_7_1,
     "Sem correspondência — DL n.º 276/2001 não impõe dever explícito de"
     " minimização de riscos; estabelece condições gerais e proibições.",
     REF_NONE,
     S,
     "Sem correspondência — @codigo não prevê o dever explícito de"
     " minimização de riscos.",
     REF_NONE,
     S,
     "Sem correspondência — @rgbeac não prevê expressamente o dever de"
     " minimização de riscos, embora as obrigações gerais o pressuponham.",
     REF_NONE,
     S),

    # -- n.º 2 (casas de acolhimento) --
    ("ART-07-2-1", "Art.º 7.º", "n.º 2, art.º 7.º",
     "Responsabilidade do operador nas casas de acolhimento (não da família)",
     EN_7_2, PT_7_2,
     SEM_ACOLH_LV,
     REF_NONE,
     S,
     SEM_ACOLH_COD,
     REF_NONE,
     S,
     RGBEAC_FAM_DEF + "\n\nConceito previsto na definição mas sem atribuição"
     " específica de responsabilidade ao operador.",
     "art.º 4.º (definições) do @rgbeac",
     P),

    ("ART-07-2-2", "Art.º 7.º", "n.º 2, art.º 7.º",
     "Limite: máx. 5 cães/gatos ou 1 ninhada por casa de acolhimento",
     EN_7_2, PT_7_2,
     SEM_ACOLH_LV,
     REF_NONE,
     S,
     SEM_ACOLH_COD,
     REF_NONE,
     S,
     "Sem correspondência — @rgbeac define «família de acolhimento temporário»"
     " mas não fixa o limite numérico de 5 animais ou 1 ninhada.",
     REF_NONE,
     S),

    ("ART-07-2-3", "Art.º 7.º", "n.º 2, art.º 7.º",
     "Fornecer à família informação sobre obrigações de bem-estar",
     EN_7_2, PT_7_2,
     SEM_ACOLH_LV,
     REF_NONE,
     S,
     SEM_ACOLH_COD,
     REF_NONE,
     S,
     SEM_RG,
     REF_NONE,
     S),

    ("ART-07-2-4", "Art.º 7.º", "n.º 2, art.º 7.º",
     "Fornecer à família informação sobre necessidades individuais dos animais",
     EN_7_2, PT_7_2,
     SEM_ACOLH_LV,
     REF_NONE,
     S,
     SEM_ACOLH_COD,
     REF_NONE,
     S,
     SEM_RG,
     REF_NONE,
     S),

    ("ART-07-2-5", "Art.º 7.º", "n.º 2, art.º 7.º",
     "Assegurar cumprimento do Regulamento nas casas de acolhimento",
     EN_7_2, PT_7_2,
     SEM_ACOLH_LV,
     REF_NONE,
     S,
     SEM_ACOLH_COD,
     REF_NONE,
     S,
     SEM_RG,
     REF_NONE,
     S),

    # -- n.º 3 (derogação Estado-Membro) --
    ("ART-07-3-1", "Art.º 7.º", "n.º 3, art.º 7.º",
     "Estado-Membro pode autorizar número maior de animais na casa de acolhimento",
     EN_7_3, PT_7_3,
     SEM_META,
     REF_NONE,
     S,
     SEM_META,
     REF_NONE,
     S,
     SEM_META,
     REF_NONE,
     S),

    ("ART-07-3-2", "Art.º 7.º", "n.º 3, art.º 7.º",
     "Condição de derogação: espaço suficiente (incluindo exterior)",
     EN_7_3, PT_7_3,
     SEM_META,
     REF_NONE,
     S,
     SEM_META,
     REF_NONE,
     S,
     SEM_META,
     REF_NONE,
     S),

    ("ART-07-3-3", "Art.º 7.º", "n.º 3, art.º 7.º",
     "Condição de derogação: número suficiente de cuidadores",
     EN_7_3, PT_7_3,
     SEM_META,
     REF_NONE,
     S,
     SEM_META,
     REF_NONE,
     S,
     SEM_META,
     REF_NONE,
     S),

    # -- n.º 4 (crueldade) --
    ("ART-07-4-1", "Art.º 7.º", "n.º 4, art.º 7.º",
     "Proibição de crueldade",
     EN_7_4, PT_7_4,
     DL276_7_3,
     "n.º 3, art.º 7.º do DL n.º 276/2001, de 17 de outubro",
     P,
     COD_5_3[:150] + " [...]",
     "n.º 3, art.º 5.º do @codigo",
     P,
     RGBEAC_12_1 + "\n" + RGBEAC_12_1b,
     "n.º 1 e al. b), n.º 1, art.º 12.º do @rgbeac",
     T),

    ("ART-07-4-2", "Art.º 7.º", "n.º 4, art.º 7.º",
     "Proibição de abusos",
     EN_7_4, PT_7_4,
     DL276_7_3 + "\n\nNota: DL n.º 276/2001 usa o termo «violências»; não"
     " distingue «crueldade» de «abusos».",
     "n.º 3, art.º 7.º do DL n.º 276/2001, de 17 de outubro",
     P,
     "3 — É proibida a violência contra animais, considerando-se como tal"
     " todos os atos que, sem necessidade, infligem a morte, o sofrimento,"
     " abuso ou lesões a um animal [...]",
     "n.º 3, art.º 5.º do @codigo",
     T,
     RGBEAC_12_1 + "\n" + RGBEAC_12_1b,
     "n.º 1 e al. b), n.º 1, art.º 12.º do @rgbeac",
     T),

    ("ART-07-4-3", "Art.º 7.º", "n.º 4, art.º 7.º",
     "Proibição de maus-tratos",
     EN_7_4, PT_7_4,
     DL276_7_3,
     "n.º 3, art.º 7.º do DL n.º 276/2001, de 17 de outubro",
     P,
     COD_5_3[:150] + " [...]",
     "n.º 3, art.º 5.º do @codigo",
     P,
     RGBEAC_12_1 + "\n" + RGBEAC_12_1b,
     "n.º 1 e al. b), n.º 1, art.º 12.º do @rgbeac",
     T),

    ("ART-07-4-4", "Art.º 7.º", "n.º 4, art.º 7.º",
     "Proibição de participação em atividades causadoras de crueldade/abusos/maus-tratos",
     EN_7_4, PT_7_4,
     DL276_7_4,
     "n.º 4, art.º 7.º do DL n.º 276/2001, de 17 de outubro",
     P,
     COD_5_4,
     "n.º 4, art.º 5.º do @codigo",
     P,
     RGBEAC_12_1u,
     "al. u), n.º 1, art.º 12.º do @rgbeac",
     T),

    # -- n.º 5 (abandono) --
    ("ART-07-5-1", "Art.º 7.º", "n.º 5, art.º 7.º",
     "Proibição de abandono",
     EN_7_5, PT_7_5,
     DL276_7_3 + "\n\nNota: Abandono é punível nos termos do art.º 68.º,"
     " n.º 2, al. c) do DL n.º 276/2001 e da Lei n.º 92/95, de 12 de setembro.",
     "n.º 3, art.º 7.º e art.º 68.º do DL n.º 276/2001",
     T,
     COD_9,
     "art.º 9.º do @codigo",
     T,
     RGBEAC_12_1e,
     "al. e), n.º 1, art.º 12.º do @rgbeac",
     T),

    # -- n.º 6 (recolocação) --
    ("ART-07-6-1", "Art.º 7.º", "n.º 6, art.º 7.º",
     "Garantir recolocação dos animais antes de cessação de atividade",
     EN_7_6, PT_7_6,
     "Sem correspondência — DL n.º 276/2001 não prevê dever de recolocação"
     " como condição prévia ao encerramento de estabelecimento.",
     REF_NONE,
     S,
     SEM_COD,
     REF_NONE,
     S,
     "Sem correspondência — @rgbeac não prevê este dever explícito de"
     " recolocação prévia ao encerramento de estabelecimento.",
     REF_NONE,
     S),

    ("ART-07-6-2", "Art.º 7.º", "n.º 6, art.º 7.º",
     "Modalidades de recolocação: assumir propriedade ou transferir",
     EN_7_6, PT_7_6,
     SEM_DL276,
     REF_NONE,
     S,
     SEM_COD,
     REF_NONE,
     S,
     SEM_RG,
     REF_NONE,
     S),

    # -- n.º 7 (cuidadores) --
    ("ART-07-7-1", "Art.º 7.º", "n.º 7, art.º 7.º",
     "Número suficiente de cuidadores de animais",
     EN_7_7, PT_7_7,
     "Sem correspondência — DL n.º 276/2001 não estabelece rácios de"
     " cuidadores por animal ou por estabelecimento.",
     REF_NONE,
     S,
     SEM_COD,
     REF_NONE,
     S,
     "a) [...] — Atenção, supervisão, controlo [...]",
     "al. a), n.º 1, art.º 10.º do @rgbeac",
     P),

    ("ART-07-7-2", "Art.º 7.º", "n.º 7, art.º 7.º",
     "Cuidadores com competências exigidas (art.º 12.º do Regulamento)",
     EN_7_7, PT_7_7,
     "Sem correspondência — DL n.º 276/2001 não define competências"
     " exigidas aos tratadores de animais em estabelecimentos.",
     REF_NONE,
     S,
     SEM_COD,
     REF_NONE,
     S,
     "Sem correspondência — @rgbeac não define, neste artigo, os requisitos"
     " de competência dos cuidadores; remissão para diplomas regulamentares.",
     REF_NONE,
     S),

    # -- n.º 8 (monitorização) --
    ("ART-07-8-1", "Art.º 7.º", "n.º 8, art.º 7.º",
     "Monitorização por indicadores de comportamento baseados nos animais",
     EN_7_8, PT_7_8,
     "Sem correspondência — legislação vigente não prevê monitorização"
     " sistemática por indicadores comportamentais baseados nos animais.",
     REF_NONE,
     S,
     SEM_COD,
     REF_NONE,
     S,
     "Sem correspondência — @rgbeac não prevê sistema de monitorização"
     " por indicadores de comportamento.",
     REF_NONE,
     S),

    ("ART-07-8-2", "Art.º 7.º", "n.º 8, art.º 7.º",
     "Monitorização por indicadores de aparência física",
     EN_7_8, PT_7_8,
     SEM_DL276,
     REF_NONE,
     S,
     SEM_COD,
     REF_NONE,
     S,
     SEM_RG,
     REF_NONE,
     S),

    ("ART-07-8-3", "Art.º 7.º", "n.º 8, art.º 7.º",
     "Adotar ações com base nos resultados da monitorização",
     EN_7_8, PT_7_8,
     SEM_DL276,
     REF_NONE,
     S,
     SEM_COD,
     REF_NONE,
     S,
     SEM_RG,
     REF_NONE,
     S),

    # -- n.º 9 (atos delegados) --
    ("ART-07-9-1", "Art.º 7.º", "n.º 9, art.º 7.º",
     "Habilitação da Comissão para atos delegados (indicadores de monitorização)",
     EN_7_9, PT_7_9,
     SEM_META,
     REF_NONE,
     S,
     SEM_META,
     REF_NONE,
     S,
     SEM_META,
     REF_NONE,
     S),
]


# =====================================================================
# CRIAR WORKBOOK
# =====================================================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Art. 6-7 — Quadro de Aplicação"

# Larguras das colunas
for col_idx, (name, width) in enumerate(COLS, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.sheet_format.defaultRowHeight = 70

# =====================================================================
# ESTILOS
# =====================================================================
hdr_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=9)
art_font  = Font(name='Calibri', bold=True, color='1A237E', size=10)
cell_font = Font(name='Calibri', size=9)
cob_font  = Font(name='Calibri', bold=True, size=9)
id_font   = Font(name='Calibri', size=8, color='546E7A')

wrap_top    = Alignment(wrap_text=True, vertical='top')
wrap_center = Alignment(wrap_text=True, vertical='center', horizontal='center')
center_v    = Alignment(wrap_text=True, vertical='center', horizontal='left')


def cob_fill_style(cob_text):
    if cob_text == T:
        return fill(C_TOTAL)
    elif cob_text == P:
        return fill(C_PARCIAL)
    else:
        return fill(C_SEM)


# =====================================================================
# LINHA DE CABEÇALHO
# =====================================================================
row = 1
for col_idx, (name, _) in enumerate(COLS, 1):
    c = ws.cell(row=row, column=col_idx, value=name)
    c.fill = fill(C_HEADER)
    c.font = hdr_font
    c.alignment = wrap_center
    c.border = thin_border()
ws.row_dimensions[row].height = 40
ws.freeze_panes = 'A2'

# =====================================================================
# FUNÇÃO: SEPARADOR DE ARTIGO
# =====================================================================
def write_art_separator(ws, row_num, artigo, titulo):
    c = ws.cell(row=row_num, column=1, value=f"▶  {artigo}   {titulo}")
    c.fill = fill(C_ART)
    c.font = art_font
    c.alignment = Alignment(vertical='center', horizontal='left',
                             indent=1)
    c.border = thin_border()
    ncols = len(COLS)
    ws.merge_cells(start_row=row_num, start_column=1,
                   end_row=row_num, end_column=ncols)
    ws.row_dimensions[row_num].height = 22
    return row_num + 1


# =====================================================================
# ESCREVER DADOS
# =====================================================================
ARTICLE_TITLES = {
    "Art.º 6.º": "Artigo 6.º — Princípios gerais de bem-estar",
    "Art.º 7.º": "Artigo 7.º — Obrigações gerais em matéria de bem-estar",
}

current_row = 2
prev_art = None

for data in ROWS:
    (id_, artigo, ref_reg, det,
     en, pt,
     lv_txt, lv_ref, lv_cob,
     cod_txt, cod_ref, cod_cob,
     rg_txt,  rg_ref,  rg_cob) = data

    # Separador de artigo
    if artigo != prev_art:
        current_row = write_art_separator(
            ws, current_row, artigo, ARTICLE_TITLES.get(artigo, artigo))
        prev_art = artigo

    values = [
        id_, artigo, ref_reg, det, en, pt,
        lv_txt, lv_ref, lv_cob,
        cod_txt, cod_ref, cod_cob,
        rg_txt, rg_ref, rg_cob,
    ]

    for col_idx, val in enumerate(values, 1):
        c = ws.cell(row=current_row, column=col_idx, value=val)
        c.border = thin_border()

        # Colunas de cobertura (9, 12, 15)
        if col_idx in (9, 12, 15):
            cob = values[col_idx - 1]
            c.fill = cob_fill_style(cob)
            c.font = cob_font
            c.alignment = wrap_center
        elif col_idx == 1:   # ID
            c.font = id_font
            c.alignment = wrap_center
        elif col_idx in (2, 3):  # Artigo, Ref
            c.font = cell_font
            c.alignment = wrap_center
        elif col_idx == 4:   # Determinação
            c.font = Font(name='Calibri', size=9, bold=True)
            c.alignment = wrap_top
        elif col_idx in (5, 6):  # Texto EN / PT
            c.font = Font(name='Calibri', size=9, italic=True)
            c.alignment = wrap_top
        else:
            c.font = cell_font
            c.alignment = wrap_top

    ws.row_dimensions[current_row].height = 90
    current_row += 1


# =====================================================================
# LEGENDA
# =====================================================================
current_row += 1  # linha em branco

# Título da legenda
c = ws.cell(row=current_row, column=1, value="LEGENDA — Indicadores de Cobertura")
c.fill = fill(C_HEADER)
c.font = hdr_font
c.alignment = wrap_center
c.border = thin_border()
ws.merge_cells(start_row=current_row, start_column=1,
               end_row=current_row, end_column=4)
ws.row_dimensions[current_row].height = 20
current_row += 1

legends = [
    (T,   C_TOTAL,   "A determinação do Regulamento tem correspondência plena na legislação/proposta nacional."),
    (P,   C_PARCIAL, "Correspondência parcial — cobre o tema mas com menor detalhe, âmbito diferente, ou terminologia distinta."),
    (S,   C_SEM,     "Sem correspondência — a determinação não tem paralelo na legislação/proposta nacional. Requer criação ou adaptação de norma."),
]
for label, color, descr in legends:
    ws.cell(row=current_row, column=1, value=label).fill = fill(color)
    ws.cell(row=current_row, column=1).font = Font(name='Calibri', bold=True, size=9)
    ws.cell(row=current_row, column=1).alignment = wrap_center
    ws.cell(row=current_row, column=1).border = thin_border()
    c2 = ws.cell(row=current_row, column=2, value=descr)
    c2.font = Font(name='Calibri', size=9)
    c2.alignment = Alignment(wrap_text=True, vertical='center')
    c2.border = thin_border()
    ws.merge_cells(start_row=current_row, start_column=2,
                   end_row=current_row, end_column=4)
    ws.row_dimensions[current_row].height = 30
    current_row += 1


# =====================================================================
# CONFIGURAÇÕES FINAIS
# =====================================================================
ws.freeze_panes = 'E2'
ws.sheet_view.zoomScale = 80

# =====================================================================
# GUARDAR
# =====================================================================
output = 'quadro_aplicacao_regulamento.xlsx'
wb.save(output)
print(f"✓ Ficheiro criado: {output}")
print(f"  Linhas de dados: {len(ROWS)}")
print(f"  Artigos cobertos: 6 e 7")
