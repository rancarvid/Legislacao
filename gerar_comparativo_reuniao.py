"""
Gera ficheiro Excel + HTML comparativo para reunião artigo a artigo do @regulamento.
Exemplo com 3 artigos: Identificação, Bem-estar/Detenção, Reprodução.
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import json
import os

# ---------------------------------------------------------------------------
# DADOS — verbatim dos documentos
# ---------------------------------------------------------------------------

ARTIGOS = [
    {
        "id": "ART-05",
        "tema": "Isenções de Obrigações",
        "regulamento": {
            "ref": "Art.º 5.º do Regulamento 2023/0447",
            "titulo": "Exemptions from the obligations set out in this Chapter",
            "texto": "1.\tA breeding establishment where no more than two litters per calendar year are produced for placing on the market shall be subject to only the obligations laid down in Article 6, Article 7(1) (2), (3) and (4). Article 8, Article 9, Article 11, Article 14(2), (3) and (4), Article 15(3), (4) and (8), Article 16(1). points (b), (c) and (d), Article 17(2), (3), (5) and (7), Article 18, Article 19(1) and points 3 and 4.3 of Annex I.\n\n2.\tA shelter in which not more than a combined total of 15 dogs or cats are kept at any given time, or any foster home shall be subject to only the obligations laid down in Article 6, Article 7(1), (2), (3), (4), and(5), Article 9, Article 11, Article 14(2), (3) and (4), Article 15(3), (4) and (8), Article 16(1). points (a), (b), (c) and (d), Article 17(2), (3), (5) and (7), Article 18 and point 4.3 of Annex I.",
            "traducao": "1 — Um estabelecimento de criação que produz no máximo duas ninhadas por ano civil para colocação no mercado fica sujeito apenas às obrigações estabelecidas no artigo 5.º, artigo 6.º (n.º 1, alíneas 1b, 1c e 1d), artigo 6a, artigo 7.º, artigo 8.º, artigo 11.º (n.ºs 2, 3 e 3a), artigo 12.º (n.ºs 3, 4 e 7), artigo 13.º (n.º 2, alíneas b, c e d), artigo 14.º (n.ºs 2, 3, 4 e 5a), artigo 15.º, artigo 15a (n.º 1) e ponto 3 e 4.3 do Anexo I.\n\n2 — Um abrigo, onde são mantidos no máximo 15 cães ou gatos em qualquer momento, ou qualquer lar de acolhimento fica sujeito apenas às obrigações estabelecidas no artigo 5.º, artigo 6.º (n.ºs 1, 1a, 1b, 1c e 1d), artigo 7.º, artigo 8.º, artigo 11.º (n.ºs 2, 3 e 3a), artigo 12.º (n.ºs 3, 4 e 7), artigo 13.º (n.º 2, alíneas a, b, c e d), artigo 14.º (n.ºs 2, 3, 4 e 5a), artigo 15.º e ponto 4.3 do Anexo I."
        },
        "rgbeac": {
            "ref": "Não se aplica",
            "texto": "Não se aplica"
        },
        "codigo": {
            "ref": "Não se aplica",
            "texto": "Não se aplica"
        },
        "legislacao": {
            "ref": "Não se aplica",
            "texto": "Não se aplica"
        },
        "divergencia": {
            "legislacao": "Não se aplica",
            "codigo": "Não se aplica",
            "rgbeac": "Não se aplica",
            "sumario": "Artigo de isenções do Regulamento europeu — não tem correspondência directa em legislação portuguesa."
        },
        "necessidade_alteracao": "Não",
        "notas": "Artigo relativo a isenções no âmbito do Regulamento europeu."
    },
    {
        "id": "ART-06",
        "tema": "Princípios Gerais de Bem-Estar",
        "regulamento": {
            "ref": "Art.º 6.º do Regulamento 2023/0447",
            "titulo": "General welfare principles",
            "texto": "Operators shall apply the following general welfare principles with respect to dogs or cats bred or kept in their establishment:\n\n(a)\tdogs and cats are provided with water and feed of a quality and quantity that affords them appropriate nutrition and hydration;\n\n(b)\tdogs and cats are kept in a physical environment that is appropriate and regularly cleaned, that is secure and comfortable, especially in terms of space, air quality, temperature, light, protection against adverse climatic conditions and that is big enough to prevent overcrowding and to afford them ease of movement;\n\n(c)\tdogs and cats are kept safe, clean and in good health, and diseases, injuries, and pain due, in particular, to management, handling practices and breeding practices, are prevented;\n\n(d)\tdogs and cats are kept in an environment that enables them to exhibit species-specific and social non-harmful behaviour, and to establish a positive relationship with human beings;\n\n(e)\tdogs and cats are kept in such a way as to optimise their mental state by preventing or reducing negative stimuli in duration and intensity, as well as by maximising opportunities for positive stimuli in duration and intensity, preventing the development of abnormal repetitive or other behaviours indicative of negative animal welfare, and taking into consideration the individual animal’s needs in the domains referred to in points (a) to (d).",
            "traducao": "Os operadores devem aplicar os seguintes princípios gerais de bem-estar aos cães e gatos criados ou detidos nos seus estabelecimentos:\n\n(a) os cães e gatos são alimentados e abebeirados com água e ração de qualidade e quantidade adequadas a uma nutrição e hidratação apropriadas;\n\n(b) os cães e gatos são mantidos num ambiente físico adequado, regularmente limpo, seguro e confortável, especialmente em termos de espaço, qualidade do ar, temperatura, iluminação, proteção face a condições climáticas adversas e facilidade de movimentação, prevenindo a sobrelotação;\n\n(c) os cães e gatos são mantidos seguros, limpos e com boa saúde, prevenindo doenças, lesões e dor, nomeadamente através de práticas de maneio, manuseamento e reprodução adequadas;\n\n(d) os cães e gatos são mantidos num ambiente que lhes permite exibir comportamentos específicos da espécie e comportamentos sociais não nocivos, e estabelecer uma relação positiva com os seres humanos;\n\n(e) os cães e gatos são mantidos de forma a otimizar o seu estado mental, prevenindo ou reduzindo estímulos negativos em duração e intensidade, maximizando oportunidades de estímulos positivos, prevenindo o desenvolvimento de comportamentos repetitivos anormais, tendo em conta as necessidades individuais do animal nos domínios referidos nas alíneas (a) a (d)."
        },
        "rgbeac": {
            "ref": "al. a) do n.º 1 do art.º 10.º do RGBEAC (proposta, jun. 2025)",
            "texto": "1 — O detentor do animal de companhia deve:\n\na) Assegurar o bem-estar do animal, de acordo com sua espécie, raça, idade e necessidades físicas e etológicas, proporcionando-lhe:\n— Atenção, supervisão, controlo, exercício físico e estímulo mental;\n— Alimentos saudáveis, adequados e convenientes ao seu normal desenvolvimento e acesso permanente a água potável;\n— Condições higiossanitárias que atendam, no mínimo, ao estabelecido no presente decreto-lei e na demais legislação aplicável;\n— Liberdade de movimento, sendo proibidos todos os sistemas de contenção permanentes;\n— Abrigo adequado, em termos de tamanho e qualidade, com vista a proteger de condições atmosféricas adversas, incluindo frio, chuva, sol ou calor excessivos, com cama seca, limpa e confortável;\n— Contato social adequado a cada espécie, de acordo com a sua idade e atividade."
        },
        "codigo": {
            "ref": "n.ºs 1, 2 e 3 do art.º 5.º do Código do Animal (DL n.º 214/2013)",
            "texto": "1 — As condições de detenção e de alojamento para reprodução, criação, manutenção e acomodação dos animais de companhia devem salvaguardar os seus parâmetros de bem-estar animal.\n\n2 — Nenhum animal deve ser detido como animal de companhia se não estiverem asseguradas as condições referidas no número anterior ou se não se adaptar ao cativeiro.\n\n3 — É proibida a violência contra animais, considerando-se como tal todos os atos que, sem necessidade, infligem a morte, o sofrimento, abuso ou lesões a um animal."
        },
        "legislacao": {
            "ref": "n.º 1 do art.º 7.º e n.º 1 do art.º 9.º do DL n.º 276/2001, de 17 de outubro",
            "texto": "Artigo 7.º, n.º 1 — As condições de detenção e de alojamento para reprodução, criação, manutenção e acomodação dos animais de companhia devem salvaguardar os seus parâmetros de bem-estar animal, nomeadamente nos termos dos artigos seguintes.\n\nArtigo 9.º, n.º 1 — A temperatura, a ventilação e a luminosidade e obscuridade das instalações devem ser as adequadas à manutenção do conforto e bem-estar das espécies que albergam."
        },
        "divergencia": {
            "legislacao": "O DL n.º 276/2001 (art.ºs 7.º e 9.º) consagra os mesmos parâmetros mas desenvolve-os em artigos separados (alojamento, alimentação, ambiente) sem os denominar como domínios nem os sistematizar de forma unificada segundo o referencial OMSA.",
            "codigo": "O @codigo (art.º 5.º) refere 'parâmetros de bem-estar animal' genericamente, sem adotar formalmente a nomenclatura dos 5 domínios OMSA nem a sua sistematização explícita.",
            "rgbeac": "O @rgbeac (art.º 10.º, n.º 1, al. a)) articula obrigações equivalentes ao nível do detentor de forma descritiva, sem sistematização nos 5 domínios nem referência ao quadro OMSA.",
            "sumario": "Alinhamento substancial de conteúdo; sem necessidade de alteração imediata. Recomenda-se a adoção formal dos 5 domínios OMSA como referencial estruturante em futura revisão legislativa ou em normas de orientação técnica da DGAV."
        },
        "necessidade_alteracao": "Não",
        "notas": ""
    },
    {
        "id": "ART-07",
        "tema": "Bem-Estar e Detenção",
        "regulamento": {
            "ref": "Art.º 7.º do Regulamento 2023/0447",
            "titulo": "General welfare obligations",
            "texto": "1.\tOperators shall be responsible for the welfare of dogs and cats kept in establishments under their responsibility and control, and shall be responsible for minimising any risks to the welfare of those animals.\n\n2.\tIn the case of foster homes, the responsibility shall lie with the operator on whose behalf dogs or cats are kept. Such operators shall not place more than a combined total of five dogs or cats or one litter with or without mother in a foster home at any given time and shall provide the foster family with adequate information on the animal welfare obligations as well as the individual needs of the dogs or cats, and shall ensure that the relevant obligations laid down in this Regulation are complied with in foster homes.\n\n\tThe Member State in which the foster home is located may allow a greater number of dogs, cats or litters to be placed in the foster home, provided that the premises of the foster home have sufficient space, including outdoor space, and that the number of animal carers in the foster home is sufficient, to ensure the welfare of the dogs or cats.\n\n3.\tOperators shall not subject any dog or cat to cruelty, abuse or mistreatment, including by making them participate in activities likely to result in cruelty to or abuse or mistreatment of the dogs or cats bred or kept by the operator.\n\n4.\tOperators shall not abandon the dogs or cats bred or kept by them.\n\n5.\tBefore operators cease activities at an establishment, they shall ensure that the dogs or cats kept there are rehomed, either by taking up the pet ownership themselves or by transferring the responsibility for, or the ownership of, the dogs and cats to other operators or acquirers.\n\n6.\tOperators shall ensure that dogs and cats are handled by a number of animal carers sufficient to meet the welfare needs of dogs or cats kept in their establishments, and that those carers have the competences required under Article 12.\n\n7.\tOperators shall ensure the welfare of the dogs or cats for which they are responsible by monitoring animal-based indicators concerning behaviour and physical appearance, and by taking actions based on the results of such monitoring.\n\n8.\tThe Commission is empowered to adopt delegated acts in accordance with Article 28 supplementing this Regulation by laying down the animal-based indicators concerning behaviour and physical appearance that operators are to use for monitoring, in accordance with paragraph 7 of this Article, and the methods by which operators are to measure them.",
            "traducao": "1. Os operadores são responsáveis pelo bem-estar dos cães ou gatos mantidos nos estabelecimentos sob a sua responsabilidade e controlo e devem minimizar quaisquer riscos para o seu bem-estar.\n\n2.\tNo caso de famílias de acolhimento, a responsabilidade recai sobre o operador em nome de quem os cães ou gatos são mantidos. Esses operadores não devem colocar mais do que um total de cinco cães ou gatos ou uma ninhada com ou sem mãe numa família de acolhimento em qualquer momento e devem fornecer à família de acolhimento informação adequada sobre as obrigações de bem-estar animal, bem como as necessidades individuais dos animais, e devem assegurar que as obrigações relevantes estabelecidas por este Regulamento são cumpridas em famílias de acolhimento. Os Estados-Membros onde a família de acolhimento está localizada podem prever um número maior de cães, gatos ou ninhadas a serem colocadas na família de acolhimento, desde que as instalações da família de acolhimento providenciem espaço suficiente, incluindo espaço ao ar livre, e que o número de cuidadores de animais na família de acolhimento seja suficiente, para assegurar o bem-estar dos cães ou gatos.\n\n3.\tOs operadores não devem sujeitar nenhum cão ou gato a crueldade, abuso ou maus-tratos, incluindo através de participação em atividades que possam resultar em crueldade, abuso ou maus-tratos, aos cães ou gatos criados ou mantidos pelo operador.\n\n4.\tOs operadores não devem abandonar cães ou gatos.\n\n5.\tOs operadores que estejam prestes a cessar as atividades do seu estabelecimento devem assegurar a reintegração dos cães ou gatos mantidos aí, quer através da assunção da posse do animal de companhia, quer através da transferência da responsabilidade ou propriedade dos cães e gatos para outros operadores ou adquirentes.\n\n6.\tOs operadores devem assegurar que os cães ou gatos são tratados por um número adequado de cuidadores de animais e de modo a satisfazer as necessidades de bem-estar dos cães ou gatos mantidos nos seus estabelecimentos e que têm as competências exigidas no artigo 12.º.\n\n7.\tOs operadores devem assegurar o bem-estar dos cães ou gatos sob a sua responsabilidade através da monitorização de indicadores baseados no comportamento e aparência física, e através da adoção de ações com base nos resultados de tal monitorização.\n\n8.\tA Comissão tem competência para adotar atos delegados em conformidade com o artigo 28.º que complementem este Regulamento através do estabelecimento de indicadores baseados no bem-estar dos animais relativos ao comportamento e aparência física e dos métodos da sua medição."
        },
        "rgbeac": {
            "ref": "Art.º 10.º do RGBEAC (proposta, jun. 2025)",
            "texto": "1 — O detentor do animal de companhia deve:\n\na) Assegurar o bem-estar do animal, de acordo com a sua espécie, raça, idade e necessidades físicas e etológicas, proporcionando-lhe:\n— Atenção, supervisão, controlo, exercício físico e estímulo mental;\n— Alimentos saudáveis, adequados e convenientes ao seu normal desenvolvimento e acesso permanente a água potável;\n— Condições higiossanitárias que atendam, no mínimo, ao estabelecido no presente decreto-lei e na demais legislação aplicável;\n— Liberdade de movimento, sendo proibidos todos os sistemas de contenção permanentes."
        },
        "codigo": {
            "ref": "Art.º 5.º do Código do Animal (DL n.º 214/2013)",
            "texto": "1 — As condições de detenção e de alojamento para reprodução, criação, manutenção e acomodação dos animais de companhia devem salvaguardar os cinco domínios do bem-estar animal (Nutrição, Ambiente, Saúde, Comportamento e Psicológico), estabelecidos pela Organização Mundial da Saúde Animal (OMSA).\n\n2 — Nenhum animal deve ser detido como animal de companhia se não estiverem asseguradas as condições referidas no número anterior ou as demais previstas no presente diploma.\n\n3 — É proibida a violência contra animais, considerando-se como tal todos os atos que, sem necessidade, infligem a morte, o sofrimento, a dor, a angústia ou ferimentos a um animal."
        },
        "legislacao": {
            "ref": "n.ºs 1, 2 e 3 do art.º 7.º do DL n.º 276/2001, de 17 de outubro",
            "texto": "1 — As condições de detenção e de alojamento para reprodução, criação, manutenção e acomodação dos animais de companhia devem salvaguardar os seus parâmetros de bem-estar animal, nomeadamente nos termos dos artigos seguintes.\n\n2 — Nenhum animal deve ser detido como animal de companhia se não estiverem asseguradas as condições referidas no número anterior ou se não se adaptar ao cativeiro.\n\n3 — São proibidas todas as violências contra animais, considerando-se como tais os atos consistentes em, sem necessidade, se infligir a morte, o sofrimento ou lesões a um animal."
        },
        "divergencia": {
            "legislacao": "O DL n.º 276/2001 (art.º 7.º) centra a responsabilidade no detentor em geral, sem distinguir o contexto de acolhimento temporário nem fixar limites numéricos de animais por família. O conceito de família de acolhimento é inexistente na legislação vigente.",
            "codigo": "O @codigo não prevê o conceito de família de acolhimento nem qualquer limite numérico de animais por unidade. A responsabilidade do operador como figura distinta do detentor particular também não é contemplada.",
            "rgbeac": "O @rgbeac menciona famílias de acolhimento temporário mas não fixa o limite de 5 animais por família de acolhimento nem especifica que a responsabilidade jurídica recai sobre o operador e não sobre a família.",
            "sumario": "Lacuna transversal a toda a legislação nacional. Necessidade de: (1) criar o conceito de 'família de acolhimento' como extensão da atividade do operador; (2) fixar limite numérico de animais por família (máximo 5, ou 1 ninhada com mãe); (3) atribuir responsabilidade jurídica ao operador, não à família de acolhimento."
        },
        "necessidade_alteracao": "Sim",
        "notas": ""
    },
    {
        "id": "ART-08",
        "tema": "Estratégias de Criação — Conformação e Consanguinidade",
        "regulamento": {
            "ref": "Art.º 8.º-A do Regulamento 2023/0447",
            "titulo": "Breeding strategies obligations",
            "texto": "1.\tOperators of breeding establishments shall ensure that their breeding strategies minimise the risk of producing dogs or cats with genotypes associated with effects detrimental to the health and welfare of those animals.\n\n2.\tOperators of breeding establishments shall not use for reproduction dogs or cats that have excessive conformational traits leading to a high risk of detrimental effects on the welfare of such dogs or cats, or of their offspring. Before selecting a dog or cat that might have an excessive conformational trait for breeding, the operator shall consult a veterinarian or an independent qualified person acting under the responsibility of a veterinarian. That veterinarian or independent qualified person shall assess whether the dog or cat has an excessive conformational trait.\n\n3.\tThe Commission is empowered to adopt delegated acts in accordance with Article 28 supplementing this Regulation by:\n\n(a)\tdefining characteristics of the genotypes referred to in paragraph 1 that are to be excluded from reproduction, and the methods for their assessment and the record keeping requirements;\n\n(b)\tdefining excessive conformational traits referred to in paragraph 2 of this Article that are to be excluded from reproduction, the methods for their assessment and the record keeping requirements.\n\n\tWhen adopting those delegated acts, the Commission shall take into account the scientific opinion of the European Food Safety Authority (EFSA), as well as any social and economic impacts of those delegated acts.\n\n\tThe delegated acts concerning the excessive conformational traits shall be adopted by 1 July 2030. The delegated acts concerning the genotypes shall be adopted by 1 July 2036.\n\n4.\tThe following shall be prohibited when managing the reproduction of dogs and cats:\n\n(a)\tbreeding between parents and offspring, between siblings, between half-siblings or between grandparents and grandchildren, unless approved by the competent authority based on a specific need to preserve local breeds with a limited genetic pool;\n\n(b)\tbreeding for the purpose of producing hybrids.",
            "traducao": "1.\tOs operadores de estabelecimentos de criação devem assegurar que as suas estratégias de criação minimizam o risco de produzir cães ou gatos com genótipos associados a efeitos prejudiciais para a sua saúde e bem-estar.\n\n2.\tOs operadores de estabelecimentos de criação não devem utilizar para reprodução cães ou gatos que apresentem traços conformacionais excessivos que conduzam a um risco elevado de efeitos prejudiciais para o bem-estar desses animais ou das suas crias. Antes da seleção para reprodução de um animal potencialmente afetado por traço conformacional excessivo, o operador deve consultar um médico veterinário ou uma pessoa qualificada independente sob responsabilidade veterinária. O médico veterinário ou a pessoa qualificada independente devem avaliar se o animal tem um traço conformacional excessivo.\n\n3.\tA Comissão tem competência para adotar atos delegados em conformidade com o artigo 28.º que complementem este Regulamento, definindo:\n\n(a)\tcaracterísticas dos genótipos a que se refere o n.º 1 que devem ser excluídos da reprodução, e os métodos para a sua avaliação e requisitos de manutenção de registos;\n\n(b)\ttraços conformacionais excessivos a que se refere o n.º 2 do presente artigo que devem ser excluídos da reprodução, os métodos para a sua avaliação e requisitos de manutenção de registos.\n\n\tAo adotar esses atos delegados, a Comissão tem em conta os pareceres científicos da Autoridade Europeia para a Segurança dos Alimentos (EFSA), bem como os impactos sociais e económicos desses atos delegados.\n\n\tOs atos delegados relativos aos traços conformacionais excessivos devem ser adotados até 1 de julho de 2030. Os atos delegados relativos aos genótipos devem ser adotados até 1 de julho de 2036.\n\n4.\tSão proibidos na gestão da reprodução de cães e gatos:\n\n(a)\to cruzamento entre progenitores e descendentes, entre irmãos, entre meios-irmãos ou entre avós e netos, exceto com aprovação da autoridade competente por razão de necessidade específica de preservação de raças locais com reserva genética limitada;\n\n(b)\to cruzamento para produção de híbridos."
        },
        "rgbeac": {
            "ref": "n.ºs 1 e 2 do art.º 34.º do RGBEAC (proposta, jun. 2025)",
            "texto": "1 — As intervenções cirúrgicas com o objetivo da modificação da aparência ou com fins não curativos ou não destinados a impedir a reprodução dos animais são proibidas, nos termos do disposto na alínea i) do n.º 1 do artigo 12.º.\n\n2 — O detentor de animal sujeito a intervenção curativa que modifique a sua aparência deve possuir documento comprovativo da necessidade da mesma, passado pelo médico veterinário que a ela procedeu, sob a forma de atestado do qual constem a identificação do médico veterinário, o número da cédula profissional e a sua assinatura, ou, no caso de animais importados, documento comprovativo da necessidade dessa intervenção, emitida pelo médico veterinário que a ela procedeu, legalizado pela autoridade competente do respetivo país."
        },
        "codigo": {
            "ref": "n.º 2, als. a), b) e c) do art.º 8.º do Código do Animal (DL n.º 214/2013)",
            "texto": "2 — A reprodução de animais obedece ao seguinte:\n\na) Os animais só devem ser utilizados na reprodução depois de atingida a maturidade reprodutiva para a espécie e raça devendo, no caso dos cães e gatos, seguir os parâmetros referidos no anexo I ao presente diploma, do qual faz parte integrante, não sendo autorizado, no caso das fêmeas, o acasalamento em cios sucessivos;\n\nb) Deve ser respeitada a regra do porte semelhante dos progenitores, para prevenir a possibilidade de distócia;\n\nc) Devem ser excluídos da reprodução, os animais que revelem defeitos genéticos e malformações, designadamente monorquidia e displasia da anca nos cães e rim poliquístico nos gatos, ou alterações comportamentais."
        },
        "legislacao": {
            "ref": "art.º 17.º e n.º 1 do art.º 18.º do DL n.º 276/2001, de 17 de outubro",
            "texto": "Artigo 17.º — As intervenções cirúrgicas, nomeadamente as destinadas ao corte de caudas nos canídeos, têm de ser executadas por um médico veterinário.\n\nArtigo 18.º, n.º 1 — Os detentores de animais de companhia que os apresentem com quaisquer amputações que modifiquem a aparência dos animais ou com fins não curativos devem possuir documento comprovativo, passado pelo médico veterinário que a elas procedeu, da necessidade dessa amputação, nomeadamente discriminando que as mesmas foram feitas por razões médico-veterinárias ou no interesse particular do animal ou para impedir a reprodução."
        },
        "divergencia": {
            "legislacao": "O DL n.º 276/2001 (art.ºs 17.º e 18.º) exige intervenção veterinária em cirurgias e documentação para amputações, mas não prevê qualquer restrição à reprodução por traços conformacionais excessivos nem proibição de consanguinidade próxima ou hibridação.",
            "codigo": "O @codigo (art.º 8.º, n.º 2) exclui da reprodução animais com defeitos genéticos e malformações (displasia, rim poliquístico) e proíbe cios sucessivos, mas não prevê o conceito de traços conformacionais excessivos nem a proibição de consanguinidade.",
            "rgbeac": "O @rgbeac (art.º 34.º) proíbe intervenções cirúrgicas de modificação da aparência, mas não contém qualquer norma sobre estratégia genética de criação, conformação excessiva ou consanguinidade.",
            "sumario": "Lacuna normativa transversal. Necessidade de criar norma específica que: (1) defina traços conformacionais excessivos (a regular por ato delegado europeu até 2030); (2) proíba consanguinidade próxima (pais/filhos, irmãos, avós/netos); (3) proíba a produção de híbridos; (4) imponha consulta veterinária prévia ao acasalamento de animais potencialmente afetados."
        },
        "necessidade_alteracao": "Sim",
        "notas": ""
    },
    {
        "id": "ART-09",
        "tema": "Reprodução e Criação",
        "regulamento": {
            "ref": "Art.º 9.º do Regulamento 2023/0447",
            "titulo": "Notification and registration of establishments",
            "texto": "1.\tOperators shall notify the competent authorities of their activity, providing at least the following information for each of their establishments:\n\n(a)\tthe name, address and contact details of the operator;\n\n(b)\tthe location of the establishment;\n\n(c)\tthe type of establishment: breeding establishment, selling establishment, shelter or foster home;\n\n(d)\tthe species and, for breeding establishments, the breeds of the dogs or cats kept in the establishment;\n\n(e)\tthe capacity of the establishment, expressed as the maximum number of dogs and cats which can be kept in the establishment;\n\n(f)\tfor breeding establishments, the estimated number of litters to be placed on the market per year.\n\n2.\tOperators shall notify the competent authority of:\n\n(a)\tany changes concerning the information referred to in paragraph 1;\n\n(b)\twhere applicable, the planned date of a cessation of their activities, at the latest five working days before that date.\n\n3.\tMember States shall use the information provided in accordance with Article 84 of Regulation (EU) 2016/429. Operators shall not be required to notify the information already submitted in accordance with Article 84 of Regulation (EU) 2016/429 again.\n\n4.\tThe competent authority shall maintain a register of establishments. The competent authority may use for that purpose the register provided for in Article 101(1), point (a), of Regulation (EU) 2016/429.",
            "traducao": "1.\tOs operadores devem notificar as autoridades competentes da sua atividade, fornecendo pelo menos as seguintes informações para cada um dos seus estabelecimentos:\n\n(a)\to nome, morada e contactos do operador;\n\n(b)\ta localização do estabelecimento;\n\n(c)\to tipo de estabelecimento: estabelecimento de criação, estabelecimento de venda, abrigo ou família de acolhimento;\n\n(d)\tas espécies e, para os estabelecimentos de criação, as raças dos cães ou gatos mantidos no estabelecimento;\n\n(e)\ta capacidade do estabelecimento, expressa no número máximo de cães e gatos que podem ser alojados no estabelecimento;\n\n(f)\tpara os estabelecimentos de criação, o número estimado de ninhadas a colocar no mercado por ano.\n\n2.\tOs operadores devem notificar a autoridade competente de:\n\n(a)\tquaisquer alterações relativas às informações referidas no n.º 1;\n\n(b)\tquando aplicável, a data prevista de cessação das suas atividades, o mais tardar cinco dias úteis antes dessa data.\n\n3.\tOs Estados-Membros devem utilizar as informações fornecidas em conformidade com o artigo 84.º do Regulamento (UE) 2016/429. Os operadores não são obrigados a notificar novamente as informações já apresentadas em conformidade com o artigo 84.º do Regulamento (UE) 2016/429.\n\n4.\tA autoridade competente deve manter um registo dos estabelecimentos. A autoridade competente pode utilizar para esse efeito o registo previsto no artigo 101.º, n.º 1, alínea a), do Regulamento (UE) 2016/429."
        },
        "rgbeac": {
            "ref": "Art.º 43.º do RGBEAC (proposta, jun. 2025)",
            "texto": "1 — Os centros de bem-estar animal procedem à esterilização dos animais recolhidos que se presuma terem sido abandonados, nos termos do artigo 41.º.\n\n2 — As câmaras municipais devem, sempre que necessário e sob a responsabilidade do médico veterinário municipal, incentivar e promover programas de esterilização de animais de companhia, nomeadamente os cães e gatos, em complementaridade com os centros de bem-estar animal.\n\n3 — Os requisitos mínimos das instalações adequadas à realização de esterilizações nos centros de bem-estar animal são estabelecidos em portaria do membro do Governo responsável pela área da agricultura."
        },
        "codigo": {
            "ref": "Art.º 8.º do Código do Animal (DL n.º 214/2013)",
            "texto": "1 — A reprodução de animais deve ser realizada de forma planeada.\n\n2 — A reprodução de animais obedece ao seguinte:\n\na) Os animais só devem ser utilizados na reprodução depois de atingida a maturidade reprodutiva para a espécie e raça devendo, no caso de cães e gatos, as fêmeas ter pelo menos dois anos de idade no momento da primeira cobertura;\n\nb) Deve ser respeitada a regra do porte semelhante dos progenitores, para prevenir a possibilidade de distócia;\n\nc) Devem ser excluídos da reprodução os animais que revelem defeitos genéticos e malformações, designadamente monorquidia e displasia."
        },
        "legislacao": {
            "ref": "n.º 1 do art.º 3.º-A do DL n.º 276/2001, de 17 de outubro",
            "texto": "1 — A mera comunicação prévia a que se refere a alínea a) do n.º 1 do artigo anterior é dirigida à DGAV e deve conter os seguintes elementos, quando aplicáveis:\n\na) O nome ou a denominação social do interessado;\nb) A localização do alojamento e a sua designação comercial;\nc) O número de identificação fiscal ou de pessoa coletiva do interessado;\nd) Municípios integrantes, no caso dos centros de recolha intermunicipais;\ne) Caracterização das atividades a exercer;\nf) Indicação do médico veterinário responsável pelo alojamento;\ng) O número de celas de quarentena para isolamento de animais por suspeita de raiva, no caso dos centros de recolha;\nh) A capacidade máxima de animais e respetivas espécies a alojar;\ni) O número de animais detidos, espécies e raças;\nj) Declaração de responsabilidade, subscrita pelo interessado, relativa ao cumprimento da legislação aplicável aos animais de companhia, nomeadamente em matéria de instalações, equipamentos, higiene, saúde e bem-estar dos animais."
        },
        "divergencia": {
            "legislacao": "O DL n.º 276/2001 (art.º 3.º-A) prevê comunicação prévia à DGAV com elementos parcialmente equivalentes (capacidade máxima, espécies, raças). Não exige a estimativa anual de ninhadas a colocar no mercado (al. ea) do @regulamento), constituindo lacuna parcial. É o diploma mais próximo do @regulamento neste eixo.",
            "codigo": "O @codigo regula condições de reprodução (art.º 8.º) mas não prevê qualquer sistema de notificação ou registo de estabelecimentos criadores junto da autoridade competente.",
            "rgbeac": "O @rgbeac trata da esterilização de errantes e CED, mas não contém norma sobre notificação de criadores ou registo de estabelecimentos com fins de reprodução comercial.",
            "sumario": "A @legislacao vigente é o diploma base adequado para a transposição. Necessidade de ajuste: acrescentar a obrigação de estimativa anual de ninhadas a colocar no mercado e alinhar os restantes elementos informativos com a lista exaustiva do art.º 7.º do @regulamento."
        },
        "necessidade_alteracao": "Sim",
        "notas": ""
    },
    {
        "id": "ART-10",
        "tema": "Aprovação de Estabelecimentos de Criação",
        "regulamento": {
            "ref": "Art.º 10.º do Regulamento 2023/0447",
            "titulo": "Approval of breeding establishments",
            "texto": "1.\tOperators of breeding establishments that either produce or intend to produce more than five litters per calendar year or that keep more than a combined total of five bitches or queens at any given time shall place dogs or cats on the market only after their establishment has been approved by the competent authority.\n\n2.\tThe competent authority shall perform on-site inspections to verify that the establishment meets the requirements of this Regulation. Member States may allow such inspections to be carried out remotely, provided that the means of distance communication used provides sufficient evidence for the competent authority to perform reliable inspections. The competent authority shall grant certificates of approval only to breeding establishments that meet the requirements of this Regulation.\n\n3.\tThe competent authority shall maintain a publicly available list including the following information for each approved establishment:\n\n(a)\tthe name, contact details and, where available, the URL of the website of the establishment;\n\n(b)\tthe address of the establishment;\n\n(c)\tthe name of the operator;\n\n(d)\tthe species and, if relevant, the breeds related to the establishment activities approved;\n\n(e)\tthe unique approval number assigned to the establishment by the competent authority and the date of the approval and cessation of activities.",
            "traducao": "1.\tOs operadores de estabelecimentos de criação que produzem ou pretendem produzir mais de cinco ninhadas por ano civil ou que mantêm mais do que um total combinado de cinco cadelas ou gatas em qualquer momento podem colocar cães ou gatos no mercado apenas após a aprovação do seu estabelecimento pela autoridade competente.\n\n2.\tA autoridade competente deve realizar inspeções no local para verificar que o estabelecimento atende aos requisitos do presente Regulamento. Os Estados-Membros podem permitir que tais inspeções sejam realizadas remotamente, desde que o meio de comunicação à distância utilizado forneça provas suficientes para a autoridade competente realizar inspeções fiáveis. A autoridade competente deve conceder certificados de aprovação apenas aos estabelecimentos de criação que cumprem os requisitos do presente Regulamento.\n\n3.\tA autoridade competente deve manter uma lista disponível ao público incluindo as seguintes informações para cada estabelecimento aprovado:\n\n(a)\to nome, dados de contacto e, se disponível, o URL do site do estabelecimento;\n\n(b)\ta morada do estabelecimento;\n\n(c)\to nome do operador;\n\n(d)\ta(s) espécie(s) e, se relevante, a(s) raça(s) relacionada(s) com as atividades do estabelecimento aprovadas;\n\n(e)\to número de aprovação único atribuído ao estabelecimento pela autoridade competente e a data da aprovação e cessação de atividades."
        },
        "rgbeac": {
            "ref": "Não se aplica",
            "texto": "Não se aplica"
        },
        "codigo": {
            "ref": "Não se aplica",
            "texto": "Não se aplica"
        },
        "legislacao": {
            "ref": "Não se aplica",
            "texto": "Não se aplica"
        },
        "divergencia": {
            "legislacao": "Conceito novo de ‘aprovação’ com limiar (>5 ninhadas/ano ou >5 cadelas/gatas)",
            "codigo": "Não implementado",
            "rgbeac": "Não implementado",
            "sumario": "Artigo completamente novo no Regulamento. Introduz aprovação formal de criadores com limiar numérico."
        },
        "necessidade_alteracao": "Sim",
        "notas": "Análise e correspondência nacional pendente de complementação."
    },
    {
        "id": "ART-11",
        "tema": "Detenção Responsável e Informação",
        "regulamento": {
            "ref": "Art.º 11.º do Regulamento 2023/0447",
            "titulo": "Obligation of informing on responsible ownership",
            "texto": "1.\tOperators shall provide to the acquirer of a dog or cat written information necessary to enable the acquirer to ensure the animal’s welfare, including information on responsible ownership and on the specific needs of the animal in terms of feeding, care, health and housing, as well as information on its behavioural needs and health history.\n\n2.\tThe written information on the dog or cat’s health history referred to in the first paragraph shall include at least:\n\n(a)\tthe animal’s vaccination status;\n\n(b)\tany medical conditions or predispositions to diseases, including allergies, that are known to the operator, and any diagnostic test results for the dog or cat that are available to the operator.\n\nWhere the information on the animal’s health history is set out in a document required under Regulation (EU) 2016/429, the operator shall transmit that document to the acquirer.",
            "traducao": "1.\tOs operadores devem fornecer ao adquirente de um cão ou gato informação escrita necessária para lhe permitir assegurar o bem-estar do cão ou gato, incluindo informação sobre detenção responsável e sobre as necessidades específicas do cão ou gato em termos de alimentação, cuidados, saúde e alojamento, bem como informação sobre as suas necessidades comportamentais e historial de saúde.\n\n2.\tA informação escrita sobre o historial de saúde do cão ou gato referida no n.º 1 deve incluir pelo menos:\n\n(a)\to estado de vacinação do cão ou gato;\n\n(b)\tquaisquer condições médicas ou predisposições a doenças, incluindo alergias, conhecidas pelo operador, e quaisquer resultados de testes de diagnóstico para o cão ou gato que estejam disponíveis para o operador.\n\nCaso a informação sobre o historial de saúde do cão ou gato esteja prevista num documento exigido nos termos do Regulamento (UE) 2016/429, o operador deve transmitir esse documento ao adquirente."
        },
        "rgbeac": {
            "ref": "Art.º 8.º do RGBEAC (proposta, jun. 2025)",
            "texto": "1 — O detentor do animal de companhia deve possuir informação sobre as obrigações inerentes à sua detenção, incluindo direitos e responsabilidades, normas de bem-estar, cuidados de saúde, comportamento, necessidades específicas da espécie/raça, duração de vida esperada, custos de manutenção, e proibição de abandono.\n\n2 — Os comerciantes devem fornecer por escrito ao adquirente informação completa antes da transferência do animal, incluindo identidade e dados de contacto do comerciante, dados de identificação do animal, cuidados de saúde e estado de vacinação, e certificado de origem ou de compatibilidade quando aplicável."
        },
        "codigo": {
            "ref": "Art.º 57.º do Código do Animal (DL n.º 214/2013)",
            "texto": "1 — O detentor do animal deve ter acesso a informação sobre:\n\na) As obrigações inerentes à sua detenção, direitos e responsabilidades;\nb) As normas de bem-estar, cuidados de saúde e comportamento esperado;\nc) As necessidades específicas da espécie ou raça;\nd) A duração de vida esperada;\ne) Os custos de manutenção;\nf) A proibição de abandono.\n\n2 — Os criadores e comerciantes devem fornecer informação escrita completa ao adquirente antes da transferência do animal, incluindo dados de identificação e cuidados de saúde."
        },
        "legislacao": {
            "ref": "Art.º 20.º do DL n.º 276/2001, de 17 de outubro",
            "texto": "1 — O proprietário ou detentor de animal de companhia deve ter acesso a informação adequada sobre as suas obrigações relativas ao bem-estar do animal.\n\n2 — Os detentores de animais de companhia que os comercializem devem fornecer informação ao adquirente sobre o estado de saúde do animal e vacinas efetuadas."
        },
        "divergencia": {
            "legislacao": "O DL n.º 276/2001 (art.º 20.º) é genérico e não especifica a forma escrita nem detalhes de conteúdo. O @regulamento exige informação escrita sobre necessidades específicas de bem-estar (alimentação, cuidados, saúde, alojamento, comportamento) e condiciona a transferência à transmissão dessa informação.",
            "codigo": "O @codigo (art.º 57.º) exige informação escrita mas com âmbito mais amplo (duração de vida, custos) do que o @regulamento, que se foca especificamente em bem-estar e saúde.",
            "rgbeac": "O @rgbeac (art.º 8.º) aproxima-se do conteúdo do @regulamento, exigindo informação escrita antes da transferência, incluindo dados de saúde e vacinação, mas ainda sem o enfoque específico em bem-estar comportamental.",
            "sumario": "Necessidade de alteração: (1) formalizar requisito de informação escrita como condição de transferência de animal; (2) especificar conteúdo obrigatório sobre bem-estar, saúde e necessidades comportamentais; (3) harmonizar entre @codigo e @rgbeac quanto a âmbito e forma."
        },
        "necessidade_alteracao": "Sim",
        "notas": ""
    },
    {
        "id": "ART-12",
        "tema": "Competências de Cuidadores e Bem-Estar Animal",
        "regulamento": {
            "ref": "Art.º 12.º do Regulamento 2023/0447",
            "titulo": "Animal welfare competences of animal carers",
            "texto": "1.	Animal carers, other than volunteers in shelters and interns who are acting under the responsibility of a competent animal carer, shall have the following competences as regards the dogs and cats they are handling:\n\n(a)	an understanding of the animals’ biological behaviour and their physiological and ethological needs;\n\n(b)	the ability to recognise the animals’ expressions, including any sign of suffering, and to identify and to take the appropriate mitigating measures  in such cases;\n\n(c)	the ability to apply good animal management practices, including operant conditioning and positive reinforcement, to use and maintain the equipment used for the dogs or cats under their care and to minimise any risks to the welfare of those dogs or cats, preventing them from suffering;\n\n(d)	knowledge of the carers’ obligations under this Regulation.\n\n2.	The competences referred to in paragraph 1 may be acquired through education, training or professional experience. Only documented education, training or professional experience shall be taken into account when determining whether an animal carer has the competences referred to in paragraph 1.\n\n3.	Operators shall ensure that at least one animal carer, other than a volunteer or intern, at the establishment has completed the training courses referred to in Article 22. Operators shall ensure that that animal carer transfers his or her knowledge to the other animal carers of the establishment.\n\n4.	The Commission shall adopt implementing acts, laying down minimum requirements concerning the formal education, training or professional experience referred to in paragraph 2 of this Article necessary to determine whether an animal carer has the competences referred to in paragraph 1 and for the training courses referred to in paragraph 3.\n\nThe implementing act concerning the training courses referred to in paragraph 3 shall be adopted by ... [3 years from the date of entry into force of this Regulation].\n\nThose implementing acts shall be adopted in accordance with the examination procedure referred to in Article 29.",
            "traducao": "1. Os cuidadores de animais, com exceção de voluntários em abrigos e estagiários sob a responsabilidade de um cuidador competente, devem ter as seguintes competências no que diz respeito aos cães e gatos que manuseiam:\n(a) compreensão do seu comportamento biológico e das suas necessidades fisiológicas e etológicas;\n(b) capacidade de reconhecer as suas expressões, incluindo qualquer sinal de sofrimento, e de identificar e adotar as medidas mitigantes apropriadas nesses casos;\n(c) capacidade de aplicar boas práticas de maneio de animais, incluindo condicionamento operante e reforço positivo, utilizar e manter o equipamento utilizado para os cães ou gatos sob os seus cuidados e minimizar quaisquer riscos para o bem-estar dos cães ou gatos, prevenindo sofrimento;\n(d) conhecimento das suas obrigações sob este Regulamento.\n\n2. As competências referidas no parágrafo 1 podem ser adquiridas através de educação, formação ou experiência profissional. A educação, formação ou experiência profissional devem ser documentadas.\n\n3.\tOs operadores devem assegurar que pelo menos um cuidador de animais, que não seja um voluntário ou estagiário, no estabelecimento tenha completado os cursos de formação referidos no artigo 22.º e que o cuidador transfira os conhecimentos aos outros cuidadores de animais do estabelecimento.\n\n4.\tA Comissão estabelecerá, por meio de atos de execução, os requisitos mínimos relativos à educação formal, formação ou experiência profissional para adquirir as competências referidas no parágrafo 2 e para os cursos de formação referidos no n.º 3. Esses atos de execução serão adotados de acordo com o procedimento de exame referido no artigo 29.º. O ato de execução relativo aos cursos de formação referidos no n.º 3 deve ser adotado por [3 anos da data de entrada em vigor do Regulamento]."
        },
        "rgbeac": {
            "ref": "Art.ºs 69.º, 84.º e 90.º do RGBEAC (proposta, jun. 2025)",
            "texto": "Artigo 69.º — Formação\n\nA reprodução, criação, manutenção, venda ou treino de animais de companhia depende de aprovação em formação sobre detenção responsável de animais de companhia, bem como sobre as necessidades fisiológicas e etológicas específicas da espécie animal em causa, ministrada pelo ICNF, I.P. ou entidades por este certificadas.\n\nArtigo 84.º — Pessoal\n\nO pessoal responsável pelas tarefas referidas no artigo 82.º deve possuir os conhecimentos e a experiência adequados para as executar.\n\nArtigo 90.º — Pessoal\n\nO pessoal auxiliar deve possuir os conhecimentos e a experiência adequada, o qual fica, contudo, sob a orientação do médico veterinário responsável."
        },
        "codigo": {
            "ref": "Art.º 19.º do Código do Animal (DL n.º 214/2013)",
            "texto": "Artigo 19.º — Pessoal auxiliar\n\nOs alojamentos devem dispor de pessoal auxiliar que possua os conhecimentos e a aptidão necessária para assegurar os cuidados adequados aos animais, o qual fica sob a orientação do médico veterinário responsável."
        },
        "legislacao": {
            "ref": "Art.º 13.º do DL n.º 276/2001, de 17 de outubro",
            "texto": "Artigo 13.º — Maneio\n\n1 — A observação diária dos animais e o seu maneio, a organização da dieta e o tratamento médico-veterinário devem ser assegurados por pessoal técnico competente e em número adequado à quantidade e espécies animais que alojam.\n\n2 — O maneio deve ser feito por pessoal que possua formação teórica e prática específica ou sob a supervisão de uma pessoa competente para o efeito.\n\n3 — Todos os animais devem ser alvo de inspeção diária, sendo de imediato prestados os primeiros cuidados."
        },
        "divergencia": {
            "legislacao": "O DL n.º 276/2001 (art.º 13.º) exige pessoal 'técnico competente' com 'formação teórica e prática específica' e 'inspeção diária', mas não detalha competências concretas. O @regulamento especifica 4 competências estruturadas (comportamento biológico, reconhecimento de sofrimento, maneio e bem-estar, conhecimento de obrigações) e exige documentação de educação/formação/experiência.",
            "codigo": "O @codigo (art.º 19.º) é genérico: exige apenas 'conhecimentos e aptidão necessária' sob orientação do médico veterinário responsável. Não detalha competências concretas nem exigências de formação formal.",
            "rgbeac": "O @rgbeac (art.ºs 69.º, 84.º, 90.º) aproxima-se mais do @regulamento: exige 'conhecimentos e experiência adequados', menciona 'necessidades fisiológicas e etológicas', exige formação certificada pelo ICNF. Lacuna: não detalha as 4 competências específicas do @regulamento (reconhecimento de sofrimento, condicionamento operante, etc.).",
            "sumario": "Alinhamento parcial do @rgbeac com @regulamento. Para @codigo e @legislacao, necessidade de: (1) detalhar as 4 competências específicas (comportamento, reconhecimento de sofrimento, maneio positivo, conhecimento de obrigações); (2) exigir documentação formal de educação/formação/experiência; (3) requerer que operador designe formador responsável que transfira conhecimento; (4) implementar requisitos mínimos de formação via regulamento delegado."
        },
        "necessidade_alteracao": "Sim",
        "notas": ""
    },
    {
        "id": "ART-13",
        "tema": "Avaliação e Supervisão de Bem-Estar",
        "regulamento": {
            "ref": "Art.º 13.º do Regulamento 2023/0447",
            "titulo": "Advisory welfare visits",
            "texto": "1.\tOperators shall:\n\n(a)\tensure that the establishments for which they are responsible receive a visit by a veterinarian for the purpose of identifying and assessing any risk factor for the welfare of the dogs or cats and advising the operator on measures to address those risks initially by ... [date three years after the date of entry into force of this Regulation] or one year following the notification of the new establishment, and thereafter when appropriate, based on a risk analysis by the competent authorities, or on an annual basis if Member States so provide in their national law;\n\n(b)\tkeep the records of the findings of the veterinarian’s visit referred to in point (a) and of their follow up actions for at least four years, from the day of the visit, and shall make those records available to the competent authorities upon request as well as to the veterinarians that perform subsequent advisory visits.\n\n2.\tBy ... [date 24 months from the date of entry into force of this Regulation], the Commission shall adopt delegated acts in accordance with Article 28 supplementing this Article by laying down the minimum criteria to be assessed by the veterinarian during the advisory welfare visit.",
            "traducao": "Os operadores devem:\n\n(a) assegurar que os estabelecimentos sob a sua responsabilidade recebem uma visita de um médico veterinário no prazo de um ano a partir da data de aplicação do presente Regulamento ou no prazo de um ano após notificação de novo estabelecimento, com o objetivo de identificar e avaliar quaisquer fatores de risco para o bem-estar dos cães ou gatos e aconselhar o operador sobre medidas para resolver esses riscos; depois, as visitas do médico veterinário devem ocorrer quando apropriado, com base numa análise de risco pelas autoridades competentes; os Estados-Membros podem estabelecer que as visitas de aconselhamento de bem-estar sejam anuais;\n\n(b) manter registos dos resultados da visita do médico veterinário referida na alínea (a) e das ações de acompanhamento por um período mínimo de 4 anos, a partir da data da visita, e disponibilizá-los às autoridades competentes a pedido e ao médico veterinário que realiza visitas de aconselhamento subsequentes.\n\nDentro de 24 meses a partir da data de entrada em vigor do presente Regulamento, a Comissão deve adotar atos delegados em conformidade com o artigo 28.º que complementem o presente artigo de forma a estabelecer critérios mínimos a serem avaliados durante a visita de aconselhamento de bem-estar."
        },
        "rgbeac": {
            "ref": "Art.º 56.º do RGBEAC (proposta, jun. 2025)",
            "texto": "Artigo 56.º — Médico veterinário responsável pelo alojamento\n\n1 — Os titulares da exploração de alojamentos para hospedagem, com exceção dos alojamentos para hospedagem com fins higiénicos, devem ter ao seu serviço um médico veterinário responsável pelo alojamento.\n\n2 — Ao médico veterinário responsável pelo alojamento compete:\na) A elaboração de parecer relativo à verificação das condições higiossanitárias e de bem-estar animal exigidas no presente decreto-lei;\nb) A elaboração e a execução de programas e ações que visem a saúde e o bem-estar dos animais e o seu acompanhamento, bem como a emissão de pareceres relativos à saúde e ao bem-estar dos animais;\nc) A orientação técnica do pessoal responsável pela observação, maneio e prestação de cuidados aos animais;\nd) A colaboração com as autoridades competentes em todas as ações que estas determinem."
        },
        "codigo": {
            "ref": "Art.º 32.º do Código do Animal (DL n.º 214/2013)",
            "texto": "Artigo 32.º — Médico veterinário responsável pelo alojamento\n\n1 — Os titulares da exploração de alojamentos para hospedagem sem fins lucrativos e com fins lucrativos de animais, com exceção dos com fins higiénicos, necessitam de ter ao seu serviço um médico veterinário que seja responsável pelo alojamento.\n\n2 — Ao médico veterinário responsável pelo alojamento compete:\na) A elaboração e execução de programas que visem a saúde e o bem-estar dos animais e o seu acompanhamento, bem como a emissão de pareceres relativos à saúde e ao bem-estar dos animais;\nb) A orientação técnica do pessoal que cuida dos animais;\nc) A colaboração com as autoridades competentes em todas as ações que estas determinarem."
        },
        "legislacao": {
            "ref": "Art.º 4.º do DL n.º 276/2001, de 17 de outubro",
            "texto": "Artigo 4.º — Médico veterinário responsável pelo alojamento\n\n1 — Os titulares da exploração de alojamentos para hospedagem sem fins lucrativos e com fins lucrativos de animais, com exceção dos alojamentos para hospedagem com fins higiénicos, devem ter ao seu serviço um médico veterinário que seja responsável pelo alojamento.\n\n2 — Ao médico veterinário responsável pelo alojamento compete:\na) A elaboração e a execução de programas e ações que visem a saúde e o bem-estar dos animais e o seu acompanhamento, bem como a emissão de pareceres relativos à saúde e ao bem-estar dos animais;\nb) A orientação técnica do pessoal que cuida dos animais;\nc) A colaboração com as autoridades competentes em todas as ações que estas determinarem."
        },
        "divergencia": {
            "legislacao": "O DL n.º 276/2001 (art.º 4.º) estabelece a obrigação de ter médico veterinário responsável que execute 'programas e ações' para saúde e bem-estar, mas não especifica que essa avaliação deve ocorrer em prazo determinado (ex.: 1 ano) ou que os registos devem ser mantidos por período específico (4 anos). O @regulamento é mais prescritivo neste aspecto.",
            "codigo": "O @codigo (art.º 32.º) replica quase verbatim o art.º 4.º do DL n.º 276/2001, mantendo as mesmas lacunas: não fixa prazos para avaliações de bem-estar nem obrigações de manutenção de registos estruturados.",
            "rgbeac": "O @rgbeac (art.º 56.º) reforça a obrigação com 'parecer relativo à verificação das condições higiossanitárias e de bem-estar', mas igualmente sem prazos específicos para avaliações ou duração de retenção de registos. Ambos focam-se em 'programas' interno, não em 'visitas periódicas externas'.",
            "sumario": "Lacuna estrutural em toda a legislação nacional: enquanto o @regulamento exige 'visita de um veterinário' num prazo específico (1 ano) com manutenção de registos por 4 anos e transmissão entre veterinários, a legislação nacional concentra-se em ter um 'médico veterinário responsável' no estabelecimento. Necessidade de alteração: (1) formalizar obrigação de 'visita de avaliação' por veterinário dentro de 1 ano após notificação; (2) exigir avaliações periódicas conforme risco; (3) obrigatória manutenção de registos por 4 anos; (4) garantir transmissão de informações ao próximo veterinário avaliador; (5) estabelecer critérios mínimos de avaliação (bem-estar comportamental, alojamento, saúde, etc.)."
        },
        "necessidade_alteracao": "Sim",
        "notas": ""
    },
    {
        "id": "ART-14",
        "tema": "Alimentação e Hidratação",
        "regulamento": {
            "ref": "Art.º 14.º do Regulamento 2023/0447",
            "titulo": "Feeding and watering",
            "texto": "1.	Operators  shall ensure that dogs and cats are fed in accordance with the requirements laid down in point 1 of Annex I  .\n\n2.	In addition, operators  shall ensure that dogs and cats are adequately fed and hydrated by supplying:\n\n(a)	clean and fresh water, ad libitum;\n\n(b)	feed of sufficient quantity and quality to meet the physiological, nutritional and metabolic needs  of the dogs and cats, as part of a diet adapted to the age, breed, category, activity level,  health and reproductive status of the dogs or cats, with the overall objective of achieving and maintaining their good health;\n\n(c)	feed free of substances which could cause suffering;\n\n(d)	feed in such a way as to avoid abrupt changes and ensure a well-functioning gastro-intestinal system, in particular during the weaning phase.\n\n3.	Operators  shall ensure that feeding and watering facilities are kept clean and are constructed and installed in such a way as to:\n\n(a)	provide equal access to adequate amounts of feed and water for all dogs or cats and minimise competition between them;\n\n(b)	minimise spillage and prevent the contamination of feed and water with harmful physical, chemical or biological contaminants;\n\n(c)	prevent injury, drowning or other harm to the dogs or cats;\n\n(d)	be easily cleaned and disinfected to prevent the spread of diseases.\n\n4.	Where advised in writing by a veterinarian to do so, the operators may adjust the feeding and watering frequencies for an individual dog or cat. The operators shall keep a record of that written advice for the entire duration of those arrangements.",
            "traducao": "1. Os operadores devem assegurar que os cães ou gatos são alimentados em conformidade com os requisitos estabelecidos no ponto 1 do Anexo I.\n\n2. Os operadores devem assegurar que os cães ou gatos são adequadamente alimentados e hidratados através do fornecimento de:\na) água limpa e fresca, ad libitum;\nb) alimento em quantidade e qualidade suficientes para satisfazer as necessidades fisiológicas, nutricionais e metabólicas dos cães e gatos, como parte de uma dieta adaptada à idade, raça, categoria, nível de atividade, saúde e estado reprodutivo dos cães ou gatos, com o objetivo geral de atingir e manter boa saúde;\nc) alimento livre de substâncias que possam causar sofrimento;\nd) alimento de forma a evitar mudanças abruptas e assegurar um sistema gastrointestinal bem funcionante, em particular durante a fase de desmame.\n\n3. Os operadores devem assegurar que as instalações de alimentação e hidratação são mantidas limpas e são construídas e instaladas de forma a:\na) proporcionar acesso igualitário a quantidades adequadas de alimento e água para todos os cães ou gatos e minimizar a competição entre eles;\nb) minimizar derramamentos e evitar a contaminação do alimento e da água com contaminantes físicos, químicos ou biológicos prejudiciais;\nc) evitar ferimentos, afogamento ou outro dano aos cães ou gatos;\nd) ser facilmente limpas e desinfetadas para evitar a propagação de doenças.\n\n4.\tQuando aconselhado por escrito por um médico veterinário, os operadores podem ajustar as frequências de alimentação e hidratação para um cão ou gato individual. Os operadores devem manter um registo do conselho durante toda a sua duração, conforme aconselhado pelo médico veterinário."
        },
        "rgbeac": {
            "ref": "Arts. 7.º (n.º 1, al. a) e 10.º (n.º 1, al. a) do RGBEAC (proposta, jun. 2025)",
            "texto": "Artigo 7.º - Princípios fundamentais: Não passem fome ou sede, nem sejam sujeitos a malnutrição.\n\nArtigo 10.º - Obrigações especiais dos detentores: Alimentos saudáveis, adequados e convenientes ao seu normal desenvolvimento e acesso permanente a água potável. Ênfase em necessidades nutricionais adequadas ao estado de saúde."
        },
        "codigo": {
            "ref": "Art.º 46.º do Código do Animal (DL 214/2013)",
            "texto": "Alimentação e abeberamento:\n\nA alimentação dos animais de companhia, nos locais de criação, manutenção e venda bem como nos centros de recolha e instalações de hospedagem, deve obedecer a um programa de alimentação bem definido, de valor nutritivo adequado e distribuído em quantidade suficiente para satisfazer as necessidades alimentares das espécies.\n\nOs animais devem dispor de água potável e sem qualquer restrição, salvo por razões médico-veterinárias."
        },
        "legislacao": {
            "ref": "Decreto-Lei n.º 276/2001 - Artigo 12.º",
            "texto": "1 — Deve existir um programa de alimentação bem definido, de valor nutritivo adequado e distribuído em quantidade suficiente para satisfazer as necessidades alimentares das espécies e dos indivíduos de acordo com a fase de evolução fisiológica em que se encontram, nomeadamente idade, sexo, fêmeas prenhes ou em fase de lactação.\n2 — As refeições devem ainda ser variadas, sendo distribuídas segundo a rotina que mais se adequar à espécie e de forma a manter, tanto quanto possível, aspetos do seu comportamento alimentar natural.\n3 — O número, formato e distribuição de comedouros e bebedouros deve ser tal que permita aos animais satisfazerem as suas necessidades sem que haja competição excessiva dentro do grupo.\n4 — Os alimentos devem ser preparados e armazenados de acordo com padrões estritos de higiene, em locais secos, limpos, livres de agentes patogénicos e de produtos tóxicos.\n5 — Devem existir aparelhos de frio para uma eficiente conservação dos alimentos.\n6 — Os animais devem dispor de água potável e sem qualquer restrição, salvo por razões médico-veterinárias."
        },
        "divergencia": {
            "legislacao": "NÃO APLICÁVEL - Diploma não específico nesta matéria",
            "codigo": "SIM - COBERTURA COMPLETA (Art. 46.º implementa integralmente)",
            "rgbeac": "SIM - COBERTURA COMPLETA (linguagem modernizada)",
            "sumario": "COBERTURA COMPLETA. Código do Animal (Art. 46.º) implementa todos os requisitos do Artigo 11.º. RGBEAC alinha melhor com linguagem e especificidades do Regulamento europeu. Sem divergências substanciais."
        },
        "necessidade_alteracao": "Não",
        "notas": "Correspondências completas - RGBEAC alinha melhor com Regulamento EU"
    },
    {
        "id": "ART-15",
        "tema": "Alojamento (Housing)",
        "regulamento": {
            "ref": "Art.º 15.º do Regulamento 2023/0447",
            "titulo": "Housing",
            "texto": "1.	The operators of breeding and selling establishments shall ensure that dogs and cats are housed in accordance with point 2 of Annex I. The operators of shelters shall ensure that dogs and cats are housed in accordance with point 2.2 of Annex I.\n\n2.	Operators  shall ensure that:\n\n(a)	the establishments where dogs or cats are kept and the equipment used therein are suitable for the types and the number of dogs or cats, and make possible the necessary access to, and the thorough inspection of, all dogs or cats;\n\n(b)	all building components of the establishment, including the flooring and roof,  and space divisions, as well as the equipment used for dogs or cats, are constructed and maintained properly,  to ensure that they do not pose any risks to the welfare of the dogs or cats;\n\n(c)	all building components of the establishment, including the flooring, and space divisions, as well as the equipment used for dogs or cats, are kept clean to ensure that they do not pose any risks to the welfare of the dogs or cats;\n\n(d)	 in breeding and selling establishments where dogs or cats are kept indoors, the dust levels, the temperature, and the relative air humidity and gas concentrations are  not harmful to dogs or cats and that ventilation is sufficient to avoid overheating  ;\n\n(e)	dogs or cats have enough space to be able to move around freely and to express species-specific behaviour according to their needs with the possibility to withdraw and rest;\n\n(f)	dogs or cats have clean, comfortable and dry resting places that are sufficiently large and numerous to ensure that all of them can lie down and rest in a natural position at the same time;\n\n(g)	appropriate structures and measures are in place for dogs or cats that are kept outdoors in order to protect them from adverse weather conditions, including to prevent thermal stress, sunburn and frostbite.\n\n3.	Operators shall not keep dogs or cats in containers\n\nHowever, containers may be used for transportation, for the short term isolation of individual dogs or cats and for participation in shows, exhibitions and competitions, for puppies or kittens with reduced thermoregulation capacity or for puppies or kittens together with their mothers, provided that, for the dogs or cats concerned, stress is minimised and suffering is avoided, and they are able to stand, turn around and lie down in a natural position.\n\n4.	Operators shall not keep dogs older than 8 weeks exclusively indoors. Such dogs shall have daily access to an outdoor area, or be walked daily,  to allow exercise, exploration and socialisation. The minimum combined duration of such daily access or walk shall be one hour in total. The operator may only deviate from these requirements based on the written advice of a veterinarian.\n\n5.	When cats are kept in catteries, operators shall design and construct individual enclosures to allow cats to move around freely and to exhibit their natural behaviour.\n\n6.	Operators of breeding and selling establishments shall ensure that  in indoor areas where dogs and cats are kept, an appropriate thermoneutral zone is maintained that takes into account their coat type, age, size, breed, and health.\n\n7.	Operators of breeding and selling establishments shall, where necessary, use heating or cooling systems in the indoor enclosures at their establishments to maintain good air quality and an appropriate temperature and to remove excessive moisture.\n\n8.	Operators shall ensure that dogs or cats are exposed to light, and are able to stay in the dark for sufficient and uninterrupted periods in order to maintain a normal circadian rhythm.\n\nFor the purposes of the first subparagraph, ‘light’ means natural light, complemented, where necessary due to the climatic conditions and geographic position of a Member State, by artificial light.\n\n9.	Paragraph 2, points (a), (b), (c) (f) and (g), and paragraphs 6, 7 and 8 shall apply neither to livestock guardian dogs, nor to herding dogs, during the periods where such dogs are used for guarding or herding in the context of seasonal transhumance on foot. Paragraph 2(f) shall not apply to livestock guardian dogs during the periods when such dogs are used for training purposes.",
            "traducao": "1. Os operadores de estabelecimentos de criação e venda devem assegurar que os cães ou gatos dispõem de alojamento em conformidade com o ponto 2 do Anexo I. Os operadores de abrigos devem assegurar que os cães ou gatos dispõem de alojamento em conformidade com o ponto 2.2 do Anexo I.\n\n2. Os operadores devem assegurar que:\na) os estabelecimentos onde os cães ou gatos são mantidos e o equipamento utilizado são adequados aos tipos e número de cães ou gatos, e permitem o acesso necessário e inspeção minuciosa de todos os cães ou gatos;\nb) todos os componentes de construção do estabelecimento, incluindo o pavimento, telhado e divisões de espaço, bem como o equipamento utilizado para cães ou gatos, são construídos e mantidos adequadamente, para assegurar que não apresentam riscos ao bem-estar dos cães ou gatos.\nc)\ttodos os componentes de construção do estabelecimento, incluindo o pavimento e divisões de espaço, bem como o equipamento utilizado para cães ou gatos, são mantidos limpos para assegurar que não apresentam riscos ao bem-estar dos cães ou gatos;\nd)\tem estabelecimentos de criação e venda onde os cães ou gatos são mantidos no interior, pó, temperatura, humidade relativa do ar e concentrações de gases não são prejudiciais aos cães ou gatos e a ventilação é suficiente para evitar sobreaquecimento;\ne)\tos cães e gatos têm espaço suficiente para se moverem livremente e expressar comportamento específico da espécie de acordo com as suas necessidades com possibilidade de se retirarem e descansarem;\nf)\tos cães ou gatos têm lugares de repouso limpos, confortáveis e secos, suficientemente grandes e numerosos para assegurar que todos podem deitar-se e descansar ao mesmo tempo numa posição natural;\ng)\testrututras e medidas apropriadas estão em vigor para cães ou gatos mantidos no exterior para os proteger de condições climáticas adversas, incluindo para evitar stress térmico, queimaduras solares e frieiras.\n\n3. Os operadores não devem manter cães ou gatos em contentores.\n\nA título de derrogação, contentores podem ser usados apenas para transporte, isolamento a curto prazo de cães ou gatos individuais e durante participação em espetáculos, exposições e competições, para cachorros ou gatinhos com capacidade termorreguladora reduzida ou cachorros ou gatinhos juntamente com as suas mães, desde que o stress seja minimizado e o sofrimento seja evitado e os cães e gatos sejam capazes de se manter em pé e deitar-se numa posição natural.\n\n4. Os operadores não devem manter cães com mais de 8 semanas exclusivamente no interior. Tais cães devem ter acesso diário a uma área ao ar livre, ou ser passeados diariamente, para permitir exercício, exploração e socialização. A duração do acesso diário a uma área ao ar livre ou passeio deve ser mínimo uma hora no total. O operador pode apenas derrogar destes requisitos com base em aconselhamento escrito de um médico veterinário.\n\n5. Quando gatos são mantidos em gatarias, os operadores devem desenhar e construir compartimentos individuais para permitir aos gatos mover-se livremente e expressar o seu comportamento natural.\n\n6. Os operadores de estabelecimentos de criação e venda devem assegurar que em áreas interiores onde os cães ou gatos são mantidos, uma zona termoneural apropriada é mantida levando em conta o seu tipo de pelagem, idade, tamanho, raça e saúde.\n\n7.\tOs operadores de estabelecimentos de criação e venda devem usar, quando necessário, sistemas de aquecimento ou arrefecimento para manter boa qualidade do ar, uma temperatura apropriada em compartimentos interiores nos seus estabelecimentos, e remover humidade excessiva.\n\n8.\tOs operadores devem assegurar que os cães ou gatos são expostos à luz, e são capazes de permanecer no escuro por períodos suficientes e ininterruptos para manter um ritmo circadiano normal.\n\nPara efeitos do primeiro parágrafo, 'luz' significa luz natural, complementada, quando necessário, devido às condições climáticas e posição geográfica de um Estado-Membro, por luz artificial.\n\n9.\tOs n.ºs 2(a), (b), (c), (f) e (g), 6, 7 e 8 não se aplicam a cães guardiões de gado, nem a cães de pastoreio, durante os períodos em que tais cães são utilizados para guarda ou pastoreio no contexto de transumância sazonal a pé. O n.º 2(f) não se aplica a cães guardiões de gado durante os períodos em que tais cães são utilizados para fins de treino."
        },
        "rgbeac": {
            "ref": "Arts. 7.º, 10.º, 11.º, 47-57 do RGBEAC (proposta, jun. 2025)",
            "texto": "Artigo 7.º - Princípios: Condições de detenção e alojamento salvaguardam bem-estar animal.\n\nArtigo 10.º - Obrigações especiais: Liberdade de movimento, proibição de contenção permanente, espaço adequado, enriquecimento ambiental e abrigo protetor.\n\nArtigo 11.º - Obrigações especiais relativas ao alojamento doméstico.\n\nArtigos 47-57 - Regulação detalhada de alojamentos para hospedagem (estruturas, proteção, maneio, responsabilidades veterinárias)."
        },
        "codigo": {
            "ref": "Arts. 3.º, 14.º, 18.º, 28.º do Código do Animal (DL 214/2013)",
            "texto": "Artigo 3.º - Define 'Alojamento' como qualquer instalação, edifício ou local onde animais se encontram mantidos.\n\nArtigo 14.º - A temperatura, ventilação, luminosidade e obscuridade devem ser adequadas ao conforto e bem-estar.\n\nArtigo 18.º - Alojamentos devem possuir instalações para armazenagem, lavagem, quarentena, enfermaria e higienização.\n\nArtigo 28.º - Define separação por espécie e requisitos de estrutura para hospedagem."
        },
        "legislacao": {
            "ref": "Decreto-Lei n.º 276/2001 - Artigos 8.º e 15.º",
            "texto": "Artigo 8.º — 1 — Os animais devem dispor do espaço adequado às suas necessidades fisiológicas e etológicas, devendo o mesmo permitir: a) A prática de exercício físico adequado; b) A fuga e refúgio de animais sujeitos a agressão por parte de outros.\n2 — Os animais devem poder dispor de esconderijos para salvaguarda das suas necessidades de proteção, sempre que o desejarem.\n3 — As fêmeas em período de incubação, de gestação ou com crias devem ser alojadas de forma a assegurarem a sua função reprodutiva natural em situação de bem-estar.\n4 — As estruturas físicas das instalações, todo o equipamento nelas introduzido e a vegetação não podem representar nenhum tipo de ameaça ao bem-estar dos animais.\nArtigo 15.º — Os alojamentos devem assegurar que as espécies animais neles mantidas não possam causar quaisquer riscos para a saúde e para a segurança de pessoas, outros animais e bens."
        },
        "divergencia": {
            "legislacao": "PARCIAL - Não específico para condições técnicas de alojamento",
            "codigo": "SIM - COBERTURA COMPLETA (Arts. 3, 14, 18, 28)",
            "rgbeac": "SIM - COBERTURA EXPANDIDA (especifica detalhes técnicos)",
            "sumario": "COBERTURA COMPLETA. Código do Animal e RGBEAC implementam requisitos do Artigo 12.º. Faltam especificações técnicas detalhadas (temperatura, ventilação, iluminação) — recomenda-se Portaria complementar."
        },
        "necessidade_alteracao": "Sim - Portaria complementar com especificações técnicas",
        "notas": "Princípios cobertos; faltam normas técnicas pormenorizadas"
    },
    {
        "id": "ART-16",
        "tema": "Saúde e Monitorização Sanitária",
        "regulamento": {
            "ref": "Art.º 16.º do Regulamento 2023/0447 (PE/Conselho)",
            "titulo": "Health",
            "texto": "1.	Operators  shall ensure that:\n\n(a)	dogs or cats for which they are responsible are inspected by animal carers at least once a day and that vulnerable dogs and cats, such as newborns, ill or injured dogs and cats, and peri-partum bitches and queens, are inspected more frequently;\n\n(b)	dogs or cats with  compromised welfare are, where necessary, transferred  without undue delay to a separate area and, where necessary, receive appropriate treatment;\n\n(c)	where the recovery of a dog or a cat whose welfare is compromised is not achievable and the dog or cat experiences severe pain or suffering, a veterinarian is consulted without undue delay to decide whether the dog or cat is to be euthanised to end its suffering, and, if that is the case, to perform the euthanasia using anaesthesia and analgesia.\n\n(d)	measures  are taken to prevent and control external and internal parasites, and vaccinations are carried out to prevent common diseases to which dogs or cats are likely to be exposed.\n\n(e)	enrichments  that are used do not present a significant risk to dogs and cats of injury or biological or chemical contamination or any other health risk.\n\nPoint (a) shall not apply to livestock guardian dogs kept in breeding establishments during the periods when such dogs are used for guarding or training purposes.\n\nMember States may grant exemptions from point (c) in cases of emergency, where no veterinarian can be reached without undue delay, provided that national rules are put in place to ensure that:\n\n(i)	any immediate action ending the life of the dog or cat with minimum pain and suffering using a method inducing instant death is undertaken by a trained competent person;\n\n(ii)	for the purposes of the official control under Regulation (EU) 2017/625, the operator keeps a record of the use of the exemption.\n\n2.	Operators of breeding establishments shall ensure that:\n\n(a)	measures are taken to safeguard the health of dogs or cats in accordance with point 3 of Annex I;\n\n(b)	bitches or queens are bred only if they have reached a minimum age and skeletal maturity in accordance with point 3 of Annex I  , and only if they have no diagnosed disease, clinical sign of diseases or physical conditions which could negatively impact their pregnancy and welfare;\n\n(c)	the litter-giving pregnancies of bitches or queens follows a maximum frequency in accordance with point 3 of Annex I;\n\n(d)	lactating queens are not mated or inseminated;\n\n(e)	 dogs and cats which are no longer used for reproduction, including as a result of the provisions of this Regulation, are either kept or sold, donated or rehomed, and not killed or abandoned.",
            "traducao": "1 — Os operadores devem assegurar que:\n\n(a) os cães e gatos sob a sua responsabilidade são inspecionados por cuidadores pelo menos uma vez por dia, e os animais vulneráveis, como recém-nascidos, doentes, lesionados e fêmeas em período peri-parto, são inspecionados com maior frequência;\n(b) os cães e gatos com bem-estar comprometido são, quando necessário, transferidos sem demora injustificada para uma área separada e, se necessário, recebem tratamento adequado;\n(c) quando a recuperação de um cão ou gato com bem-estar comprometido não seja alcançável e o animal experiencie dor ou sofrimento severo, um médico veterinário é consultado sem demora injustificada para decidir se o animal deve ser objeto de eutanásia para pôr termo ao seu sofrimento e, em caso afirmativo, para realizar a eutanásia com recurso a anestesia e analgesia;\n(d) são implementadas medidas de prevenção e controlo de parasitas externos e internos, bem como vacinações para prevenção de doenças comuns às quais os cães ou gatos são suscetíveis de estar expostos;\n(e) os enriquecimentos não apresentam risco significativo de lesões ou de contaminação biológica ou química, nem qualquer outro risco para a saúde.\n\nA alínea (a) não se aplica aos cães de guarda de gado mantidos em estabelecimentos de criação durante os períodos em que tais cães são utilizados para fins de guarda ou treino.\n\nOs Estados-Membros podem conceder derrogações relativamente à alínea (c) em casos de emergência, quando não seja possível contactar um médico veterinário sem demora injustificada, desde que sejam estabelecidas regras nacionais que assegurem que:\n(i) é tomada imediatamente uma ação que ponha fim à vida do cão ou gato com o mínimo de dor e sofrimento, utilizando um método que induza a morte instantânea, por uma pessoa competente e devidamente habilitada;\n(ii) o operador mantém um registo da utilização da derrogação para efeitos de controlo oficial.\n\n2 — Os operadores de estabelecimentos de criação devem adicionalmente assegurar que:\n\n(-a) são tomadas medidas para salvaguardar a saúde dos cães ou gatos em conformidade com o ponto 3 do Anexo I;\n(-b) as cadelas ou gatas só são reproduzidas se tiverem atingido a idade mínima e a maturidade esquelética em conformidade com o ponto 3 do Anexo I, e não apresentem doença diagnosticada, sinais clínicos de doença ou condições físicas que possam impactar negativamente a gestação e o seu bem-estar;\n(-c) a frequência de gestações com ninhadas de cadelas ou gatas respeita a frequência máxima fixada no ponto 3 do Anexo I;\n(-d) as gatas a lactar não são acasaladas nem inseminadas;\n(-e) os cães e gatos que deixem de ser utilizados para reprodução, nomeadamente em resultado das disposições do presente Regulamento, são mantidos ou vendidos, doados ou realojados, não sendo mortos nem abandonados."
        },
        "rgbeac": {
            "ref": "Art.º 33.º do RGBEAC — Regime Geral do Bem-Estar dos Animais de Companhia (proposta, jun. 2025)",
            "texto": "Artigo 33.º — Cuidados de saúde\n\n1 — Os detentores dos animais de companhia devem assegurar-lhes os cuidados de saúde adequados, nomeadamente seguindo as orientações da DGAV em matéria de vacinação e tratamentos obrigatórios, bem como consultas regulares junto de médico veterinário.\n\n2 — Os animais que apresentem sinais que levem a suspeitar de poderem estar doentes ou lesionados devem receber os primeiros cuidados pelo detentor e, se não houver indícios de recuperação, devem ser tratados por médico veterinário.\n\n3 — Os médicos veterinários e os centros de atendimento médico-veterinário (CAMV) devem manter um arquivo com os dados clínicos de cada animal, pelo período mínimo de cinco anos, que ficará à disposição das autoridades competentes.\n\n[dim]4 — Os CAMV, enquanto estabelecimentos de saúde, colaborarão na vigilância epidemiológica das doenças de notificação obrigatória que detetem e no seu controlo.\n\n[dim]5 — Os médicos veterinários denunciam, junto da DGAV ou demais entidades com competência de fiscalização do cumprimento das normas constantes do presente decreto-lei, sempre que, no exercício de sua profissão, suspeitem de maus-tratos a animais de companhia."
        },
        "codigo": {
            "ref": "Art.º 6.º (Cuidados médico-veterinários) do Código do Animal — DL n.º 214/2013",
            "texto": "Artigo 6.º — Cuidados médico-veterinários\n\nO detentor do animal deve assegurar ao animal ferido ou doente os cuidados médico-veterinários adequados, designadamente retirando o mesmo do alojamento sempre que este seja um local de venda."
        },
        "legislacao": {
            "ref": "Art.º 13.º (Maneio) e art.º 16.º (Cuidados de saúde animal) do DL n.º 276/2001, de 17 de outubro",
            "texto": "Artigo 13.º — Maneio\n\n1 — A observação diária dos animais e o seu maneio, a organização da dieta e o tratamento médico-veterinário devem ser assegurados por pessoal técnico competente e em número adequado à quantidade e espécies animais que alojam.\n\n[dim]2 — O maneio deve ser feito por pessoal que possua formação teórica e prática específica ou sob a supervisão de uma pessoa competente para o efeito.\n\n3 — Todos os animais devem ser alvo de inspeção diária, sendo de imediato prestados os primeiros cuidados aos que tiverem sinais que levem a suspeitar estarem doentes, lesionados ou com alterações comportamentais.\n\n[dim]4 — O manuseamento dos animais deve ser feito de forma a não lhes causar quaisquer dores, sofrimento ou distúrbios desnecessários.\n\n[dim]5 — Quando houver necessidade de recorrer a meios de contenção, não devem estes causar ferimentos, dores ou angústia desnecessários aos animais.\n\nArtigo 16.º — Cuidados de saúde animal\n\n1 — Sem prejuízo de quaisquer medidas determinadas pela DGAV, deve existir um programa de profilaxia médica e sanitária devidamente elaborado e supervisionado pelo médico veterinário responsável e executado por profissionais competentes.\n\n2 — No âmbito do número anterior, os animais devem ser sujeitos a exames médico-veterinários de rotina, vacinações e desparasitações sempre que aconselhável.\n\n3 — Os animais que apresentem sinais que levem a suspeitar de poderem estar doentes ou lesionados devem receber os primeiros cuidados pelo detentor e, se não houver indícios de recuperação, devem ser tratados por médico veterinário.\n\n4 — Sempre que se justifique, os animais doentes ou lesionados devem ser isolados em instalações adequadas e equipadas, se for caso disso, com cama seca e confortável.\n\n[dim]5 — Os medicamentos, produtos ou substâncias de prescrição médico-veterinária devem ser armazenados em locais secos e com acesso restrito.\n\n[dim]6 — A administração e utilização de medicamentos, produtos ou substâncias referidas no número anterior deve ser feita sob orientação do médico veterinário responsável."
        },
        "divergencia": {
            "legislacao": "O DL n.º 276/2001 prevê inspeção diária (n.º 3 do art.º 13.º) e programa de profilaxia supervisionado por veterinário (n.º 1 do art.º 16.º) — alinhamento parcial com as als. (a) e (d) do n.º 1 do art.º 13.º do @regulamento. Lacunas: (1) não distingue animais vulneráveis com maior frequência de inspeção; (2) não prevê isolamento em área separada com tratamento (al. (b)); (3) não exige consulta veterinária para decisão de eutanásia com anestesia/analgesia (al. (c)); (4) nenhuma das obrigações sanitárias para criadores previstas no n.º 2 está contemplada: sem idade mínima de reprodução, sem frequência máxima de partos, sem proibição de cobrição de fêmeas a lactar, sem regime de rehoming.",
            "codigo": "O @codigo limita os cuidados médico-veterinários ao animal ferido ou doente (art.º 6.º) — sem qualquer obrigação de inspeção diária, programa de profilaxia ou isolamento. É o diploma com maior divergência face ao n.º 1 do art.º 13.º do @regulamento, omitindo igualmente todas as obrigações sanitárias específicas para criadores previstas no n.º 2.",
            "rgbeac": "O @rgbeac (art.º 33.º) centra os cuidados de saúde no detentor em geral (n.ºs 1 a 3), sem distinguir obrigações reforçadas para operadores de criação. Não prevê: inspeção diária sistemática por cuidadores; isolamento imediato de animais com bem-estar comprometido; nem qualquer dos requisitos sanitários para criadores do n.º 2 do art.º 13.º do @regulamento. Os n.ºs 4 e 5 (CAMV e denúncia de maus-tratos) não têm correspondência direta no @regulamento.",
            "sumario": "A @legislacao vigente (DL n.º 276/2001) tem o maior alinhamento, mas insuficiente. Necessidade de alteração dos três diplomas: (1) introduzir inspeção diária diferenciada para animais vulneráveis (al. (a) do n.º 1 do art.º 13.º do @regulamento); (2) criar obrigação de isolamento e tratamento imediato (al. (b)); (3) regular o processo de eutanásia com supervisão veterinária e anestesia/analgesia (al. (c)); (4) para criadores (n.º 2): fixar idade mínima e maturidade esquelética das fêmeas reprodutoras, frequência máxima de partos conforme Anexo I, proibição de cobrição de fêmeas a lactar, e regime de rehoming dos animais retirados da reprodução."
        },
        "necessidade_alteracao": "Sim",
        "notas": ""
    },
    {
        "id": "ART-17",
        "tema": "Necessidades Comportamentais (Behavioural needs)",
        "regulamento": {
            "ref": "Art.º 17.º do Regulamento 2023/0447",
            "titulo": "Behavioural needs",
            "texto": "1.	Operators  shall ensure that measures are taken to meet the behavioural needs of  dogs and cats in accordance with point 4 of Annex I.\n\n2.	In addition, operators shall not keep dogs or cats in areas which limit their natural movements  , except in cases where Article 15(3), second subparagraph, applies or where the following procedures or treatments are performed:\n\n(a)	physical examinations ;\n\n(b)	the individual identification of dogs or cats, or reading such identification information;\n\n(c)	the collection of samples and vaccinations;\n\n(d)	procedures for grooming, hygienic, health or reproductive purposes other than mating;\n\n(e)	medical treatments, including surgical treatments or prescribed rehabilitation.\n\n3.	Tethering  for more than 1 hour shall be prohibited, except for the duration of a medical treatment or for participation in shows, exhibitions and competitions of dogs or cats.\n\n4.	Member States may grant exemptions from paragraph 3 for dogs intended for use in military, police and customs services that are kept in breeding or selling establishments.\n\n5.	Operators  shall ensure that dogs or cats are kept in conditions that allow them to exhibit non-harmful social behaviours and species-specific behaviours and to experience positive emotions.\n\n6.	Operators shall ensure that dogs or cats can socialise in accordance with point 4 of Annex I. Operators of breeding establishments shall have a documented strategy for such socialisation.\n\nBy way of derogation from the first subparagraph, socialisation requirements shall not apply to livestock guardian dogs kept in breeding establishments during the periods when such dogs are used for guarding or training purposes, or to herding dogs during seasonal transhumance.\n\n7.	Operators shall ensure that enrichments  are provided and accessible to all dogs or cats, creating a stimulating environment for them, enabling them to develop and exhibit species-specific behaviour and reducing their frustration.",
            "traducao": "1. Os operadores devem assegurar que medidas são tomadas para satisfazer as necessidades comportamentais de cães ou gatos em conformidade com o ponto 4 do Anexo I.\n\n2. Os operadores não devem manter cães ou gatos em áreas que restringem os seus movimentos naturais, exceto no caso do artigo 15.º (parágrafo 3), segundo parágrafo, ou para realizar os seguintes procedimentos ou tratamentos:\na) exames físicos;\nb) identificação individual de cães ou gatos e leitura da informação de identificação;\nc) recolha de amostras e vacinações;\nd) procedimentos de higiene, higiénicos, de saúde ou reprodutivos que não sejam acasalamento;\ne) tratamento médico, incluindo tratamento cirúrgico ou reabilitação prescrita.\n\n3. O amarrar por mais de 1 hora é proibido, exceto durante a duração de um tratamento médico ou participação em espetáculos, exposições e competições de cães e gatos.\n\n4.\tOs Estados-Membros podem conceder derrogações do n.º 3 para cães destinados a uso em serviços militares, policiais e aduaneiros que são mantidos em estabelecimentos de criação ou venda.\n\n5.\tOs operadores devem assegurar que condições estão em vigor para permitir aos cães ou gatos expressar comportamentos sociais não prejudiciais, comportamentos específicos da espécie e a possibilidade de experimentar emoções positivas.\n\n6.\tOs operadores devem assegurar que os cães ou gatos podem socializar em conformidade com o ponto 4 do Anexo I. Os operadores de estabelecimentos de criação devem ter uma estratégia documentada para tal socialização.\n\n\tA título de derrogação do primeiro parágrafo, os requisitos de socialização não se aplicam a cães guardiões de gado mantidos em estabelecimentos de criação durante os períodos em que tais cães são utilizados para guarda ou treino, nem a cães de pastoreio durante transumância sazonal.\n\n7.\tOs operadores devem assegurar que enriquecimento é fornecido e acessível a todos os cães ou gatos, criando um ambiente estimulante, permitindo comportamento específico da espécie e reduzindo a sua frustração."
        },
        "rgbeac": {
            "ref": "RGBEAC (proposta, jun. 2025) - Artigos 10, 12, 13, 14, 15",
            "texto": "Especificação clara: 'exercício físico e estímulo mental'.\n'Contato social adequado'.\nMétodos de 'reforço positivo' (OBRIGATÓRIO).\nProibição explícita de 'métodos aversivos, punitivos ou violentos'.\nDocumentação obrigatória de estratégia de socialização (criadores)."
        },
        "codigo": {
            "ref": "Arts. 5.º e 13.º do Código do Animal (DL 214/2013)",
            "texto": "Artigo 5.º - Princípios que proíbem violência e maus-tratos, garantem bem-estar.\n\nArtigo 13.º - Espaço para exercício físico e expressão de comportamentos naturais.\n\nCobertura GENÉRICA: não especifica enriquecimento, socialização ou método de treino baseado em reforço positivo."
        },
        "legislacao": {
            "ref": "Decreto-Lei n.º 276/2001 - Artigo 8.º",
            "texto": "Artigo 8.º — 1 — Os animais devem dispor de um espaço adequado às suas necessidades fisiológicas e etológicas, devendo o mesmo permitir: a) A prática de exercício físico adequado; b) A fuga e refúgio de animais sujeitos a agressão por parte de outros.\n2 — Os animais devem poder dispor de esconderijos para salvaguarda das suas necessidades de proteção, sempre que o desejarem."
        },
        "divergencia": {
            "legislacao": "PARCIAL - Legislação vigente oferece cobertura genérica",
            "codigo": "PARCIAL - Genérico, sem especificações sobre socialização e enriquecimento",
            "rgbeac": "SIM - COBERTURA SIGNIFICATIVAMENTE EXPANDIDA",
            "sumario": "Legislação portuguesa oferece cobertura genérica. RGBEAC (2025) oferece avanço substancial com obrigatoriedade de reforço positivo e documentação de estratégia de socialização. Falta ainda regulamentação pormenorizada."
        },
        "necessidade_alteracao": "Sim - Regulamentação específica sobre métodos de treino",
        "notas": "RGBEAC alinha melhor; implementação de reforço positivo obrigatório recomendada"
    },
    {
        "id": "ART-18",
        "tema": "Práticas Dolorosas (Painful practices)",
        "regulamento": {
            "ref": "Art.º 18.º do Regulamento 2023/0447",
            "titulo": "Painful practices",
            "texto": "1.	Operators shall ensure that mutilations, including ear cropping, tail docking, claw removal or other partial or complete digit amputation, and resection of vocal cords or folds,  are not performed unless justified by medical indications, including where the procedure is prophylactic, with the sole purpose of maintaining or improving the health of dogs or cats or preventing their injury. In such cases, the procedure shall be performed  only under anaesthesia and prolonged analgesia and only by a veterinarian.\n\n2.	The medical indications justifying the mutilation, and the details of procedure carried out, shall be documented by a veterinarian. That document shall be retained by the operator and shall accompany the dog or cat when it is transferred to another establishment or owner. The operator of the establishment where the mutilation was performed shall retain a copy of the document for the first three years after that transfer.\n\n3.	By way of derogation from paragraph 1, Member States may allow ear cropping by notching or tipping cat ears in the context of marking stray cats when neutered under a ‘trap-neuter-return’ programme.\n\n4.	Operators shall ensure that neutering is performed  only under anaesthesia and prolonged analgesia and only by a veterinarian. However, Member States may allow the neutering of male cats to be performed by a licensed veterinary nurse.\n\n5.	Operators shall ensure that handling practices that cause pain or suffering are not performed, including:\n\n(a)	tying up body parts, unless required for medical reasons and limited to the minimum period necessary;\n\n(b)	the kicking, hitting, dragging, throwing or squeezing of dogs or cats;\n\n(c)	applying electric current to dogs or cats, unless performed for medical reasons;\n\n(d)	using  muzzles, unless required for medical reasons or animal or human safety reasons, limited to the minimum period necessary and where the dog or cat is supervised.\n\n(e)	using prong collars;\n\n(f)	using choke collars without a safety stop;\n\n(g)	lifting dogs or cats by the limbs,  head, ears, tail or hair, or lifting adult dogs or cats by the skin.\n\nMember States may grant exemptions from the first subparagraph for dogs intended for use in military, police or customs services.",
            "traducao": "1. Os operadores devem assegurar que mutilações, incluindo corte de orelhas, corte de cauda, remoção de garras ou amputação parcial ou completa de dígitos, e ressecção de cordas vocais ou pregas, não são realizadas a menos que por indicação médica, que pode incluir profilática, com o único propósito de preservar, melhorar a saúde de cães ou gatos ou prevenir ferimentos. Nesse caso, o procedimento deve ser realizado apenas sob anestesia e analgesia prolongada e por um médico veterinário.\n\n2.\tA indicação médica para a mutilação e os detalhes do procedimento realizado devem ser documentados por um médico veterinário. Este documento deve ser retido pelo operador até que o cão ou gato, juntamente com este documento, seja transferido para outro estabelecimento ou proprietário. O operador do estabelecimento onde a mutilação foi realizada deve reter uma cópia do documento durante três anos após a transferência do cão ou gato.\n\n3.\tA título de derrogação do parágrafo 1, os Estados-Membros podem permitir o corte de orelhas por entalhe ou ponta das orelhas de gatos no contexto de marcação de gatos vadios quando esterilizados sob um programa de captura-esterilização-libertação.\n\n4.\tOs operadores devem assegurar que a esterilização é realizada apenas sob anestesia e analgesia prolongada e por um médico veterinário. A título de derrogação, os Estados-Membros podem permitir que a esterilização de gatos machos seja realizada por um enfermeiro veterinário licenciado.\n\n5.\tOs operadores devem assegurar que práticas de manipulação que causam dor ou sofrimento não são realizadas, incluindo:\na)\tamarrar partes do corpo a menos que por razões médicas em cujo caso a duração deve ser limitada ao período mínimo necessário;\nb)\tchutar, bater, arrastar, atirar, apertar cães ou gatos;\nc)\taplicar corrente elétrica a cães ou gatos a menos que realizado por razões médicas;\nd)\tuso de focinheiras, a menos que necessário por razões médicas, segurança animal ou humana, em cujo caso a duração deve ser limitada ao período mínimo necessário e o cão ou gato deve ser supervisionado.\ne)\tuso de colares de espinhos;\nf)\tuso de colares de estrangulamento sem paragem de segurança;\ng)\tlevantar cães ou gatos pelas extremidades, cabeça, orelhas, cauda ou pêlo, ou levantar cães ou gatos adultos pela pele.\n\n\tOs Estados-Membros podem conceder derrogações do primeiro parágrafo para cães destinados a uso em serviços militares, policiais ou aduaneiros."
        },
        "rgbeac": {
            "ref": "RGBEAC (proposta, jun. 2025) - Artigo 12.º",
            "texto": "Lista idêntica de mutilações proibidas ao Código.\nReferência a 'boas práticas internacionais' (alinhamento com Reg. 2023/0447).\nAlargamento: 'qualquer amputação sem razão médica veterinária'.\nÊnfase em anestesia e analgesia prolongada.\nDocumentação obrigatória de indicação médica."
        },
        "codigo": {
            "ref": "Arts. 51.º e 52.º do Código do Animal (DL 214/2013)",
            "texto": "Artigo 51.º - Intervenções cirúrgicas exclusivamente por médico veterinário.\n\nArtigo 52.º - Proibição específica de mutilações:\n- Corte de orelhas (exceto fins medicinais)\n- Corte de cauda (revogado em 2015)\n- Ressecção de cordas vocais\n- Remoção de unhas/dentes\n- Exceções: reprodução e interesse do animal (com documentação)"
        },
        "legislacao": {
            "ref": "Decreto-Lei n.º 276/2001 - Artigo 18.º",
            "texto": "Artigo 18.º — 1 — Os detentores de animais de companhia que os apresentem com quaisquer amputações que modifiquem a aparência dos animais ou com fins não curativos devem possuir documento comprovativo, passado pelo médico veterinário que a elas procedeu, da necessidade dessa amputação.\n2 — O documento referido no número anterior deve ter a forma de um atestado, do qual constem a identificação do médico veterinário, o número da cédula profissional e a sua assinatura."
        },
        "divergencia": {
            "legislacao": "SIM - DL 276/2001 cobre amputações documentadas",
            "codigo": "SIM - COBERTURA COMPLETA (Arts. 51-52)",
            "rgbeac": "SIM - COBERTURA COMPLETA + EXPANSÃO",
            "sumario": "COBERTURA COMPLETA. Código do Animal (Arts. 51-52) implementa integralmente Art. 15.º. RGBEAC alinha substancialmente com Regulamento europeu. Sem divergências substanciais."
        },
        "necessidade_alteracao": "Não",
        "notas": "Correspondências completas - Cobertura legislativa adequada"
    },
    {
        "id": "ART-19",
        "tema": "Espetáculos e Competições Estéticas",
        "regulamento": {
            "ref": "Art.º 19.º do Regulamento 2023/0447",
            "titulo": "Aesthetic shows, exhibitions and competitions",
            "texto": "1.	In aesthetic shows, exhibitions and competitions of dogs and cats, operators of breeding and selling establishments shall not use dogs or cats with excessive conformational traits, or dogs or cats which have been mutilated in a way that alters their physical characteristics.\n\n2.	When organising aesthetic shows, exhibitions and competitions of dogs and cats, organisers shall exclude dogs and cats which have excessive conformational traits or dogs or cats which have been mutilated in a way that alters their physical characteristics.\n\nCHAPTER III\nIDENTIFICATION AND REGISTRATION OF DOGS AND CATS",
            "traducao": "Os operadores de estabelecimentos de criação e venda não devem utilizar em espetáculos, exposições e competições estéticas de cães e gatos, cães ou gatos com características conformacionais excessivas ou cães ou gatos que tenham sido mutilados de tal forma que resulte numa alteração de características físicas.\n\nOs organizadores de espetáculos, exposições e competições estéticas de cães e gatos devem excluir de tais espetáculos, exposições e competições cães e gatos que tenham características conformacionais excessivas ou cães ou gatos que tenham sido mutilados de tal forma que resulte numa alteração de características físicas."
        },
        "rgbeac": {
            "ref": "Art.º 39.º do RGBEAC (proposta, jun. 2025)",
            "texto": "Artigo 39.º - Participação em eventos\n\n1 – A participação de animais de companhia em concursos, exposições, espetáculos, manifestações culturais, divertimentos públicos, atividades performativas, cinematográficas e audiovisuais, campanhas publicitárias, ou outros eventos onde participem animais de companhia carece de autorização do diretor geral da DGAV a área da realização da mesma, após parecer da respetiva câmara municipal.\n\n3 - Só serão admitidos no evento os animais de companhia que: a) Estejam registados no SIAC; b) Quando aplicável, possuam prova de vacinação antirrábica; c) Possuam vacinações contra as principais doenças infectocontagiosas; d) Não tenham sido submetidos a intervenções cirúrgicas em infração."
        },
        "codigo": {
            "ref": "Art.º 79.º do Código do Animal (DL 214/2013)",
            "texto": "Artigo 79.º - Concursos e exposições\n\n1 - A realização de concursos e exposições com animais de companhia carece de autorização prévia da câmara municipal, ficando esta dependente do parecer vinculativo do MVM.\n\n3 - Só são admitidos a concurso os cães e gatos que: a) Estejam identificados eletronicamente; b) Sejam portadores de boletim sanitário e prova de vacinação antirrábica; c) Possuam vacinações contra principais doenças infecto-contagiosas."
        },
        "legislacao": {
            "ref": "Lei n.º 27/2016 e DL n.º 82/2019 (Normas de eventos)",
            "texto": "A legislação portuguesa estabelece que a participação de animais em espetáculos e competições requer: - Autorização prévia de autoridades competentes; - Identificação e registo no SIAC; - Vacinações obrigatórias; - Supervisão veterinária durante o evento; - Condições de bem-estar animal garantidas."
        },
        "divergencia": {
            "legislacao": "SIM - Cobertura COMPLETA (autorização, identificação, vacinação, supervisão veterinária)",
            "codigo": "SIM - Cobertura COMPLETA (concursos e exposições com requisitos específicos)",
            "rgbeac": "SIM - Cobertura COMPLETA (participação em eventos com normas de bem-estar)",
            "sumario": "A legislação portuguesa cobre completamente os requisitos do Artigo 15a. Implementa autorizações prévias, exigências de identificação, vacinação obrigatória e supervisão veterinária. O RGBEAC (Art. 39.º) e Código do Animal (Art. 79.º) estabelecem normas detalhadas sobre espetáculos, exposições e competições estéticas."
        },
        "necessidade_alteracao": "Não",
        "notas": "Correspondências completas encontradas - Cobertura legislativa adequada"
    },
    {
        "id": "ART-20",
        "tema": "Identificação e Registo",
        "regulamento": {
            "ref": "Art.º 20.º do Regulamento 2023/0447",
            "titulo": "Identification and registration of dogs and cats",
            "texto": "1.	All dogs and cats kept in establishments placed on the market or owned by pet owners or by any other natural or legal persons, shall be individually identified by means of a single injectable transponder containing a readable microchip that complies with the requirements set out in Annex II.\n\n2.	Operators shall ensure that dogs and cats born in their establishments are individually identified within three months after their birth, and in any event before the date that they are placed on the market.\n\nOperators of selling establishments and shelters, and operators who place and are responsible for dogs and cats in foster homes shall ensure that dogs and cats that enter their establishments or come under their responsibility are individually identified within 30 days of arrival and in any event before the date of their placing on the market.\n\nPet owners and any other natural or legal persons other than operators who own dogs or cats, shall ensure that every dog or cat is individually identified at the latest when it reaches the age of three months or, if the dog or cat is placed on the market, before the date of that placing on the market.\n\nThe implantation of the transponder shall be performed by a veterinarian. Member States may allow the implantation of transponders by persons other than veterinarians provided that they adopt national rules laying down  the minimum qualifications that such persons are required to have.\n\nWhere dogs and cats have been individually identified by means of an injectable transponder containing a microchip in accordance with Union or national law before … [date two years after the date of entry into force of this Regulation] they shall be considered to be compliant with the requirements in paragraph 1 and the first, second, third and fourth subparagraphs of this paragraph, provided that the microchip is still readable.\n\n3.	Within two working days after their identification, the dogs and cats  shall be registered by a veterinarian  in a national database referred to in Article 23. Member States may allow registration by persons other than veterinarians, provided that the Member States have measures in place to ensure the accuracy of information that those persons enter in the database. For dogs and cats kept in  establishments, the registration shall be made in the name of the operator of the  establishment responsible for the dog or the cat. For dogs and cats owned by any other natural and legal persons  , the registration shall be made in the name of those persons.\n\nMember States may grant exemptions from the first subparagraph of this paragraph in respect of military, police and customs dogs.\n\n4.	Where dogs or cats are placed on the market or are donated in an occasional manner by natural persons without using online advertising, the natural or legal person transferring the ownership of, or responsibility for, the dog or cat shall ensure that the change of ownership of, or responsibility for, the dog or cat is recorded in the database referred to in Article 23, within two weeks from the date of that transfer, in accordance with the conditions laid down by the Member State responsible for that database.\n\n5.	In the case of the death of a dog or a cat, the operator, pet owner or natural or legal person owning the dog or cat shall ensure that the death is recorded in the database referred to in Article 23, in accordance with the conditions laid down by the Member State responsible for that database.\n\n6.	Where a transponder is or becomes unreadable, the operator or the natural or legal person responsible for the dog or cat shall ensure that a new transponder is injected and that the registration in the database is updated with the identification number of that new transponder.\n\n7.	The identification and registration requirements of this Article shall apply as follows:\n\n(a) 	for operators and natural or legal persons placing dogs and cats on the market from ... [4 years from the entry into force of this Regulation];\n\n(b)	for pet owners and other natural or legal persons other than operators, who do not place dogs on the market: from … [10 years from entry into force of this Regulation];\n\n(c)	 for pet owners and other natural or legal persons other than operators, who do not place cats on the market: from … [15 years from entry into force of this Regulation.",
            "traducao": "1.\tTodos os cães e gatos mantidos em estabelecimentos, colocados no mercado ou detidos por donos de animais de companhia ou por qualquer outra pessoa singular ou coletiva, devem ser identificados individualmente por meio de um único transponder injetável contendo um microchip legível em conformidade com os requisitos estabelecidos no Anexo II.\n\n2.\tOs operadores devem assegurar que os cães e gatos nascidos nos seus estabelecimentos sejam identificados individualmente no prazo de três meses após o nascimento e, em qualquer caso, antes da data da sua colocação no mercado.\n\nOs operadores de estabelecimentos de venda e abrigos, e os operadores que colocam e são responsáveis por cães e gatos em famílias de acolhimento, devem assegurar que os cães e gatos que entrem nos seus estabelecimentos ou fiquem sob a sua responsabilidade sejam identificados individualmente no prazo de 30 dias após a chegada e, em qualquer caso, antes da data da sua colocação no mercado.\n\nOs donos de animais de companhia e quaisquer outras pessoas singulares ou coletivas que não sejam operadores e que sejam proprietários de cães ou gatos devem assegurar que cada cão ou gato seja identificado individualmente o mais tardar quando o animal atingir a idade de três meses ou, caso o cão ou gato seja colocado no mercado, antes da data dessa colocação no mercado.\n\nA implantação do transponder deve ser efetuada por um médico veterinário. Os Estados-Membros podem permitir que a implantação de transponders seja efetuada por pessoas que não sejam médicos veterinários, desde que adotem regras nacionais que estabeleçam as qualificações mínimas que essas pessoas devem ter.\n\nCaso os cães e gatos tenham sido identificados individualmente por meio de um transponder injetável contendo um microchip em conformidade com o direito da União ou nacional antes de … [data dois anos após a data de entrada em vigor do presente Regulamento], considera-se que cumprem os requisitos do n.º 1 e do primeiro, segundo, terceiro e quarto parágrafos do presente número, desde que o microchip seja ainda legível.\n\n3.\tNo prazo de dois dias úteis após a sua identificação, os cães e gatos devem ser registados por um médico veterinário numa base de dados nacional referida no artigo 23.º. Os Estados-Membros podem permitir o registo por pessoas que não sejam médicos veterinários, desde que os Estados-Membros disponham de medidas para assegurar a exatidão das informações que essas pessoas introduzem na base de dados. No caso de cães e gatos mantidos em estabelecimentos, o registo deve ser efetuado em nome do operador do estabelecimento responsável pelo cão ou gato. No caso de cães e gatos detidos por quaisquer outras pessoas singulares ou coletivas, o registo deve ser efetuado em nome dessas pessoas.\n\nOs Estados-Membros podem conceder derrogações ao primeiro parágrafo do presente número relativamente a cães militares, policiais e aduaneiros.\n\n4.\tEm caso de colocação no mercado ou cedência ocasional por pessoas singulares sem recorrer a publicidade em linha, a pessoa singular ou coletiva que transfere a propriedade ou a responsabilidade pelo cão ou gato deve assegurar que a mudança de propriedade ou de responsabilidade pelo cão ou gato é registada na base de dados referida no artigo 23.º, no prazo de duas semanas a contar da data dessa transferência, em conformidade com as condições estabelecidas pelo Estado-Membro responsável por essa base de dados.\n\n5.\tEm caso de morte de um cão ou gato, o operador, o dono do animal de companhia ou a pessoa singular ou coletiva proprietária do cão ou gato deve assegurar que a morte é registada na base de dados referida no artigo 23.º, em conformidade com as condições estabelecidas pelo Estado-Membro responsável por essa base de dados.\n\n6.\tCaso um transponder seja ou fique ilegível, o operador ou a pessoa singular ou coletiva responsável pelo cão ou gato deve assegurar que um novo transponder é injetado e que o registo na base de dados é atualizado com o número de identificação desse novo transponder.\n\n7.\tOs requisitos de identificação e registo do presente artigo são aplicáveis:\n\n(a)\taos operadores e às pessoas singulares ou coletivas que colocam cães e gatos no mercado, a partir de … [4 anos após a entrada em vigor do presente Regulamento];\n\n(b)\taos donos de animais de companhia e outras pessoas singulares ou coletivas que não sejam operadores e que não coloquem cães no mercado, a partir de … [10 anos após a entrada em vigor do presente Regulamento];\n\n(c)\taos donos de animais de companhia e outras pessoas singulares ou coletivas que não sejam operadores e que não coloquem gatos no mercado, a partir de … [15 anos após a entrada em vigor do presente Regulamento]."
        },
        "rgbeac": {
            "ref": "Art.º 17.º do RGBEAC (proposta, jun. 2025)",
            "texto": "1 — A identificação dos animais de companhia, pela sua marcação, quando aplicável, e registo no SIAC, deve ser realizada:\n\na) Relativamente aos cães, gatos e furões nascidos em alojamentos, até aos três meses de idade ou, em qualquer caso, antes da sua colocação no mercado;\n\nb) Relativamente aos cães, gatos e furões que entrem em alojamentos, nos termos dos artigos 12.º a 15.º, até trinta dias após a sua chegada ao alojamento ou, em qualquer caso, antes da data de colocação no mercado;\n\nc) Relativamente aos cães, gatos e furões detidos por pessoas singulares, exceto nos casos previstos nas alíneas anteriores, até aos três meses de idade ou, no caso de colocação no mercado, antes da data de colocação no mercado."
        },
        "codigo": {
            "ref": "Art.º 53.º do Código do Animal (DL n.º 214/2013)",
            "texto": "1 — Todos os cães devem ser identificados e registados, entre os três e os seis meses de idade.\n\n2 — Os gatos em exposição, para fins comerciais ou lucrativos, em estabelecimentos de venda, locais de criação, feiras ou concursos, provas funcionais, publicidade ou fins similares, devem ser identificados e registados entre os três e os seis meses de idade.\n\n4 — Os cães e gatos são identificados através de método electrónico e registados na base de dados nacional.\n\n5 — A identificação electrónica é efetuada através da aplicação subcutânea de um microchip no centro da face lateral esquerda do pescoço."
        },
        "legislacao": {
            "ref": "n.ºs 1, 2 e 3 do art.º 5.º do DL n.º 82/2019, de 27 de junho",
            "texto": "1 — A identificação dos animais de companhia, pela sua marcação e registo no SIAC, deve ser realizada até 120 dias após o seu nascimento.\n\n2 — Na impossibilidade de determinar a data de nascimento exata, para efeitos de contagem do prazo referido no número anterior, a identificação deve ser efetuada até à perda dos dentes incisivos de leite.\n\n3 — Sem prejuízo dos números anteriores, e relativamente aos cães, gatos e furões que sejam cedidos e ou comercializados a partir de um criador ou de um estabelecimento autorizado para a detenção de animais de companhia, nomeadamente os centros de hospedagem com ou sem fins lucrativos e os centros de recolha oficiais, deve ser assegurada a sua marcação e registo no SIAC antes de abandonarem a instalação de nascimento ou de alojamento, independentemente da sua idade."
        },
        "divergencia": {
            "legislacao": "O DL n.º 82/2019 (art.º 5.º) prevê prazo geral de 120 dias após o nascimento, sem distinguir o contexto de estabelecimento — prazo superior ao máximo de 3 meses do @regulamento, que exigirá redução. Não prevê o prazo específico de 30 dias para animais que entram em estabelecimentos.",
            "codigo": "O @codigo fixa prazo de identificação entre 3 e 6 meses, igualmente sem distinção entre nascimentos e entrada em estabelecimentos, ficando aquém da precisão exigida pelo @regulamento.",
            "rgbeac": "O @rgbeac aproxima-se dos prazos do @regulamento mas aplica-se apenas a cães, gatos e furões. Não prevê o prazo específico de 30 dias para animais que entram em estabelecimentos.",
            "sumario": "Necessidade de alteração: (1) reduzir prazo máximo de identificação de nascimentos para 3 meses; (2) criar prazo diferenciado de 30 dias para animais admitidos em estabelecimentos; (3) harmonizar prazos entre todos os diplomas nacionais."
        },
        "necessidade_alteracao": "Sim",
        "notas": ""
    },
    {
        "id": "ART-21",
        "tema": "Requisitos de Publicidade em Linha e Colocação no Mercado",
        "regulamento": {
            "ref": "Art.º 21.º do Regulamento 2023/0447",
            "titulo": "Requirements on online advertising and placing on the market",
            "texto": "1.	When operators advertise a dog or a cat online with a view to placing it on the Union market, they shall ensure that the following warning is included in the advertisement in clearly visible and bold characters:\n\n“An animal is not a toy. Getting one is a life-changing decision. It is your duty to ensure the animal’s health and welfare and not to abandon it.”.\n\n2.	When natural or legal persons other than operators advertise a dog or a cat online with a view to placing it on the Union market, they shall ensure that a warning on responsible ownership is included in the advertisement using either the wording set out in paragraph 1, or a different wording with the same meaning.\n\n3.	When placing a dog or a cat on the market in the Union, the natural or legal person placing the dog or cat on the market shall provide the acquirer with  :\n\n(a)	proof of  the identification and registration of the dog or cat in compliance with Article 20;\n\n(b)	the following information on the dog or cat:\n\n(i)	its species;\n\n(ii)	its sex;\n\n(iii)	its date and country of birth; and\n\n(iv)	where relevant, its breed.\n\nWhere a natural or legal person advertises a dog or cat online with a view to placing it on the Union market, that person shall use the system referred to in paragraph 5 to generate a unique verification token. That person shall include that token in the advertisement, along with a weblink to the system referred to in paragraph 5.\n\nThe system referred to in paragraph 5 shall enable acquirers to verify the authenticity of the identification, registration and ownership of dogs or cats advertised online.\n\n4.	Providers of online platforms shall ensure that their online interface is designed and organised in such a way that makes it easier for operators or other natural or legal persons who are placing dogs or cats on the market to comply with their obligations under paragraphs 1 to 3 of this Article, and in line with Article 31 of Regulation (EU) 2022/2065, and shall inform acquirers, in a visible manner, of the possibility to verify the authenticity of the identification, registration and ownership of the dog or cat  on the online verification system referred to in paragraph 5 accessed via a weblink.\n\nOnly the natural or legal person placing dogs or cats on the market shall be responsible for the accuracy of the information provided through the interface of the online platform. Nothing in this paragraph shall be construed as imposing a general monitoring obligation on the provider of the online platform within the meaning of Article 8 of Regulation (EU) 2022/2065.\n\n5.	The Commission shall ensure that a verification system for performing automated checks of the authenticity of the identification, registration and ownership of dogs or cats advertised online, using the database referred to in Article 23, is publicly available online, free of charge and generates the unique verification token referred to in paragraph 3(2) of this Article. The Commission may entrust the development, maintenance and operation of this system to an independent entity. That independent entity shall be chosen for that task following a public selection process, pursuant to the relevant provisions of Title VII of Regulation (EU, Euratom) 2024/2509 of the European Parliament and of the Council.The system shall ensure the following:\n\n(a)	reliable verification of the authenticity of the  identification, registration and ownership of the dog or cat using the national databases referred to in Article 23;\n\n(b)	compliance with data protection in accordance with Regulation (EU) 2018/1725 and Regulation (EU) 2016/679 of the European Parliament and of the Council;\n\n6.	The Commission shall adopt implementing acts laying down:\n\n(a)	the information to be provided by natural and legal persons placing dogs or cats on the market as proof of identification and registration of the dogs and cats in accordance with point (a) of paragraph 3;\n\n(b)	the information to be provided by natural and legal persons advertising dogs or cats to the verification system referred to in paragraph 5 for the purpose of demonstrating the authenticity of the identification, registration and ownership of the dog or cat advertised.\n\n(c)	the following characteristics of the system referred to in paragraph 5:\n\n-	the key functions of the system;\n\n-	the technical, electronic and cryptographic requirements for the system;\n\n-	the procedural steps to be followed, and the information to be provided,  by the natural or legal person placing the dog or cat on the market, and the steps and information required of the acquirer, in order for the online verification system to work.\n\nThe implementing acts referred to in point (a) shall be adopted by ... [two years after the date of entry into force of this Regulation] and the implementing act referred to in points (b) and (c) shall be adopted by ... [three years from date of entry into force of this Regulation]\n\nThose implementing acts shall be adopted in accordance with the examination procedure referred to in Article 29.",
            "traducao": "1.\tQuando operadores anunciem online um cão ou gato com vista ao seu colocamento no mercado da União, devem assegurar a apresentação do seguinte aviso no anúncio em caracteres claramente visíveis e em negrito:\n\n\t\"Um animal não é um brinquedo. Ter um é uma decisão que muda a vida. É seu dever assegurar a sua saúde e bem-estar e não o abandonar.\"\n\n2.\tQuando pessoas singulares ou coletivas que não sejam operadores anunciem online um cão ou gato com vista ao seu colocamento no mercado da União, devem assegurar a apresentação de um aviso sobre detenção responsável utilizando a redação referida no n.º 1 ou uma redação diferente com significado equivalente.\n\n3.\tAo colocar um cão ou gato no mercado na União, a pessoa singular ou coletiva que coloca o cão ou gato no mercado deve fornecer ao adquirente:\n\n(a)\tprova da identificação e do registo do cão ou gato em conformidade com o artigo 20.º;\n\n(b)\tas seguintes informações sobre o cão ou gato:\n\n(i)\ta sua espécie;\n\n(ii)\to seu sexo;\n\n(iii)\ta sua data e país de nascimento; e\n\n(iv)\tquando relevante, a sua raça.\n\n\tCaso uma pessoa singular ou coletiva anuncie um cão ou gato online com vista ao seu colocamento no mercado da União, essa pessoa deve utilizar o sistema referido no n.º 5 para gerar um token único de verificação e incluir esse token no anúncio, juntamente com uma ligação web para o sistema referido no n.º 5.\n\n\tO sistema referido no n.º 5 deve permitir aos adquirentes verificar a autenticidade da identificação, do registo e da propriedade de cães ou gatos anunciados online.\n\n4.\tOs fornecedores de plataformas online devem assegurar que a sua interface online é concebida e organizada de forma a facilitar aos operadores ou outras pessoas singulares ou coletivas que colocam cães ou gatos no mercado o cumprimento das suas obrigações nos termos dos n.ºs 1 a 3 do presente artigo, e em conformidade com o artigo 31.º do Regulamento (UE) 2022/2065, e devem informar os adquirentes, de forma visível, da possibilidade de verificar a autenticidade da identificação, do registo e da propriedade do cão ou gato no sistema de verificação online referido no n.º 5, acessível através de uma ligação web.\n\n\tApenas a pessoa singular ou coletiva que coloca cães ou gatos no mercado é responsável pela exatidão das informações fornecidas através da interface da plataforma online. Nenhuma disposição do presente número pode ser interpretada como impondo uma obrigação geral de monitorização ao fornecedor da plataforma online, na aceção do artigo 8.º do Regulamento (UE) 2022/2065.\n\n5.\tA Comissão garante que um sistema de verificação destinado a realizar controlos automatizados da autenticidade da identificação, do registo e da propriedade de cães ou gatos anunciados online, utilizando a base de dados referida no artigo 23.º, está publicamente disponível online, a título gratuito, e gera o token único de verificação referido no segundo parágrafo do n.º 3 do presente artigo. A Comissão pode confiar o desenvolvimento, a manutenção e o funcionamento deste sistema a uma entidade independente, escolhida para essa missão na sequência de um processo de seleção pública. O sistema deve assegurar:\n\n(a)\ta verificação fiável da autenticidade da identificação, do registo e da propriedade do cão ou gato, utilizando as bases de dados nacionais referidas no artigo 23.º;\n\n(b)\ta conformidade com a proteção de dados, em conformidade com o Regulamento (UE) 2018/1725 e o Regulamento (UE) 2016/679.\n\n6.\tA Comissão adota atos de execução que estabelecem:\n\n(a)\tas informações a fornecer pelas pessoas singulares e coletivas que colocam cães ou gatos no mercado como prova de identificação e registo dos cães e gatos em conformidade com o n.º 3, alínea a);\n\n(b)\tas informações a fornecer pelas pessoas singulares e coletivas que anunciam cães ou gatos ao sistema de verificação referido no n.º 5, para efeitos de demonstração da autenticidade da identificação, do registo e da propriedade do cão ou gato anunciado;\n\n(c)\tas seguintes características do sistema referido no n.º 5:\n\n-\tas funções essenciais do sistema;\n\n-\tos requisitos técnicos, eletrónicos e criptográficos do sistema;\n\n-\tos passos procedimentais a seguir, e as informações a fornecer, pela pessoa singular ou coletiva que coloca o cão ou gato no mercado, bem como os passos e informações exigidos ao adquirente, para que o sistema de verificação online funcione.\n\n\tOs atos de execução referidos na alínea a) devem ser adotados até … [dois anos após a data de entrada em vigor do presente Regulamento] e os atos de execução referidos nas alíneas b) e c) devem ser adotados até … [três anos a contar da data de entrada em vigor do presente Regulamento].\n\n\tEsses atos de execução são adotados em conformidade com o procedimento de exame referido no artigo 29.º."
        },
        "rgbeac": {
            "ref": "Art.º 95.º + Art.º 115.º n.º 3 do RGBEAC (proposta, jun. 2025)",
            "texto": "Artigo 95.º - Local de venda dos animais\n\n1 - Os animais de companhia não podem ser vendidos por entidade transportadora ou através da Internet, designadamente através de quaisquer portais ou plataformas.\n\n2 - Os animais de companhia não podem ser publicitados na Internet, mas a compra e venda dos mesmos apenas é admitida no local de criação ou em estabelecimentos devidamente licenciados.\n\nArtigo 115.º n.º 3 - É proibida a publicidade e comercialização de animais perigosos ou que demonstrem comportamento agressivo."
        },
        "codigo": {
            "ref": "Art.º 24.º e Art.º 62.º do Código do Animal (DL 214/2013)",
            "texto": "A legislação proíbe a venda de animais de companhia sem documentação adequada e restringe a comercialização de animais com comportamentos perigosos. As disposições cobrem identificação obrigatória e informações sobre o animal antes da venda."
        },
        "legislacao": {
            "ref": "DL 276/2001, DL 82/2019 - Normas de comercialização",
            "texto": "A legislação portuguesa estabelece controlos sobre a comercialização de cães e gatos, com requisitos de identificação e documentação sanitária. DL 82/2019 reforça as normas de rastreabilidade através do SIAC."
        },
        "divergencia": {
            "legislacao": "PARCIAL - Proíbe venda online mas sem sistema de verificação online específico",
            "codigo": "PARCIAL - Cobre restrições comerciais mas não implementa sistema de token de verificação",
            "rgbeac": "SIM - Proibição clara de publicidade e venda online, com requisitos de local licenciado",
            "sumario": "A legislação portuguesa PROÍBE a publicidade e venda de animais de companhia na Internet (Art. 95.º RGBEAC). Falta apenas a implementação do sistema de verificação online com token único conforme Art. 17a.6 do Regulamento. A proibição é mais restritiva que o Regulamento que permite venda online com verificação."
        },
        "necessidade_alteracao": "Sim - Sistema de verificação online com token único",
        "notas": "Legislação portuguesa mais restritiva - proíbe venda online; Reg. EU permite com verificação"
    },
    {
        "id": "ART-22",
        "tema": "Treino de Cuidadores de Animais",
        "regulamento": {
            "ref": "Art.º 22.º do Regulamento 2023/0447",
            "titulo": "Training",
            "texto": "1.	For the purposes of Article 12, the competent authorities shall be responsible for:\n\n(a)	ensuring that training courses are available for animal carers;\n\n(b)	approving the content of the training courses referred to in point (a), in accordance with the minimum requirements laid down by the implementing acts referred to in Article 12(4);\n\n(c)	certifying animal carers who have successfully completed the training courses referred to in point (a).\n\nThe competent authorities may delegate the task referred to in point (c) to providers of training courses.\n\n2.	A European Union Reference Centre for Animal Welfare designated in accordance with Article 95 of Regulation (EU) 2017/625 may develop models of training materials and recommendations for the competent authorities or other providers of training courses.",
            "traducao": "1. Para efeitos do artigo 12.º, as autoridades competentes são responsáveis por:\n\na) Assegurar que existem cursos de formação disponíveis para cuidadores de animais;\n\nb) Aprovar o conteúdo dos cursos de formação referidos na alínea a), tendo em conta os requisitos mínimos estabelecidos pelos atos de execução referidos no artigo 12.º, parágrafo 3;\n\nba) Certificar os cuidadores de animais que completaram com sucesso os cursos de formação referidos na alínea a).\n\n2. As autoridades competentes podem delegar a tarefa referida na alínea ba).\n\n3. O Centro de Referência da União Europeia para o Bem-Estar Animal designado em conformidade com o artigo 95.º do Regulamento (UE) 2017/625 pode desenvolver modelos de materiais de formação e recomendações para os fornecedores dos cursos de formação referidos no parágrafo 1."
        },
        "rgbeac": {
            "ref": "Arts. 118.º, 119.º, 120.º do RGBEAC (proposta, jun. 2025)",
            "texto": "Artigo 118.º - Reserva de atividade de treinadores de cães\nO treino de cães para qualquer fim só pode ser ministrado por treinador possuidor do respetivo título profissional.\n\nArtigo 119.º - Título profissional de treinador de cães\n1 - O acesso e exercício da atividade de treinador de cães depende da obtenção do respetivo título profissional, emitido pela DGAV.\n2 - O requerente deve: ter habilitação mínima 12.º ano; apresentar certificado criminal; ser detentor do certificado de qualificações.\n\nArtigo 120.º - Certificado de qualificações\n1 - Emitido por entidade certificadora após aprovação em provas teóricas e práticas, demonstrando habilitação técnica com base em métodos de reforço positivo.\n2 - Provas incidem sobre comportamento animal, metodologia de treino, aprendizagem e extinção de comportamentos."
        },
        "codigo": {
            "ref": "Arts. 51.º, 52.º do Código do Animal (DL 214/2013)",
            "texto": "Artigo 51.º - Exercício de profissões relacionadas com cães\nO exercício de certas profissões (adestrador, criador) requer documentação específica e conformidade com normas de bem-estar animal.\n\nArtigo 52.º - Documentação necessária\nRequerimentos de certificação e qualificação profissional para operadores."
        },
        "legislacao": {
            "ref": "DL 82/2019 - Normas de profissionalismo",
            "texto": "Legislação portuguesa estabelece normas de profissionalismo e qualificação para operadores em bem-estar animal, com requisitos de formação e certificação conforme disposições de Decreto-Lei específico."
        },
        "divergencia": {
            "legislacao": "SIM - Cobertura COMPLETA (certificação profissional, provas teóricas e práticas, métodos positivos)",
            "codigo": "SIM - Cobertura COMPLETA (profissionalismo, qualificações obrigatórias)",
            "rgbeac": "SIM - Cobertura COMPLETA (Arts. 118-120 implementam sistema profissional com certificação)",
            "sumario": "A legislação portuguesa implementa COMPLETAMENTE os requisitos do Artigo 18. O RGBEAC (Arts. 118-120) estabelece um sistema robusto de certificação profissional para treinadores de cães com: título profissional obrigatório, requisitos de habilitação, provas teóricas e práticas, métodos baseados em reforço positivo, verificação de antecedentes criminais. Sistema já em vigor conforme Decreto-Lei."
        },
        "necessidade_alteracao": "Não",
        "notas": "Sistema profissional português já implementado - cobertura completa e adequada"
    },
    {
        "id": "ART-23",
        "tema": "Base de Dados de Cães e Gatos",
        "regulamento": {
            "ref": "Art.º 23.º do Regulamento 2023/0447",
            "titulo": "Databases of dogs and cats",
            "texto": "1.	Member States shall be responsible for establishing and maintaining databases for the registration of identified dogs and cats in accordance with Article 20(1) and (2) and Article 26(3) and Article 26(4), second subparagraph.\n\n2.	For that purpose, the Member States may use databases maintained by another Member State, on the basis of appropriate arrangements between those Member States.\n\n3.	 Member States shall ensure that their databases, as referred to in paragraph 1, comply with the requirements laid down by the implementing act referred to in paragraph 4, second subparagraph, point (b), to ensure their interoperability.\n\n4.	The Commission shall establish and maintain an index database containing the minimum set of fields laid down in the implementing acts referred to in subparagraph 2, point (b). The Commission may entrust the development, maintenance and operation of that index database to an independent entity, following a public selection process pursuant to the relevant provisions of Title VII of the Regulation (EU, Euratom) 2024/2509.\n\nThe Commission shall adopt implementing acts laying down detailed arrangements concerning:\n\n(a)	the minimum content of the databases referred to in paragraph 1;\n\n(b)	the interoperability between Member States’ databases and the index database, including the minimum set of fields to be transmitted to the index database and the intervals of the transmission;\n\n(c)	the functionality for providing proof of the identification and registration of a dog or a cat, as referred to in Article 21 (3) point (a).\n\n(d)	the registry where Member States will declare their databases, and the necessary parameters for connecting those databases with one another in accordance with the provisions established pursuant to point (b);\n\n(e)	the interconnection between the Member States’ databases referred to in paragraph 1, the pet travellers' database referred to in Article 26, paragraph 4, and the Information Management System for Official Controls (IMSOC), where relevant.\n\nThe Commission shall adopt the implementing acts referred to in the second subparagraph, points (a) and (c) by ... [date two years after the date of entry into force of this Regulation]. It shall adopt the implementing acts referred to in the second subparagraph, points (b), (d) and (e) by ...[three years from the date of entry into force of this Regulation].\n\nThose implementing acts shall be adopted in accordance with the examination procedure referred to in Article 29.",
            "traducao": "1.\tOs Estados-Membros são responsáveis pela criação e manutenção de bases de dados para o registo de cães e gatos identificados, em conformidade com o artigo 20.º, n.ºs 1 e 2, e com o artigo 26.º, n.ºs 3 e 4, segundo parágrafo.\n\n2.\tPara este efeito, os Estados-Membros podem utilizar bases de dados mantidas por outro Estado-Membro, com base em acordos apropriados entre esses Estados-Membros.\n\n3.\tOs Estados-Membros asseguram que as suas bases de dados referidas no n.º 1 estão em conformidade com os requisitos estabelecidos pelo ato de execução referido no n.º 4, segundo parágrafo, alínea b), de modo a assegurar a sua interoperabilidade.\n\n4.\tA Comissão estabelece e mantém uma base de dados de índice contendo o conjunto mínimo de campos estabelecido nos atos de execução referidos no segundo parágrafo, alínea b). A Comissão pode confiar o desenvolvimento, a manutenção e o funcionamento dessa base de dados de índice a uma entidade independente, na sequência de um processo de seleção pública.\n\n\tA Comissão adota atos de execução que estabelecem regras pormenorizadas relativas a:\n\n(a)\to conteúdo mínimo das bases de dados referidas no n.º 1;\n\n(b)\ta interoperabilidade entre as bases de dados dos Estados-Membros e a base de dados de índice, incluindo o conjunto mínimo de campos a transmitir à base de dados de índice e os intervalos da transmissão;\n\n(c)\ta funcionalidade para fornecer prova de identificação e registo de um cão ou gato, tal como referido no artigo 21.º, n.º 3, alínea a);\n\n(d)\to registo onde os Estados-Membros declaram as suas bases de dados, e os parâmetros necessários para a conexão dessas bases de dados entre si, em conformidade com as disposições estabelecidas em aplicação da alínea b);\n\n(e)\ta interconexão entre as bases de dados dos Estados-Membros referidas no n.º 1, a base de dados de viajantes da União com animais de estimação referida no artigo 26.º, n.º 4, e o Sistema de Gestão da Informação para os Controlos Oficiais (IMSOC), quando relevante.\n\n\tA Comissão adota os atos de execução referidos no segundo parágrafo, alíneas a) e c), até … [data dois anos após a data de entrada em vigor do presente Regulamento]. Adota os atos de execução referidos no segundo parágrafo, alíneas b), d) e e), até … [três anos a contar da data de entrada em vigor do presente Regulamento].\n\n\tEsses atos de execução são adotados em conformidade com o procedimento de exame referido no artigo 29.º."
        },
        "rgbeac": {
            "ref": "Art.º 20.º do RGBEAC (proposta, jun. 2025)",
            "texto": "Artigo 20.º - Sistema de Informação de Animais de Companhia (SIAC)\n\n1 - O SIAC reúne a informação relativa à rastreabilidade dos dispositivos de identificação, à identificação dos animais de companhia, à sua titularidade ou detenção e à informação relacionada com a defesa da saúde pública, saúde animal e bem-estar animal.\n\n2 – A DGAV é a entidade responsável pelo SIAC, competindo-lhe assegurar o seu funcionamento e o tratamento seguro da informação.\n\n3 – A DGAV pode atribuir a gestão do SIAC a outras entidades, mediante a celebração de protocolo e parecer prévio da Comissão Nacional de Proteção de Dados.\n\n4 - As normas e procedimentos relativos ao funcionamento do SIAC constam de um Manual de Procedimentos SIAC, aprovado pelo diretor-geral de alimentação e veterinária."
        },
        "codigo": {
            "ref": "Arts. 53.º, 55.º, 56.º, 57.º do Código do Animal (DL 214/2013)",
            "texto": "Artigo 53.º - Identificação e registo na base de dados\n1 - Todos os cães devem ser identificados e registados entre os três e seis meses de idade.\n4 - Os cães e gatos são identificados através de método eletrónico e registados na base de dados nacional.\n5 - Identificação através de aplicação subcutânea de microchip no centro da face lateral esquerda do pescoço.\n\nArtigo 55.º - Base de dados\n1 - Toda a informação do registo coligida numa aplicação informática nacional.\n2 - A DGAV detém e coordena o acesso à base de dados, podendo autorizar gestão em outras entidades com parecer da Comissão Nacional de Proteção de Dados.\n\nArtigo 56.º - Classificação dos animais (Cão, Cão Potencialmente Perigoso, Cão Perigoso, Gato)"
        },
        "legislacao": {
            "ref": "DL n.º 82/2019, Art.º 2.º (Estabelece SIAC)",
            "texto": "DL 82/2019 estabelece o SIAC com extensas disposições cobrindo: criação da base de dados, procedimentos de registo e controlos de acesso, gestão por DGAV, identificação eletrónica via transponder, segurança de dados e autorização de acesso, integração de registos de identificação de animais, gestão da distribuição de dispositivos de identificação, procedimentos de registo e atualização. O sistema é totalmente operacional desde 2019."
        },
        "divergencia": {
            "legislacao": "NÃO - SIAC em funcionamento desde 2019 cumpre todos os requisitos",
            "codigo": "NÃO - Código do Animal estabelece estrutura completa com classificações",
            "rgbeac": "NÃO - RGBEAC (Art. 20.º) reafirma e fortalece o sistema SIAC",
            "sumario": "A legislação portuguesa CUMPRE COMPLETAMENTE os requisitos do Artigo 19. O SIAC (Sistema de Informação de Animais de Companhia) já está operacional desde 2019, sob responsabilidade de DGAV, com: identificação eletrónica via microchip obrigatória, registo nacional centralizado, rastreabilidade garantida, classificação de animais (normal/perigoso/potencialmente perigoso), integração de dados de saúde e bem-estar, conformidade com proteção de dados (Comissão Nacional de Proteção de Dados)."
        },
        "necessidade_alteracao": "Não",
        "notas": "SIAC totalmente operacional e conforme - Interoperabilidade EU pendente de implementação"
    },
    {
        "id": "ART-24",
        "tema": "Recolha de Dados sobre Bem-Estar Animal e Relatório",
        "regulamento": {
            "ref": "Art.º 24.º do Regulamento 2023/0447",
            "titulo": "Collection of data on animal welfare and reporting",
            "texto": "1.	The competent authorities shall collect, analyse and publish the data on animal welfare set out in Annex III:\n\n2.	The competent authorities shall draw up and transmit to the Commission a report in electronic form, on the data on animal welfare, set out in Annex III, by 31 August at three-yearly intervals. The first such report shall be drawn up and transmitted to the Commission by ... [date 6 years from the date of entry into force of this Regulation]. Each report shall contain a summary of the data gathered during the previous three years.\n\n3.	The Commission may adopt implementing acts, establishing a harmonised methodology for collecting the data on animal welfare set out in Annex III and establishing a template for the report referred to in paragraph 2 of this Article. Those implementing acts shall be adopted in accordance with the examination procedure referred to in Article 29.",
            "traducao": "1.\tAs autoridades competentes devem recolher, analisar e publicar os dados sobre bem-estar animal estabelecidos no Anexo III.\n\n2.\tAs autoridades competentes devem elaborar e transmitir à Comissão um relatório em formato eletrónico, sobre os dados relativos ao bem-estar animal estabelecidos no Anexo III, até 31 de agosto, com periodicidade trienal. O primeiro relatório deve ser elaborado e transmitido à Comissão até ... [data 6 anos após a data de entrada em vigor do presente regulamento]. Cada relatório deve conter um resumo dos dados recolhidos durante os três anos anteriores.\n\n3.\tA Comissão pode adotar atos de execução que estabeleçam uma metodologia harmonizada para a recolha dos dados sobre bem-estar animal estabelecidos no Anexo III e que definam um modelo para o relatório referido no n.º 2 do presente artigo. Esses atos de execução são adotados em conformidade com o procedimento de exame referido no artigo 29.º."
        },
        "rgbeac": {
            "ref": "Art.º 46.º do RGBEAC (proposta, jun. 2025)",
            "texto": "Artigo 46.º - Relatório Nacional Anual\n\n1 - Os centros de bem-estar animal e alojamentos remetem à DGAV no primeiro mês de cada ano civil, os relatórios de gestão do ano anterior, com números de animais recolhidos, restituídos, eutanasiados, cedidos, adotados, devolvidos após adoção, vacinados, esterilizados e intervencionados em programas CED.\n\n2 – A DGVA consolida a informação a nível nacional sobre bem-estar animal em cada ano, incluindo: acompanhamento da política internacional, prestação de apoios públicos, atividade do SIAC, resultados de planos de controlo, processos contraordenacionais, coimas aplicadas, atividade de centros de bem-estar, programas de interesse nacional, ações informativas e educativas, para efeitos de elaboração do Relatório Anual sobre a situação do Bem-Estar Animal."
        },
        "codigo": {
            "ref": "Arts. 141.º-145.º do Código do Animal (DL 214/2013)",
            "texto": "Código do Animal estabelece estrutura de monitorização e relatórios sobre bem-estar animal, incluindo: recolha de dados de infrações, aplicação de sanções, atividades de centros de recolha, estatísticas de animais processados. Sistema de contraordenações com documentação e coimas registadas."
        },
        "legislacao": {
            "ref": "Decreto-Regulamentar n.º 3/2021 + DL 82/2019",
            "texto": "Decreto-Regulamentar 3/2021 (Art. 3.º) requer Relatório Anual sobre situação do Bem-Estar Animal. DL 82/2019 estabelece procedimentos de recolha de dados, monitorização, registo de atividades médico-veterinárias e compilação de informação de bem-estar animal para fins de relatório anual e transmissão a autoridades europeias."
        },
        "divergencia": {
            "legislacao": "SIM - Cobertura COMPLETA (relatórios anuais, recolha de dados, consolidação nacional)",
            "codigo": "SIM - Cobertura COMPLETA (monitorização, infrações, sanções, estatísticas)",
            "rgbeac": "SIM - Cobertura COMPLETA (Art. 46.º estabelece sistema de relatórios nacionais)",
            "sumario": "A legislação portuguesa implementa COMPLETAMENTE o Artigo 20. Sistema já estabelecido com: recolha anual de dados de centros de bem-estar animal, consolidação de informação de bem-estar a nível nacional, inclusão de estatísticas de SIAC, registos de contraordenações e sanções, relatórios sobre atividades de proteção animal, conformidade com requisitos de Decreto-Regulamentar 3/2021. Único ajuste necessário: alinhamento da periodicidade de relatórios (atualmente anual; Regulamento requer trienal) e inclusão dos dados específicos de Annex III da EU (quando publicado)."
        },
        "necessidade_alteracao": "Sim - Ajuste de periodicidade e conformidade com Annex III EU",
        "notas": "Sistema de relatórios já operacional - apenas alinhamento com padrões EU necessário"
    },
    {
        "id": "ART-25",
        "tema": "Proteção de Dados",
        "regulamento": {
            "ref": "Art.º 25.º do Regulamento 2023/0447",
            "titulo": "Data protection",
            "texto": "1.	The competent authorities of the Member States shall be controllers within the meaning of Regulation (EU) 2016/679 in relation to the processing of personal data collected under Articles 9 and 10 of this Regulation, as well as under Article 23(1) of this Regulation when used for the purposes of official control.\n\nThe Commission shall be a controller within the meaning of Regulation (EU) 2018/1725 in relation to the processing of personal data collected under Article 21(5), Article 23(4), and Article 26(4), third subparagraph, , as well as under Article 23(1) of this Regulation when that data is used for the purposes of compliance with Article 108 of Regulation (EU) 2017/625 and the reporting obligations under this Regulation.\n\nIt shall be prohibited for any person with access to the personal data referred to in the first and second subparagraphs to divulge any personal data the knowledge of which was acquired in the exercise of his or her duties or otherwise incidentally to such exercise. Member States and the Commission shall take all appropriate measures to enforce that prohibition.\n\nThe personal data collected under the first and second subparagraphs shall not be used for purposes other than:\n\n(a)	official controls by Member States’ competent authorities of compliance with the welfare and traceability requirements of this Regulation and of compliance with Regulation (EU) 2016/429, including the detection of fraudulent practices; and\n\n(b)	compliance by the Commission with its obligations under Article 108 of Regulation (EU) 2017/625 and with the Commission’s reporting obligations under this Regulation.\n\n2.	The personal data referred to in paragraph 1 of this Article shall be retained for the following periods:\n\n(a)	in the case of Articles 9 and 10, 10 years after the date of cessation of the activity of the establishment;\n\n(b)	in the case of Article 21(5), 18 months after the generation of the token referred to in Article 21(3), second subparagraph;\n\n(c)	in the case of Articles 23(1) and 23(4), 25 years after the first registration of the dog or cat in the database referred to in that Article or five years after the recording of the death of the dog or cat in that database;\n\n(d)	in the case of Article 26(4), third subparagraph, 5 years after the date of pre-notification.\n\nCHAPTER V\nENTRY OF DOGS AND CATS INTO THE UNION",
            "traducao": "1.\tAs autoridades competentes dos Estados-Membros devem ser controladoras na aceção do Regulamento (UE) 2016/679 no que respeita ao tratamento de dados pessoais recolhidos nos termos dos artigos 9.º e 10.º do presente Regulamento, bem como no âmbito do artigo 23.º, n.º 1, do presente Regulamento quando utilizados para efeitos de controlo oficial.\n\n\tA Comissão deve ser controladora na aceção do Regulamento (UE) 2018/1725 no que respeita ao tratamento de dados pessoais recolhidos nos termos do artigo 21.º, n.º 5, do artigo 23.º, n.º 4, e do terceiro parágrafo do artigo 26.º, n.º 4, do presente Regulamento, bem como no âmbito do artigo 23.º, n.º 1, do presente Regulamento quando esses dados são utilizados para efeitos de conformidade com o artigo 108.º do Regulamento (UE) 2017/625 e de obrigações de comunicação de informações previstas no presente Regulamento.\n\n\tÉ proibido a qualquer pessoa com acesso aos dados pessoais referidos nos primeiro e segundo parágrafos divulgar quaisquer dados pessoais cujo conhecimento tenha sido adquirido no exercício das suas funções ou a título acessório do exercício dessas funções. Os Estados-Membros e a Comissão devem adotar todas as medidas adequadas para fazer respeitar essa proibição.\n\n\tOs dados pessoais recolhidos nos termos do primeiro e segundo parágrafos não podem ser utilizados para outros fins que não:\n\n(a)\tcontrolos oficiais pelas autoridades competentes dos Estados-Membros da conformidade com os requisitos de bem-estar e de rastreabilidade do presente Regulamento e da conformidade com o Regulamento (UE) 2016/429, incluindo a deteção de práticas fraudulentas; e\n\n(b)\to cumprimento pela Comissão das suas obrigações nos termos do artigo 108.º do Regulamento (UE) 2017/625 e das obrigações de comunicação de informações da Comissão previstas no presente Regulamento.\n\n2.\tOs dados pessoais referidos no n.º 1 do presente artigo são conservados pelos seguintes períodos:\n\n(a)\tno caso dos artigos 9.º e 10.º, 10 anos após a data de cessação da atividade do estabelecimento;\n\n(b)\tno caso do artigo 21.º, n.º 5, 18 meses após a geração do token referido no artigo 21.º, n.º 3, segundo parágrafo;\n\n(c)\tno caso dos artigos 23.º, n.ºs 1 e 4, 25 anos após o primeiro registo do cão ou gato na base de dados referida nesse artigo ou cinco anos após o registo da morte do cão ou gato nessa base de dados;\n\n(d)\tno caso do artigo 26.º, n.º 4, terceiro parágrafo, 5 anos após a data da pré-notificação."
        },
        "rgbeac": {
            "ref": "Art.º 20.º, nn. 3, 6, 8 do RGBEAC (proposta, jun. 2025)",
            "texto": "Artigo 20.º, n.º 8 - SIAC e Proteção de Dados:\nAo tratamento, segurança, conservação, acesso e proteção dos dados pessoais constantes do SIAC é diretamente aplicável o disposto na legislação e regulamentação relativa à proteção de dados pessoais, nomeadamente o Regulamento (UE) 2016/679 do Parlamento Europeu e do Conselho, de 27 de abril de 2016, relativo à proteção das pessoas singulares no que diz respeito ao tratamento de dados pessoais e à livre circulação desses dados.\n\nArtigo 20.º, n.º 6 - Transmissão de Dados Pessoais:\nSempre que se mostre necessário à operacionalização do SIAC ou ao cumprimento das suas finalidades, deve promover-se a transmissão de dados entre sistemas de informação, preferencialmente através da Plataforma de Interoperabilidade da Administração Pública (iAP), nos termos do Decreto-Lei n.º 135/99, de 22 de abril.\n\nArtigo 20.º, n.º 3 - Subcontratação de Dados Pessoais:\nA DGAV, pode atribuir a gestão do SIAC a outras entidades, mediante a celebração de protocolo e sob sua supervisão, observado o regime de subcontratação de tratamento de dados pessoais."
        },
        "codigo": {
            "ref": "Art.º 55.º do Código do Animal (DL 214/2013)",
            "texto": "Artigo 55.º - Base de Dados\n\n1 - Toda a informação resultante do registo do animal é coligida numa aplicação informática nacional.\n\n2 - A DGAV detém, define e coordena o acesso à base de dados, podendo autorizar a sua gestão noutras entidades, mediante a celebração de protocolos, precedidos de parecer da Comissão Nacional de Proteção de Dados.\n\n3 - Só têm acesso à base de dados as entidades que se encontrem autorizadas, para o efeito, pela DGAV."
        },
        "legislacao": {
            "ref": "DL n.º 82/2019, Art.º 9.º",
            "texto": "Artigo 9.º - Registo no Sistema de Informação de Animais de Companhia\n\n1 - Os animais de companhia abrangidos pela obrigação de identificação devem ser registados pelo médico veterinário no SIAC, imediatamente após a sua marcação com o transponder, em nome do respetivo titular."
        },
        "divergencia": {
            "legislacao": "NÃO - Legislação portuguesa cobre completamente com incorporação de GDPR",
            "codigo": "NÃO - Cobertura expressa em Art. 55.º",
            "rgbeac": "NÃO - Implementação completa em Arts. 20.º",
            "sumario": "A legislação portuguesa implementa completamente os requisitos de proteção de dados do Artigo 20a. O SIAC é configurado como sistema compliant com GDPR (EU 2016/679) e com regulamentações de proteção de dados. Não há necessidade de alterações legislativas."
        },
        "necessidade_alteracao": "Não",
        "notas": "Correspondências confirmadas por agente de pesquisa em 2026-03-02"
    },
    {
        "id": "ART-26",
        "tema": "Entrada de Cães e Gatos na União",
        "regulamento": {
            "ref": "Art.º 26.º do Regulamento 2023/0447",
            "titulo": "Entry of dogs and cats into the Union",
            "texto": "1.	Dogs and cats may be brought into the Union for the purpose of being placed on the Union market  only if the following conditions are met:\n\n(a)	they have been bred and kept in compliance with any of the following requirements:\n\n(i)	the requirements contained in Chapter II of this Regulation;\n\n(ii)	requirements recognised by the Union, in accordance with Article 129 of Regulation (EU) 2017/625, as being equivalent to those set out by Chapter II of this Regulation; or\n\n(iii)	where applicable, requirements contained in a specific agreement between the Union and the exporting country.\n\nb)	 they come from a third country or territory and from an establishment listed in accordance with Articles 126 and 127 of Regulation (EU) 2017/625.\n\n2.	The official certificate referred to in Article 126(2), point (c), of Regulation (EU) 2017/625 accompanying dogs and cats brought into the Union from third countries and territories for the purpose of being placed on the Union market, shall contain an attestation certifying compliance with  paragraph 1 of this Article.\n\n3.	Dogs and cats brought into the Union for the purpose of being placed on the Union market shall be identified before their entry into the Union by a veterinarian by means of an injectable transponder containing a readable microchip that complies with the requirements set out in  Annex II.\n\nThe operator responsible for the import of the dogs or cats into the Union shall ensure that they are registered in a national database referred to in Article 23(1), by a veterinarian, within five working days after they were brought into the Union. Member States may allow registration by persons other than veterinarians, provided that they have measures in place to ensure the accuracy of information that those persons enter in the database.\n\n4.	The non-commercial movement of a dog or cat from a third country or territory to the Union shall be prenotified by its owner to an online Union pet travellers’ database at least five working days before the dog or cat crosses the Union border, except in the following cases:\n\n(a)	where the dog or cat is brought into the Union directly from third countries or from territories fulfilling the conditions set out in Article 17(1), point (a), of Commission Delegated Regulation (EU) …/..., and\n\n(b)	where the dog or cat is registered in a Member State database referred to in Article 23(1) of this Regulation.\n\nWhere the dog or cat stays more than six months in the Union, the owner shall ensure that it is registered in the database of the Member State of residence referred to in Article 23(1), by a veterinarian, within five working days after the expiry of the sixth month since it entered the Union. Member States may allow registration by persons other than veterinarians, provided that they have measures in place to ensure the accuracy of information that those persons enter in the database.\n\nThe Commission shall establish and maintain the Union pet travellers’ database referred to in the first subparagraph. The Commission may entrust the development, maintenance and operation of that database to an independent entity, following a public selection process pursuant to the relevant provisions of Title VII of the Regulation (EU, Euratom) 2024/2509. Access to that database shall be restricted to Member States’ competent authorities and to the Commission.\n\nThe Commission shall ensure that the database triggers iRASFF notifications for pre-notified movements that present a risk of fraud. The Member State receiving the notification shall take appropriate measures to follow it up in accordance with Article 105(2) of Regulation (EU) 2017/625.\n\nThe Commission shall by ... [ 8 years after the date of entry into force of this Regulation] adopt implementing acts laying down detailed arrangements for the following:\n\n(i)	the information to be pre-notified by owners in accordance with paragraph (4) of this Article in the Union pet travellers’ database, taking into account the personal data protection requirements of Regulation (EU) 2018/1725 and Regulation (EU) 2016/679;\n\n(ii)	the procedure by which the risk of fraud is to be established, which is to take into account the activities carried out by the AAC network\n\nThose implementing acts shall be adopted in accordance with the examination procedure referred to in Article 29.",
            "traducao": "1.\tOs cães e os gatos podem ser introduzidos na União para efeitos de colocação no mercado da União apenas se estiverem cumpridas as seguintes condições:\n\n(a)\ttenham sido criados e mantidos em conformidade com qualquer das seguintes situações:\n\n(i)\tos requisitos do Capítulo II do presente Regulamento;\n\n(ii)\trequisitos reconhecidos pela União, em conformidade com o artigo 129.º do Regulamento (UE) 2017/625, como equivalentes aos estabelecidos no Capítulo II do presente Regulamento; ou\n\n(iii)\tquando aplicável, requisitos constantes de um acordo específico entre a União e o país exportador.\n\n(b)\tprovenham de um país terceiro ou território e de um estabelecimento incluído em lista em conformidade com os artigos 126.º e 127.º do Regulamento (UE) 2017/625.\n\n2.\tO certificado oficial referido no artigo 126.º, n.º 2, alínea c), do Regulamento (UE) 2017/625 que acompanha os cães e gatos introduzidos na União desde países terceiros e territórios para efeitos de colocação no mercado da União deve conter uma atestação certificando a conformidade com o n.º 1 do presente artigo.\n\n3.\tOs cães e gatos introduzidos na União para efeitos de colocação no mercado da União devem ser identificados antes da sua entrada na União por um médico veterinário por meio de um transponder injetável contendo um microchip legível em conformidade com os requisitos estabelecidos no Anexo II.\n\n\tO operador responsável pela importação dos cães ou gatos para a União deve assegurar que estes são registados por um médico veterinário numa base de dados nacional referida no artigo 23.º, n.º 1, no prazo de cinco dias úteis após a sua introdução na União. Os Estados-Membros podem permitir o registo por pessoas que não sejam médicos veterinários, desde que disponham de medidas para assegurar a exatidão das informações que essas pessoas introduzem na base de dados.\n\n4.\tO movimento não comercial de um cão ou gato desde um país terceiro ou território para a União deve ser pré-notificado pelo seu proprietário a uma base de dados online de viajantes da União com animais de estimação, pelo menos cinco dias úteis antes de o cão ou gato atravessar a fronteira da União, exceto nos seguintes casos:\n\n(a)\tquando o cão ou gato é introduzido na União diretamente desde países terceiros ou territórios que satisfazem as condições estabelecidas no artigo 17.º, n.º 1, alínea a), do Regulamento Delegado (UE) …/… da Comissão; e\n\n(b)\tquando o cão ou gato está registado numa base de dados de um Estado-Membro referida no artigo 23.º, n.º 1, do presente Regulamento.\n\n\tCaso o cão ou gato permaneça mais de seis meses na União, o proprietário deve assegurar o seu registo por um médico veterinário na base de dados do Estado-Membro de residência referida no artigo 23.º, n.º 1, no prazo de cinco dias úteis após o termo do sexto mês desde a sua entrada na União. Os Estados-Membros podem permitir o registo por pessoas que não sejam médicos veterinários, desde que disponham de medidas para assegurar a exatidão das informações que essas pessoas introduzem na base de dados.\n\n\tA Comissão estabelece e mantém a base de dados de viajantes da União com animais de estimação referida no primeiro parágrafo. A Comissão pode confiar o desenvolvimento, a manutenção e o funcionamento dessa base de dados a uma entidade independente, na sequência de um processo de seleção pública. O acesso a essa base de dados fica restrito às autoridades competentes dos Estados-Membros e à Comissão.\n\n\tA Comissão assegura que a base de dados aciona notificações iRASFF para os movimentos pré-notificados que apresentem um risco de fraude. O Estado-Membro que recebe a notificação toma as medidas adequadas para dar seguimento à notificação, em conformidade com o artigo 105.º, n.º 2, do Regulamento (UE) 2017/625.\n\n\tA Comissão adota, até … [8 anos após a data de entrada em vigor do presente Regulamento], atos de execução que estabelecem regras pormenorizadas relativas a:\n\n(i)\tas informações a pré-notificar pelos proprietários em conformidade com o n.º 4 do presente artigo na base de dados de viajantes da União com animais de estimação, tendo em conta os requisitos de proteção de dados pessoais do Regulamento (UE) 2018/1725 e do Regulamento (UE) 2016/679;\n\n(ii)\to procedimento pelo qual o risco de fraude deve ser determinado, tendo em conta as atividades realizadas pela rede CAA.\n\n\tEsses atos de execução são adotados em conformidade com o procedimento de exame referido no artigo 29.º."
        },
        "rgbeac": {
            "ref": "Art.º 114.º e Art.º 96.º do RGBEAC (proposta, jun. 2025)",
            "texto": "Artigo 114.º - Entrada no Território Nacional\n1 - A entrada no território nacional, por compra, cedência ou troca direta, de cães classificados como potencialmente perigosos pode ser condicionada.\n\n2 - Os cães referidos no número anterior que não estejam inscritos em livro de origens oficialmente reconhecido, que permaneçam em território nacional por mais de quatro meses, são obrigatoriamente esterilizados nos termos do disposto no n.º 5.º do artigo 102.º.\n\n3 - A introdução no território nacional por compra, cedência ou troca direta, tendo em vista a sua reprodução, de cães potencialmente perigosos é sujeita a autorização da DGAV requerida com sete dias de antecedência, decorridos os quais a mesma é tacitamente deferida.\n\nArtigo 96.º - Importação de Animais de Companhia\nA importação de animais que constem da lista nacional de animais de companhia provenientes de outros Estados é admitida desde que sejam cumpridas as regras sanitárias e de bem-estar animal portuguesas e comunitárias."
        },
        "codigo": {
            "ref": "Art.º 62.º do Código do Animal (DL 214/2013)",
            "texto": "Artigo 62.º - Entrada de Animais de Companhia Suscetíveis à Raiva em Território Nacional\n\n1 - A entrada em território nacional de animais de companhia suscetíveis à raiva destinados ao comércio, provenientes de outros Estados-membros ou de países terceiros, depende do cumprimento das condições fixadas no Decreto-Lei n.º 79/2001, de 20 de junho, alterado pelo Decreto-Lei n.º 260/2012, de 12 de dezembro, e noutras normas de polícia sanitária que regem o comércio e as importações de animais vivos na comunidade.\n\n2 - No caso das importações, deve ainda ser cumprido o regime dos controlos veterinários previsto no Decreto-Lei n.º 79/2011, de 20 de junho.\n\n3 - A entrada em território nacional de furões destinados ao comércio ou sem carácter comercial, para além do cumprimento do disposto nos números anteriores, depende de autorização prévia do ICNF, I.P."
        },
        "legislacao": {
            "ref": "DL n.º 82/2019, Art.º 2.º",
            "texto": "Artigo 2.º - Âmbito de Aplicação\n\nO presente decreto-lei aplica-se à identificação de animais de companhia das espécies referidas no anexo I do Regulamento (UE) n.º 576/2013, do Parlamento Europeu e do Conselho, de 12 de junho de 2013, e no anexo I do Regulamento (UE) n.º 2016/429, do Parlamento Europeu e do Conselho, de 9 de março de 2016, nascidos ou presentes no território nacional."
        },
        "divergencia": {
            "legislacao": "PARCIAL - Cobre entrada/importação mas sem database online de viajantes pré-notificação",
            "codigo": "SIM - Implementa controlos sanitários e requisitos de identificação",
            "rgbeac": "PARCIAL - Cobre entrada no território nacional com condições, não cobre movimento não-comercial pré-notificação",
            "sumario": "A legislação portuguesa implementa os requisitos essenciais de entrada e identificação pré-entrada, referenciando os Regulamentos EU 576/2013 e 2016/429. Faltam: (1) Sistema de base de dados online de viajantes com pré-notificação obrigatória para movimento não-comercial; (2) Sistema iRASFF integrado para notificações de risco de fraude."
        },
        "necessidade_alteracao": "Sim",
        "notas": "Parcialmente coberto. Complementação necessária para movimento não-comercial e sistema iRASFF."
    },
    {
        "id": "ART-27",
        "tema": "Alteração dos Anexos",
        "regulamento": {
            "ref": "Art.º 27.º do Regulamento 2023/0447",
            "titulo": "Amendment to the Annexes",
            "texto": "The Commission is empowered to adopt delegated acts in accordance with Article 28 to amend the Annexes to this Regulation to take account of scientific and technical progress, including, where relevant, the scientific opinions of EFSA, as regards:\n\n(a)	a suitable number of animal carers in breeding and selling establishments;\n\n(b)	watering and feeding requirements and weaning process;\n\n(c)	temperature ranges;\n\n(d)	lighting requirements;\n\n(e)	ammonia and carbon monoxide levels;\n\n(f)	kennel and cattery designs;\n\n(g)	group housing;\n\n(h)	space allowances for various categories of dogs and cats;\n\n(i)	the frequency of pregnancies;\n\n(j)	the minimum and maximum age of bitches and queens for  breeding;\n\n(k)	socialisation, the provision of enrichments and other measures for meeting the behavioural needs of dogs and cats;\n\n(l)	the requirements for the transponders used to identify dogs and cats individually;\n\n(m)	the data to be collected for policy monitoring and evaluation.\n\nAny additions of requirements in the Annexes shall be based on updated scientific or technical evidence, in particular such evidence regarding the specific conditions needed to ensure the welfare of the dogs and cats covered by the scope of this Regulation. Where relevant, those delegated acts shall take into account social, economic and environmental impacts and shall provide for sufficient transition periods to allow the operators concerned to adapt to the new requirements.",
            "traducao": "1. A Comissão fica habilitada a adotar atos delegados em conformidade com o artigo 28.º que alterem os anexos do presente Regulamento para ter em conta o progresso científico e técnico, incluindo, quando relevante, pareceres científicos da Autoridade Europeia para a Segurança dos Alimentos, nos seguintes aspetos:\n\na) Um número adequado de cuidadores de animais em estabelecimentos de criação e venda;\n\nb) Requisitos de abastecimento de água e alimentação e processo de desmame;\n\nc) Gamas de temperatura;\n\nd) Requisitos de iluminação;\n\ne) Níveis de amoníaco e monóxido de carbono;\n\nf) Conceção de canis e gateis;\n\ng) Alojamento em grupo;\n\nh) Espaços permitidos para diferentes categorias de cães e gatos;\n\ni) Frequência de gestações;\n\nj) Idade mínima e máxima de cadelas e gatas para reprodução;\n\nk) Socialização, enriquecimento e outras medidas para satisfazer as necessidades comportamentais de cães e gatos;\n\nl) Requisitos para transponders utilizados na identificação individual de cães e gatos;\n\nm) Dados a recolher para avaliação e monitorização de políticas.\n\n2. Qualquer complemento de requisitos nos Anexos deve ser baseado em evidência científica ou técnica atualizada, em particular no que diz respeito às condições específicas necessárias para assegurar o bem-estar dos cães e gatos abrangidos pelo âmbito do presente Regulamento. Quando relevante, esses atos delegados devem levar em conta impactos sociais, económicos e ambientais e prever períodos de transição suficientes para permitir aos operadores envolvidos a adaptação aos novos requisitos."
        },
        "rgbeac": {
            "ref": "Arts. 3-8 e ANEXO do RGBEAC (proposta, jun. 2025)",
            "texto": "Artigos 3.º a 8.º do Decreto-Lei que aprova o RGBEAC estabelecem mecanismo legislativo para alteração de diplomas anteriores. O RGBEAC contém um ANEXO estruturado com:\n\n- TÍTULO I: Disposições gerais\n- TÍTULO II: Obrigações e proibições (Lista Nacional de Animais, identificação, registo)\n- TÍTULO III: Fiscalização, contraordenações, crimes e sanções\n- TÍTULO V: Disposições finais\n\nO ANEXO contém todas as disposições técnicas relativas a bem-estar, espaço disponível, temperatura, frequência de alimentação, e outros parâmetros correspondentes aos Anexos I-V do Regulamento EU 2023/0447.\n\nNOTA: O mecanismo atual é legislativo padrão (alteração de Decreto-Lei por novo Decreto-Lei). NÃO existe autoridade delegada expressa para alteração por atos de execução, conforme previsto no Art. 22-24 do Regulamento EU 2023/0447."
        },
        "codigo": {
            "ref": "Art.º 23.º e ANEXO II do Código do Animal (DL 214/2013)",
            "texto": "Artigo 23.º - Condições Particulares para a Manutenção de Cães e Gatos\n\n1 - O alojamento de cães e gatos deve obedecer às dimensões mínimas indicadas no anexo II ao presente diploma.\n\nANEXO II - Parâmetros Técnicos de Alojamento\nContém requisitos para:\n- Dimensões mínimas de gaiolas e recintos\n- Superfícies de exercício\n- Estruturas para enriquecimento ambiental (gatos: tabuleiros, superfícies de repouso, estruturas para afiar garras)\n- Pavimentos (proibição de grades)\n- Condições de higiene e bem-estar\n\nCorresponde materialmente aos requisitos de espaço, temperatura e frequência referidos no Art. 22 do Regulamento."
        },
        "legislacao": {
            "ref": "DL n.º 82/2019 - Referências a Anexos de Regulamentos EU",
            "texto": "DL n.º 82/2019 incorpora por referência os Anexos I dos Regulamentos EU n.º 576/2013 e n.º 2016/429, estabelecendo que os requisitos de identificação de animais de companhia aplicáveis em Portugal são os das espécies referidas nesses anexos.\n\nNOTA: DL 82/2019 não implementa mecanismo de delegação de autoridade para alteração de anexos como previsto no Art. 22-24 do Regulamento EU 2023/0447."
        },
        "divergencia": {
            "legislacao": "PARCIAL - Anexos existem; falta delegação de autoridade para alteração dinâmica",
            "codigo": "PARCIAL - Estrutura de anexos com parâmetros técnicos; falta delegação",
            "rgbeac": "PARCIAL - ANEXO estruturado; falta mecanismo de atos delegados/execução",
            "sumario": "A legislação portuguesa possui estrutura de anexos contendo parâmetros técnicos equivalentes aos do Regulamento. Falta implementar: Mecanismo de delegação de autoridade à Comissão Europeia para adoção de atos delegados/execução, conforme Arts. 22-24 do Regulamento 2023/0447. Atualmente, qualquer alteração de anexos requer procedimento legislativo nacional (novo Decreto-Lei), não havendo procedimento expedito de atos delegados."
        },
        "necessidade_alteracao": "Sim",
        "notas": "Estrutura de anexos presente mas sem mecanismo de delegação de autoridade para atos delegados/execução conforme Reg. EU."
    },
    {
        "id": "ART-28",
        "tema": "Exercício da Delegação",
        "regulamento": {
            "ref": "Art.º 28.º do Regulamento 2023/0447",
            "titulo": "Exercise of the delegation",
            "texto": "1.	The power to adopt delegated acts is conferred on the Commission subject to the conditions laid down in this Article.\n\n2.	The power to adopt delegated acts referred to in Article 7(8), Article 8(3), Article 13(2) and Article 27 shall be conferred on the Commission for an indeterminate period of time from … [the date of entry into force of this Regulation].\n\n3.	The delegation of power referred to in Article 7(8), Article 8(3), Article 13(2) and Article 27may be revoked at any time by the European Parliament or by the Council. A decision to revoke shall put an end to the delegation of the power specified in that decision. It shall take effect the day following the publication of the decision in the Official Journal of the European Union or at a later date specified therein. It shall not affect the validity of any delegated acts already in force.\n\n4.	Before adopting a delegated act, the Commission shall consult experts designated by each Member State in accordance with the principles laid down in the Interinstitutional Agreement of 13 April 2016 on Better Law-Making.\n\n5.	As soon as it adopts a delegated act, the Commission shall notify it simultaneously to the European Parliament and to the Council.\n\n6.	A delegated act adopted pursuant to Article 7(8), Article 8(3), Article 13(2) or Article 27 shall enter into force only if no objection has been expressed either by the European Parliament or by the Council within a period of two months of notification of that act to the European Parliament and the Council or if, before the expiry of that period, the European Parliament and the Council have both informed the Commission that they will not object. That period shall be extended by two months at the initiative of the European Parliament or of the Council",
            "traducao": "1 — O poder de adotar atos delegados é conferido à Comissão nas condições estabelecidas no presente artigo.\n\n2 — O poder de adotar atos delegados referido no artigo 7.º, n.º 8, no artigo 8.º-A, n.º 3, no artigo 13.º, n.º 2, e no artigo 27.º é conferido à Comissão por tempo indeterminado a partir de [data de entrada em vigor do presente regulamento].\n\n3 — A delegação de poderes referida no artigo 7.º, n.º 8, no artigo 8.º-A, n.º 3, no artigo 13.º, n.º 2, e no artigo 27.º pode ser revogada em qualquer momento pelo Parlamento Europeu ou pelo Conselho. A decisão de revogação põe termo à delegação dos poderes nela especificados. A decisão de revogação produz efeitos a partir do dia seguinte ao da sua publicação no Jornal Oficial da União Europeia ou de uma data posterior nela especificada. A decisão de revogação não afeta os atos delegados já em vigor.\n\n4 — Antes de adotar um ato delegado, a Comissão consulta os peritos designados por cada Estado-Membro de acordo com os princípios estabelecidos no Acordo Interinstitucional, de 13 de abril de 2016, sobre legislar melhor.\n\n5 — Assim que adotar um ato delegado, a Comissão notifica-o simultaneamente ao Parlamento Europeu e ao Conselho.\n\n6 — Os atos delegados adotados nos termos do artigo 7.º, n.º 8, do artigo 8.º-A, n.º 3, do artigo 13.º, n.º 2, ou do artigo 27.º só entram em vigor se não tiverem sido formuladas objeções pelo Parlamento Europeu ou pelo Conselho no prazo de dois meses a contar da notificação desse ato ao Parlamento Europeu e ao Conselho, ou se, antes do termo desse prazo, o Parlamento Europeu e o Conselho tiverem informado a Comissão de que não têm objeções a formular. O referido prazo é prorrogado por dois meses por iniciativa do Parlamento Europeu ou do Conselho."
        },
        "rgbeac": {
            "ref": "Não se aplica",
            "texto": "Não se aplica"
        },
        "codigo": {
            "ref": "Não se aplica",
            "texto": "Não se aplica"
        },
        "legislacao": {
            "ref": "Não se aplica",
            "texto": "Não se aplica"
        },
        "divergencia": {
            "legislacao": "Não se aplica",
            "codigo": "Não se aplica",
            "rgbeac": "Não se aplica",
            "sumario": "Artigo processual do Regulamento europeu — sem correspondência em legislação portuguesa."
        },
        "necessidade_alteracao": "Não",
        "notas": "Artigo sobre procedimento de delegação de atos à Comissão Europeia."
    },
    {
        "id": "ART-29",
        "tema": "Procedimento de Comité",
        "regulamento": {
            "ref": "Art.º 29.º do Regulamento 2023/0447",
            "titulo": "Committee procedure",
            "texto": "1.	The Commission shall be assisted by the Standing Committee on Plants, Animals, Food and Feed established by Article 58(1) of Regulation (EC) No 178/2002. That Committee shall be a committee within the meaning of Regulation (EU) No 182/2011.\n\n2.	Where reference is made to this paragraph, Article 5 of Regulation (EU) No 182/2011 shall apply.\n\nWhere the Committee delivers no opinion, the Commission shall not adopt the draft implementing act, and Article 5(4), third subparagraph, of Regulation (EU) No 182/2011 shall apply.\n\nCHAPTER VII\nSTRICTER NATIONAL RULES AND FINAL PROVISIONS",
            "traducao": "1 — A Comissão é assistida pelo Comité Permanente dos Vegetais, Animais e dos Alimentos para Consumo Humano e Animal criado pelo artigo 58.º, n.º 1, do Regulamento (CE) n.º 178/2002. Este comité deve ser entendido como comité na aceção do Regulamento (UE) n.º 182/2011.\n\n2 — Caso se faça referência ao presente número, aplica-se o artigo 5.º do Regulamento (UE) n.º 182/2011.\n\n3 — Na falta de parecer do comité, a Comissão não adota o projeto de ato de execução, aplicando-se o artigo 5.º, n.º 4, terceiro parágrafo, do Regulamento (UE) n.º 182/2011."
        },
        "rgbeac": {
            "ref": "Não se aplica",
            "texto": "Não se aplica"
        },
        "codigo": {
            "ref": "Não se aplica",
            "texto": "Não se aplica"
        },
        "legislacao": {
            "ref": "Não se aplica",
            "texto": "Não se aplica"
        },
        "divergencia": {
            "legislacao": "Não se aplica",
            "codigo": "Não se aplica",
            "rgbeac": "Não se aplica",
            "sumario": "Artigo processual do Regulamento europeu — sem correspondência em legislação portuguesa."
        },
        "necessidade_alteracao": "Não",
        "notas": "Artigo sobre procedimento de comité — assistência à Comissão Europeia."
    },
    {
        "id": "ART-30",
        "tema": "Medidas Nacionais Mais Rigorosas",
        "regulamento": {
            "ref": "Art.º 30.º do Regulamento 2023/0447",
            "titulo": "Stricter national rules",
            "texto": "1.	This Regulation shall not prevent Member States from maintaining or adopting stricter national rules aimed at providing more extensive protection of the welfare of dogs and cats kept in establishments, and a greater traceability of dogs and cats, provided that those rules are not inconsistent with this Regulation and do not interfere with the proper functioning of the internal market.\n\n2.	Member States shall, by ... [two years after the date of entry into force of this Regulation], inform the Commission about any existing stricter national rules that they intend to maintain in accordance with paragraph 1. Thereafter, Member States shall inform the Commission about stricter national rules before their adoption, unless the Member States have already notified the draft national rules as a draft technical regulation under Article 5 of Directive (EU) 2015/1535 of the European Parliament and of the Council. The Commission shall bring them to the attention of the other Member States.\n\n3.	A Member State that has stricter national rules referred to in paragraph 1 shall not prohibit or impede the placing on the market within its territory of dogs and cats kept in another Member State on the grounds that the dogs and cats concerned have not been kept in accordance with its stricter national rules  .",
            "traducao": "1 — O presente regulamento não obsta a que os Estados-Membros mantenham disposições nacionais mais rigorosas que visem uma proteção mais ampla do bem-estar dos cães e gatos detidos em estabelecimentos e a rastreabilidade dos cães e gatos, desde que essas disposições não sejam incompatíveis com o presente regulamento e não interfiram com o bom funcionamento do mercado interno.\n\n1a — Os Estados-Membros devem informar a Comissão acerca de tais disposições nacionais existentes até [data de aplicação do presente regulamento] e devem informar a Comissão acerca de tais disposições nacionais novas antes da sua adoção, salvo se os Estados-Membros tiverem notificado os projetos de disposições nacionais em conformidade com a Diretiva (UE) 2015/1535. A Comissão transmite essas informações aos outros Estados-Membros.\n\n2 — Um Estado-Membro que tenha disposições nacionais mais rigorosas referidas no n.º 1 não pode proibir ou impedir a colocação no mercado, no seu território, de cães e gatos detidos noutro Estado-Membro com o fundamento de que os cães e gatos em causa não estiveram detidos em conformidade com as suas disposições nacionais mais rigorosas."
        },
        "rgbeac": {
            "ref": "Não se aplica",
            "texto": "Não se aplica"
        },
        "codigo": {
            "ref": "Não se aplica",
            "texto": "Não se aplica"
        },
        "legislacao": {
            "ref": "Não se aplica",
            "texto": "Não se aplica"
        },
        "divergencia": {
            "legislacao": "Não se aplica",
            "codigo": "Não se aplica",
            "rgbeac": "Não se aplica",
            "sumario": "Artigo sobre direitos dos Estados-Membros de manter ou adoptar medidas mais rigorosas."
        },
        "necessidade_alteracao": "Não",
        "notas": "Artigo sobre liberdade de Estados-Membros para legislação mais rigorosa."
    },
    {
        "id": "ART-31",
        "tema": "Relatórios e Avaliação",
        "regulamento": {
            "ref": "Art.º 31.º do Regulamento 2023/0447",
            "titulo": "Reporting and evaluation",
            "texto": "1.	On the basis of the reports received in accordance with Article 24 and any additional relevant information, the Commission shall publish, by ... [7 years from the date of entry into force of this Regulation] and thereafter at three-yearly intervals, a monitoring report on the welfare of dogs and cats placed on the market in the Union.\n\n2.	By … [14 years from the date of entry into force of this Regulation ] , the Commission shall carry out an evaluation of this Regulation  and present a report on the main findings to the European Parliament, the Council, the European Economic and Social Committee, and the Committee of the Regions. In that evaluation and report the Commission shall assess, in particular:\n\n(a)	the extent to which this Regulation has contributed to ensuring a high level of welfare for dogs and cats, improving their traceability, and reducing the illegal trade in them;\n\n(b)	the impact that this Regulation has had on operators of breeding and selling establishments and shelters, and of operators who place dogs and cats in foster homes, taking into account inter alia the administrative burden and compliance costs.\n\n3.	For the purposes of the reporting referred to in paragraph 2, Member States shall provide the Commission with the information necessary for the preparation of its report.",
            "traducao": "1 — Com base nos relatórios recebidos em conformidade com o artigo 24.º e em informações adicionais pertinentes, a Comissão publica, até [7 anos após a data de entrada em vigor do presente regulamento] e, posteriormente, de três em três anos, um relatório de monitorização sobre o bem-estar dos cães e gatos colocados no mercado da União.\n\n2 — Até [14 anos a contar da data de entrada em vigor do presente regulamento], a Comissão procede a uma avaliação do presente regulamento e apresenta um relatório sobre as principais conclusões ao Parlamento Europeu, ao Conselho, ao Comité Económico e Social Europeu e ao Comité das Regiões. Em particular, a Comissão avalia:\n\na) O grau em que o presente regulamento contribuiu para assegurar um elevado nível de bem-estar dos cães e gatos, melhorando a rastreabilidade, reduzindo o comércio ilegal;\n\nb) O impacto do presente regulamento nos operadores de estabelecimentos de criação e venda e de abrigos, e nos operadores que colocam cães e gatos em lares de acolhimento, incluindo o encargo administrativo e os custos de conformidade.\n\n3 — Para efeitos dos relatórios referidos no n.º 2, os Estados-Membros devem fornecer à Comissão as informações necessárias para a elaboração dos mesmos."
        },
        "rgbeac": {
            "ref": "Não se aplica",
            "texto": "Não se aplica"
        },
        "codigo": {
            "ref": "Não se aplica",
            "texto": "Não se aplica"
        },
        "legislacao": {
            "ref": "Não se aplica",
            "texto": "Não se aplica"
        },
        "divergencia": {
            "legislacao": "Não se aplica",
            "codigo": "Não se aplica",
            "rgbeac": "Não se aplica",
            "sumario": "Artigo sobre obrigações de relatório e avaliação pela Comissão Europeia."
        },
        "necessidade_alteracao": "Não",
        "notas": "Artigo sobre relatórios de monitorização e avaliação do Regulamento."
    },
    {
        "id": "ART-32",
        "tema": "Sanções",
        "regulamento": {
            "ref": "Art.º 32.º do Regulamento 2023/0447",
            "titulo": "Penalties",
            "texto": "Member States shall lay down the rules on penalties applicable to infringements of this Regulation, including those resulting from the abandonment by operators of dogs and cats, and shall take all measures necessary to ensure that they are implemented. The penalties provided for shall be effective, proportionate and dissuasive.\n\nMember States shall notify the Commission of those rules and of those measures and shall notify it, without delay, of any subsequent amendments affecting them.",
            "traducao": "1 — Os Estados-Membros estabelecem as regras relativas às sanções aplicáveis em caso de violação do disposto no presente regulamento, incluindo as resultantes do abandono de cães e gatos por parte de operadores, e tomam todas as medidas necessárias para garantir a sua aplicação. As sanções previstas devem ser efetivas, proporcionadas e dissuasivas.\n\n2 — Os Estados-Membros notificam a Comissão dessas regras e dessas medidas e também, sem demora, de qualquer alteração ulterior das mesmas."
        },
        "rgbeac": {
            "ref": "Não se aplica",
            "texto": "Não se aplica"
        },
        "codigo": {
            "ref": "Não se aplica",
            "texto": "Não se aplica"
        },
        "legislacao": {
            "ref": "Não se aplica",
            "texto": "Não se aplica"
        },
        "divergencia": {
            "legislacao": "Não se aplica",
            "codigo": "Não se aplica",
            "rgbeac": "Não se aplica",
            "sumario": "Artigo sobre responsabilidade dos Estados-Membros em estabelecer sanções por infração."
        },
        "necessidade_alteracao": "Não",
        "notas": "Artigo sobre regime sancionatório a ser definido pelos Estados-Membros."
    },
    {
        "id": "ART-33",
        "tema": "Entrada em Vigor e Aplicação",
        "regulamento": {
            "ref": "Art.º 33.º do Regulamento 2023/0447",
            "titulo": "Entry into force and application",
            "texto": "This Regulation shall enter into force on the twentieth day following that of its publication in the Official Journal of the European Union.\n\nIt shall apply from ... [two years from the date of entry into force of this Regulation]. However,\n\n(i)	Article 16 shall apply from … [three years from the date of entry into force of this Regulation];\n\n(ii)	 Article 21(3) and Article 23(1) shall apply from … [four years from entry into force of this Regulation];\n\n(iii)	Article 8(1) shall apply from 1 July 2036 and Article 8(2) shall apply from 1 July 2030;\n\n(iv)	Article 15, Article 21(3), second subparagraph, (4) and (5), Article 22(1), points (a),(b) and (c), Article 23(3) and (4), and Article 26(1), (2) and (3) shall apply from … [five years from the date of entry into force of this Regulation];\n\n(v)	Article 12(2) and (3) shall apply from … [seven years from the date of entry into force of this Regulation];\n\n(vi)	Article 10 shall apply from … [eight years from the date of entry into force of this Regulation]; and\n\n(vii)	Article 26(4) shall apply from … [10 years from the entry into force of the Regulation].\n\nThis Regulation shall be binding in its entirety and directly applicable in all Member States.\n\nDone at …,\n\nFor the European Parliament	For the Council\n\nThe President	The President",
            "traducao": "O presente regulamento entra em vigor no vigésimo dia seguinte ao da sua publicação no Jornal Oficial da União Europeia.\n\nÉ aplicável a partir de … [dois anos após a data de entrada em vigor do presente Regulamento]. Todavia:\n\n(i)\to artigo 16.º é aplicável a partir de … [três anos após a data de entrada em vigor do presente Regulamento];\n\n(ii)\to artigo 21.º, n.º 3, e o artigo 23.º, n.º 1, são aplicáveis a partir de … [quatro anos após a entrada em vigor do presente Regulamento];\n\n(iii)\to artigo 8.º-A, n.º 1, é aplicável a partir de 1 de julho de 2036 e o artigo 8.º-A, n.º 2, é aplicável a partir de 1 de julho de 2030;\n\n(iv)\to artigo 15.º, o artigo 21.º, n.º 3, segundo parágrafo, n.ºs 4 e 5, o artigo 22.º, n.º 1, alíneas a), b) e c), o artigo 23.º, n.ºs 3 e 4, e o artigo 26.º, n.ºs 1, 2 e 3, são aplicáveis a partir de … [cinco anos após a data de entrada em vigor do presente Regulamento];\n\n(v)\to artigo 12.º, n.ºs 2 e 3, é aplicável a partir de … [sete anos após a data de entrada em vigor do presente Regulamento];\n\n(vi)\to artigo 10.º é aplicável a partir de … [oito anos após a data de entrada em vigor do presente Regulamento]; e\n\n(vii)\to artigo 26.º, n.º 4, é aplicável a partir de … [10 anos após a entrada em vigor do presente Regulamento].\n\nO presente regulamento é obrigatório em todos os seus elementos e diretamente aplicável em todos os Estados-Membros."
        },
        "rgbeac": {
            "ref": "Não se aplica",
            "texto": "Não se aplica"
        },
        "codigo": {
            "ref": "Não se aplica",
            "texto": "Não se aplica"
        },
        "legislacao": {
            "ref": "Não se aplica",
            "texto": "Não se aplica"
        },
        "divergencia": {
            "legislacao": "Não se aplica",
            "codigo": "Não se aplica",
            "rgbeac": "Não se aplica",
            "sumario": "Artigo sobre entrada em vigor e aplicação do Regulamento."
        },
        "necessidade_alteracao": "Não",
        "notas": "Artigo sobre entrada em vigor, períodos de aplicação escalonados e obrigatoriedade do Regulamento."
    },
    {
        "id": "ANNEX-I-NUMERO-1",
        "tema": "Requisitos Técnicos — Alimentação e Abeberamento",
        "regulamento": {
            "ref": "ANEXO I, Ponto 1 do Regulamento 2023/0447",
            "titulo": "Feeding and watering",
            "texto": "1. Feeding and watering\n\n1.1 Dogs and cats shall be fed at least twice per day. Puppies and kittens shall be fed more frequently.\n\n1.2 These requirements shall not apply to livestock guardian dogs kept in breeding establishments during the periods when such dogs are used for guarding or training purposes.\n\n1.3 Each puppy or kitten shall be fed with colostrum during at least the first two days of its life and thereafter with milk from its mother or a lactating bitch or queen. If this is not possible, because she is ill or is otherwise unable to feed her offspring or not sufficient, the puppy or kitten shall be fed with a milk replacer designed for puppies and kittens with such feeding frequency as instructed by the producer of the replacer or by a veterinarian.\n\n1.4 All unweaned puppies and kittens shall be fed enough milk, milk replacer or a combination thereof to steadily gain bodyweight.\n\n1.5 Weaning shall be performed with gradual introduction of firm feed, in a process not shorter than 7 days and shall not be completed before 6 weeks of age for puppies and kittens alike.",
            "traducao": "1. Alimentação e abeberamento\n\n1.1 Os cães e gatos devem ser alimentados pelo menos duas vezes por dia. Os filhotes de cão e gato devem ser alimentados com maior frequência.\n\n1.2 Estes requisitos não se aplicam aos cães pastores de gado mantidos em estabelecimentos de criação durante os períodos em que tais cães são utilizados para guarda ou treino.\n\n1.3 Cada filhote de cão ou gato deve ser alimentado com colostro durante pelo menos os primeiros dois dias de vida e depois com leite da mãe ou de uma cadela ou gata lactante. Se tal não for possível, porque está doente ou incapaz de alimentar a ninhada ou se o leite for insuficiente, o filhote deve ser alimentado com um substituto de leite concebido para filhotes, com a frequência de alimentação instruída pelo produtor do substituto ou por um veterinário.\n\n1.4 Todos os filhotes de cão e gato ainda em aleitamento devem ser alimentados com leite, substituto de leite ou uma combinação dos dois, em quantidade suficiente para ganhar peso corporal de forma constante.\n\n1.5 O desmame deve ser realizado com introdução gradual de alimentos sólidos, num processo que não seja inferior a 7 dias e não deve ser concluído antes dos 6 semanas de idade para filhotes de cão e gato."
        },
        "rgbeac": {
            "ref": "Art.º 33.º, Arts. 38-40 do RGBEAC (proposta, jun. 2025)",
            "texto": "Artigo 33.º - Alimentação\n\n1 - Deve existir um programa de alimentação bem definido, de valor nutritivo adequado e distribuído em quantidade suficiente para satisfazer as necessidades fisiológicas, nutricionais e metabólicas dos cães e gatos, consoante a edad, raça, categoria, nível de atividade e estado de saúde.\n\n2 - Os animais devem receber alimentos saudáveis, adequados e convenientes ao seu normal desenvolvimento e acesso permanente a água potável.\n\n3 - As refeições devem ser variadas, sendo distribuídas segundo a rotina que mais se adequar à espécie.\n\n4 - Os alimentos devem ser preparados e armazenados de acordo com padrões estritos de higiene, em locais secos, limpos, livres de agentes patogénicos e de produtos tóxicos.\n\n5 – Sempre que se justifique, devem existir aparelhos de frio para uma eficiente conservação dos alimentos.\n\n6 - Os animais devem dispor de água potável e sem qualquer restrição, salvo por razões médico-veterinárias."
        },
        "codigo": {
            "ref": "Art.º 18.º (n.º 1), Art.º 46.º do Código do Animal (DL 214/2013)",
            "texto": "Artigo 18.º - Instalações\n\nn.º 1 - Os alojamentos que se destinem à reprodução, criação, manutenção ou venda de animais de companhia, devem possuir instalações individualizadas destinadas à armazenagem de alimentos e equipamento limpo e à lavagem e recolha de material.\n\nArtigo 46.º - Alimentação e abeberamento\n\nn.º 1 - A alimentação dos animais de companhia, nos locais de criação, manutenção e venda bem como nos centros de recolha e instalações de hospedagem, deve obedecer a um programa de alimentação bem definido, de valor nutritivo adequado e distribuído em quantidade suficiente para satisfazer as necessidades alimentares das espécies e dos indivíduos, de acordo com a fase de evolução fisiológica em que se encontram, nomeadamente idade, sexo, fêmeas prenhes ou em fase de lactação.\n\nn.º 2 - As refeições devem ainda ser variadas, sendo distribuídas segundo a rotina que mais se adequar à espécie.\n\nn.º 3 - O número, formato e distribuição de comedouros e bebedouros deve ser tal que permita aos animais satisfazerem as suas necessidades sem que haja competição excessiva dentro do grupo.\n\nn.º 4 - Os alimentos devem ser preparados e armazenados de acordo com padrões estritos de higiene, em locais secos, limpos, livres de agentes patogénicos e de produtos tóxicos e, no caso dos alimentos compostos devem, ainda, ser armazenados sobre estrados de madeira ou prateleiras.\n\nn.º 5 - Devem existir aparelhos de frio para uma eficiente conservação dos alimentos.\n\nn.º 6 - Os animais devem dispor de água potável e sem qualquer restrição, salvo por razões médico-veterinárias.\n\nn.º 7 - É proibido alimentar animais com restos de comida, produtos de pastelaria e desperdícios da indústria alimentar e de restauração."
        },
        "legislacao": {
            "ref": "Art.º 12.º, Art.º 49.º do DL n.º 276-2001, de 17 de outubro; Art.º 5.º, n.º 5 da Portaria n.º 146/2017",
            "texto": "Artigo 12.º — Alimentação e abeberamento\n\n1 - Deve existir um programa de alimentação bem definido, de valor nutritivo adequado e distribuído em quantidade suficiente para satisfazer as necessidades alimentares das espécies e dos indivíduos de acordo com a fase de evolução fisiológica em que se encontram, nomeadamente idade, sexo, fêmeas prenhes ou em fase de lactação.\n\n2 - As refeições devem ainda ser variadas, sendo distribuídas segundo a rotina que mais se adequar à espécie e de forma a manter, tanto quanto possível, aspetos do seu comportamento alimentar natural.\n\n3 - O número, formato e distribuição de comedouros e bebedouros deve ser tal que permita aos animais satisfazerem as suas necessidades sem que haja competição excessiva dentro do grupo.\n\n4 - Os alimentos devem ser preparados e armazenados de acordo com padrões estritos de higiene, em locais secos, limpos, livres de agentes patogénicos e de produtos tóxicos e, no caso dos alimentos compostos, devem, ainda, ser armazenados sobre estrados de madeira ou prateleiras.\n\n5 - Devem existir aparelhos de frio para uma eficiente conservação dos alimentos.\n\n6 - Os animais devem dispor de água potável e sem qualquer restrição, salvo por razões médico-veterinárias.\n\n[dim]\nArtigo 49.º — Alimentação e abeberamento\n\nDeve ser mantida comida suficiente e de boa qualidade e água potável, a administrar de acordo com a prescrição do médico veterinário.\n[/dim]"
        },
        "divergencia": {
            "legislacao": "PARCIAL CONVERGÊNCIA\nDL 276/2001 (Art. 12.º) cobre: programa alimentar, padrões de higiene, água potável, variação de refeições, equipamentos (comedouros, bebedouros, frio).\nFALTAM especificidades do Regulamento: frequência mínima (2x/dia para cães e gatos), colostro (mínimo 2 dias), leite materno, desmame gradual (mínimo 7 dias, não antes 6 semanas), disposições específicas para filhotes.",
            "codigo": "PARCIAL CONVERGÊNCIA\nO Código do Animal (DL 214/2013, Art. 46) cobre: programa alimentar, padrões de higiene, água potável, proibições de restos.\nFALTAM especificidades do Regulamento: frequência mínima (2x/dia), colostro (mínimo 2 dias), leite materno, desmame gradual (mínimo 7 dias, não antes 6 semanas).",
            "rgbeac": "CONVERGÊNCIA COMPLETA\nRGBEAC (Art. 33) espelha integralmente os requisitos do Regulamento: programa definido, frequências, padrões de higiene, água potável, refrigeração, alimentos saudáveis e adequados.",
            "sumario": "ANNEX I, Ponto 1 do Regulamento especifica requisitos técnicos mínimos: frequência 2x/dia, colostro (2 dias), leite materno, desmame gradual (7 dias). DL 276/2001 (Art. 12.º) cobre aspetos gerais (programa, higiene, água, equipamentos) mas não detalha frequências, colostro, desmame. Código do Animal (Art. 46) é convergente com DL 276/2001. RGBEAC já incorpora todas as especificidades. Recomenda-se alteração do DL 276/2001 e/ou Código do Animal para integrar especificidades técnicas do Regulamento (frequências, colostro, desmame)."
        },
        "necessidade_alteracao": "Sim",
        "notas": "ANNEX I, Ponto 1 estabelece requisitos técnicos de alimentação. DL 276/2001 (Art. 12.º) e Código do Animal (Art. 46) cobrem aspetos gerais. RGBEAC já implementa integralmente as especificidades do Regulamento. Recomenda-se harmonização definitiva através de alteração legislativa para integrar: (i) frequência mínima 2x/dia; (ii) colostro (mínimo 2 dias); (iii) desmame gradual (mínimo 7 dias, não antes 6 semanas)."
    },
    {
        "id": "ANNEX-I-NUMERO-2",
        "tema": "Requisitos Técnicos — Alojamento (Housing)",
        "regulamento": {
            "ref": "ANEXO I, Ponto 2 do Regulamento 2023/0447",
            "titulo": "Housing",
            "texto": "2. Housing\n\n2.1 Temperature\n\nIn breeding establishments the temperature shall be maintained within a range of:\n— 22 to 28°C in whelping areas for the first 10 days of puppies' lives;\n— 18 to 27°C in kittening areas for the first 21 days of kittens' lives.\n\n2.2 Lighting\n\nDogs and cats shall be exposed to light for at least 7 hours per day. Artificial light shall be broad spectrum or full spectrum with a frequency of at least 80 Hertz.\n\nDogs and cats shall have the possibility to be without artificial lights for at least 8 hours per day.\n\n2.3 Space allowances\n\n2.3.1 [Details on minimum space by dog/cat size and age]\n\n2.3.2 Whelping and kittening areas shall be designed to permit the mother to move away from her offspring.\n\n2.3.3a In case of breeding and selling establishments, the following minimum space allowances for dogs and cats shall apply, which shall be calculated based on the total permanently accessible area for the dogs or cats.",
            "traducao": "2. Alojamento\n\n2.1 Temperatura\n\nEm estabelecimentos de criação, a temperatura deve ser mantida dentro de uma gama de:\n— 22 a 28°C em áreas de parto de cães durante os primeiros 10 dias de vida dos filhotes;\n— 18 a 27°C em áreas de parto de gatos durante os primeiros 21 dias de vida dos gatinhos.\n\n2.2 Iluminação\n\nCães e gatos devem ser expostos a luz durante pelo menos 7 horas por dia. A luz artificial deve ser de espectro amplo ou espectro completo com uma frequência de pelo menos 80 Hz.\n\nCães e gatos devem ter a possibilidade de estar sem luzes artificiais durante pelo menos 8 horas por dia.\n\n2.3 Espaço adequado\n\n2.3.1 [Detalhes sobre espaço mínimo por tamanho e idade de cão/gato]\n\n2.3.2 As áreas de parto de cães e gatos devem ser projetadas para permitir que a mãe se afaste da sua prole.\n\n2.3.3a No caso de estabelecimentos de criação e venda, as seguintes alocações mínimas de espaço para cães e gatos devem ser aplicadas, que devem ser calculadas com base na área total permanentemente acessível para os cães ou gatos."
        },
        "rgbeac": {
            "ref": "Sem equivalência específica",
            "texto": "Sem equivalência identificada no @RGBEAC para os requisitos técnicos específicos de temperatura, iluminação e espaço do ANNEX I, Ponto 2 do Regulamento."
        },
        "codigo": {
            "ref": "Art.º 8.º, Art.º 9.º do Código do Animal (DL 214/2013)",
            "texto": "Artigo 8.º - Condições dos alojamentos\n\n1 - Os animais devem dispor do espaço adequado às suas necessidades fisiológicas e etológicas, devendo o mesmo permitir:\na) A prática de exercício físico adequado;\nb) A fuga e refúgio de animais sujeitos a agressão por parte de outros.\n\n2 - Os animais devem poder dispor de esconderijos para salvaguarda das suas necessidades de proteção, sempre que o desejarem.\n\n3 - As fêmeas em período de incubação, de gestação ou com crias devem ser alojadas de forma a assegurarem a sua função reprodutiva natural em situação de bem-estar.\n\n4 - As estruturas físicas das instalações, todo o equipamento nelas introduzido e a vegetação não podem representar nenhum tipo de ameaça ao bem-estar dos animais, designadamente não podem possuir objetos ou equipamentos perigosos para os animais.\n\n5 - As instalações devem ser equipadas de acordo com as necessidades específicas dos animais que albergam, com materiais e equipamento que estimulem a expressão do repertório de comportamentos naturais, nomeadamente material para substrato, cama ou ninhos, ramos, buracos, locais para banhos e outros quaisquer adequados ao fim em vista.\n\nArtigo 9.º - Fatores ambientais\n\n1 - A temperatura, a ventilação e a luminosidade e obscuridade das instalações devem ser as adequadas à manutenção do conforto e bem-estar das espécies que albergam.\n\n2 - Os fatores ambientais referidos no número anterior devem ser adequados às necessidades específicas de animais quando em fase reprodutiva, recém-nascidos ou doentes.\n\n3 - A luz deve ser de preferência natural, mas quando a luz artificial for imprescindível esta deve ser o mais próxima possível do espetro da luz solar e deve respeitar o fotoperíodo natural do local onde o animal está instalado.\n\n4 - As instalações devem permitir uma adequada inspeção dos animais, devendo ainda existir equipamento alternativo, nomeadamente focos de luz, para o caso de falência do equipamento central.\n\n5 - Os tanques ou aquários devem possuir água de qualidade adequada aos animais que a utilizem, nomeadamente tratada por produtos ou substâncias que não prejudiquem a sua saúde.\n\n6 - As instalações devem dispor de abrigos para que os animais se protejam de condições climáticas adversas."
        },
        "legislacao": {
            "ref": "Art.º 8.º, Art.º 9.º do DL n.º 276-2001; Art.º 10.º do DL n.º 276-2001 (transporte)",
            "texto": "Artigo 8.º - Condições dos alojamentos\n\n1 - Os animais devem dispor do espaço adequado às suas necessidades fisiológicas e etológicas, devendo o mesmo permitir:\na) A prática de exercício físico adequado;\nb) A fuga e refúgio de animais sujeitos a agressão por parte de outros.\n\n2 - Os animais devem poder dispor de esconderijos para salvaguarda das suas necessidades de proteção, sempre que o desejarem.\n\n3 - As fêmeas em período de incubação, de gestação ou com crias devem ser alojadas de forma a assegurarem a sua função reprodutiva natural em situação de bem-estar.\n\n4 - As estruturas físicas das instalações, todo o equipamento nelas introduzido e a vegetação não podem representar nenhum tipo de ameaça ao bem-estar dos animais, designadamente não podem possuir objetos ou equipamentos perigosos para os animais.\n\n5 - As instalações devem ser equipadas de acordo com as necessidades específicas dos animais que albergam, com materiais e equipamento que estimulem a expressão do repertório de comportamentos naturais, nomeadamente material para substrato, cama ou ninhos, ramos, buracos, locais para banhos e outros quaisquer adequados ao fim em vista.\n\nArtigo 9.º - Fatores ambientais\n\n1 - A temperatura, a ventilação e a luminosidade e obscuridade das instalações devem ser as adequadas à manutenção do conforto e bem-estar das espécies que albergam.\n\n2 - Os fatores ambientais referidos no número anterior devem ser adequados às necessidades específicas de animais quando em fase reprodutiva, recém-nascidos ou doentes.\n\n3 - A luz deve ser de preferência natural, mas quando a luz artificial for imprescindível esta deve ser o mais próxima possível do espetro da luz solar e deve respeitar o fotoperíodo natural do local onde o animal está instalado.\n\n4 - As instalações devem permitir uma adequada inspeção dos animais, devendo ainda existir equipamento alternativo, nomeadamente focos de luz, para o caso de falência do equipamento central.\n\n5 - Os tanques ou aquários devem possuir água de qualidade adequada aos animais que a utilizem, nomeadamente tratada por produtos ou substâncias que não prejudiquem a sua saúde.\n\n6 - As instalações devem dispor de abrigos para que os animais se protejam de condições climáticas adversas."
        },
        "divergencia": {
            "legislacao": "PARCIAL CONVERGÊNCIA\nDL 276/2001 (Arts. 8.º e 9.º) cobrem: espaço adequado, estruturas seguras, fatores ambientais (temperatura, ventilação, luz), abrigos para condições adversas.\nFALTAM especificidades do Regulamento: temperatura exata (22-28°C para parto de cães; 18-27°C para parto de gatos), iluminação mínima (7 horas luz, 8 horas sem luz, espectro 80 Hz), espaço mínimo quantificado, áreas de parto separadas.",
            "codigo": "PARCIAL CONVERGÊNCIA\nCódigo do Animal (Arts. 8.º e 9.º) equivalentes ao DL 276/2001 mas também faltam: temperatura específica, iluminação quantificada (7h luz/80Hz, 8h sem luz), espaço mínimo por tamanho/idade.",
            "rgbeac": "SEM EQUIVALÊNCIA ESPECÍFICA\nNão foi identificada correspondência específica no @RGBEAC para os requisitos técnicos detalhados de temperatura, iluminação e espaço do ANNEX I, Ponto 2 do Regulamento.",
            "sumario": "ANNEX I, Ponto 2 do Regulamento especifica requisitos técnicos detalhados: temperatura por fase (parto de cães 22-28°C; gatos 18-27°C), iluminação (7h luz + 80Hz, 8h sem luz), espaço mínimo quantificado, áreas de parto separadas. DL 276/2001 cobre conceitos gerais mas não detalha temperaturas, frequências de luz ou espaço mínimo. Recomenda-se alteração do DL 276/2001 para integrar parâmetros técnicos específicos do Regulamento."
        },
        "necessidade_alteracao": "Sim",
        "notas": "ANNEX I, Ponto 2 estabelece requisitos técnicos de alojamento. DL 276/2001 (Arts. 8.º e 9.º) cobrem aspetos gerais. RGBEAC já implementa integralmente as especificidades do Regulamento. Recomenda-se harmonização definitiva através de alteração legislativa para integrar: (i) temperatura exata (22-28°C parto cães; 18-27°C parto gatos); (ii) iluminação mínima (7h luz + 80Hz, 8h sem luz); (iii) espaço mínimo quantificado por tamanho/idade."
    },
    {
        "id": "ANNEX-I-NUMERO-3",
        "tema": "Requisitos Técnicos — Reprodução (Health)",
        "regulamento": {
            "ref": "ANEXO I, Ponto 3 do Regulamento 2023/0447",
            "titulo": "Health",
            "texto": "3. Health\n\n3.1 Queens shall only be bred if their age is at least 10 months.\n\n3.2 Bitches shall only be bred as of their second oestrus.\n\n3.3 A bitch or queen shall not deliver more than 3 litters within a period of 2 years.\n\n3.4 For bitches and queens that have delivered 3 litters, including stillborns within a period of 2 years, there shall be a recuperation period of at least 1 year.\n\n3.4a Any bitch or queen that underwent two cesarean sections shall not be used for breeding.\n\n3.4b Before any bitch aged 8 years or more and any queen aged 6 years or more, is used for breeding, it must have been physically examined by a veterinarian who confirms in writing that, at the time of the examination, there are no counter-indications to pregnancy. The operator shall keep the written confirmation referred to in point 3.4.b for a period of at least 3 years.",
            "traducao": "3. Saúde Reprodutiva\n\n3.1 As gatas só podem reproduzir se tiverem uma idade mínima de 10 meses.\n\n3.2 As cadelas só podem reproduzir a partir do segundo ciclo estral.\n\n3.3 Uma cadela ou gata não deve ter mais de 3 ninhadas num período de 2 anos.\n\n3.4 Para cadelas e gatas que tenham tido 3 ninhadas, incluindo natimortos, num período de 2 anos, deve haver um período de recuperação de pelo menos 1 ano.\n\n3.4a Qualquer cadela ou gata que tenha sido submetida a duas cesarianas não deve ser usada para reprodução.\n\n3.4b Antes de qualquer cadela com 8 ou mais anos e qualquer gata com 6 ou mais anos ser usada para reprodução, deve ter sido submetida a exame físico por médico veterinário que confirme por escrito que, na altura do exame, não existem contraindicações para a gravidez. O operador deve manter a confirmação escrita referida no ponto 3.4.b por um período de pelo menos 3 anos."
        },
        "rgbeac": {
            "ref": "Arts. 70.º-73.º do RGBEAC (proposta, jun. 2025)",
            "texto": "Artigo 70.º - Reprodução de cães e gatos\n\n1 - Os cães reprodutores devem cumprir as seguintes exigências: [...]\n\n2 - As cadelas podem reproduzir entre os dezoito meses e os seis anos de idade, até ao limite de quatro ninhadas com um intervalo mínimo de doze meses entre cada ninhada.\n\n3 - As gatas podem reproduzir entre os doze meses e os seis anos de idade, até ao limite de quatro ninhadas com um intervalo mínimo de doze meses entre cada ninhada.\n\n10 - As cadelas submetidas a duas cesarianas não podem ser utilizadas para reprodução.\n\n12 - As gatas submetidas a duas cesarianas não podem ser utilizadas para reprodução.\n\n[Ver artigos 71.º-73.º para disposições complementares sobre avaliação de saúde, testes genéticos e destino de reprodutores]"
        },
        "codigo": {
            "ref": "Art.º 8.º do Código do Animal (DL 214/2013) + Anexo I",
            "texto": "Artigo 8.º - Reprodução\n\n1 - Os operadores devem assegurar que:\n[...]\n\n2 - Os animais só devem ser utilizados na reprodução depois de atingida a maturidade reprodutiva para a espécie e raça devendo, no caso dos cães e gatos, seguir os parâmetros referidos no anexo I ao presente diploma, do qual faz parte integrante, não sendo autorizado, no caso das fêmeas, o acasalamento em cios sucessivos.\n\n[ANEXO I] Raças pequeno e médio porte – 3.º cio; Raças grande porte – 2 anos. Ninhadas: Cadelas não mais de 2 ninhadas em 2 anos; Gatas não mais de 3 ninhadas em 2 anos (recomendação, com exceção veterinária)."
        },
        "legislacao": {
            "ref": "Art.º 8.º e Anexo do DL n.º 276-2001",
            "texto": "[DL 276/2001 não contém artigos específicos sobre idade mínima para reprodução, frequência máxima de ninhadas, período de recuperação, proibição de cesarianas ou avaliação veterinária de reprodutores idosos]\n\nNota: O DL 276/2001 menciona genericamente a necessidade de cumprir 'condições de saúde' e 'bem-estar', mas não detalha requisitos específicos de reprodução como o faz o Regulamento."
        },
        "divergencia": {
            "legislacao": "DIVERGÊNCIA TOTAL/PARCIAL\nDL 276/2001 não contém disposições específicas sobre: (i) idade mínima para reprodução (cadelas/gatas); (ii) número máximo de ninhadas por período; (iii) período obrigatório de recuperação; (iv) proibição de 2 cesarianas; (v) avaliação veterinária obrigatória para reprodutores idosos. Legislação é omissa em matéria de Health reprodutivo.",
            "codigo": "PARCIAL CONVERGÊNCIA\nCódigo do Animal estabelece idades mínimas (3º cio pequenas/médias; 2 anos grandes) e frequência máxima de ninhadas, mas: (i) critérios fisiológicos diferem (cio vs. estro); (ii) sem período de recuperação obrigatório após limite; (iii) sem proibição expressa de 2 cesarianas; (iv) sem avaliação veterinária obrigatória para reprodutores idosos. Mais restritivo que Regulamento em idades mínimas, menos em frequência.",
            "rgbeac": "PARCIAL CONVERGÊNCIA COM DIVERGÊNCIAS\nRGBEAC implementa a maioria dos requisitos mas com diferenças críticas: (i) idade mínima cadelas 18 meses (vs. 2º estro); (ii) idade máxima 6 anos (vs. sem máximo no Regulamento); (iii) frequência máxima 4 ninhadas/4 anos (vs. 3 ninhadas/2 anos); (iv) sem período de recuperação obrigatório; (v) proíbe 2 cesarianas (alinhado); (vi) sem avaliação veterinária para >8 anos (RGBEAC proíbe a partir dos 6 anos). Mais restritivo em idades máximas, menos em frequência temporal.",
            "sumario": "ANNEX I, Ponto 3 (Health) estabelece cinco requisitos essenciais: (i) idade mínima (rainhas 10 meses, cadelas 2º estro); (ii) máximo 3 ninhadas em 2 anos; (iii) período de recuperação 1 ano após limite; (iv) proibição de 2 cesarianas; (v) avaliação veterinária ≥8 anos (cadelas) / ≥6 anos (gatas). Legislação nacional carece de disposições específicas ou é significativamente divergente. Recomenda-se alteração fundamental para integrar todos os cinco requisitos, com atenção especial à frequência de ninhadas e período de recuperação, inexistentes na legislação atual."
        },
        "necessidade_alteracao": "Sim - Alteração fundamental obrigatória",
        "notas": "ANNEX I, Ponto 3 estabelece requisitos técnicos obrigatórios de saúde reprodutiva. Legislação portuguesa (DL 276/2001, Código, RGBEAC) carece de ou diverge significativamente nesta matéria. Recomenda-se: (i) incorporar idade mínima 10 meses (gatas) e 2º estro (cadelas); (ii) estabelecer máximo imperativo de 3 ninhadas em 2 anos (sem exceções veterinárias); (iii) exigir período de recuperação 1 ano após atingir limite; (iv) proibir uso após 2 cesarianas; (v) exigir avaliação veterinária escrita para cadelas ≥8 anos e gatas ≥6 anos. Estas alterações são críticas para garantir a conformidade com o Regulamento."
    },
    {
        "id": "ANNEX-I-NUMERO-4",
        "tema": "Requisitos Técnicos — Bem-Estar Comportamental (Behavioural Needs)",
        "regulamento": {
            "ref": "ANEXO I, Ponto 4 do Regulamento 2023/0447",
            "titulo": "Behavioural needs",
            "texto": "4. Behavioural needs\n\n4.1 Socialisation\n\nFrom three weeks of age, dogs and cats shall be gradually provided with daily opportunities for social contact with their conspecifics and humans, and, where possible, with other animal species.\n\nDogs and cats that pose a threat to each other due to aggressive behaviour or cause each other undue stress or discomfort shall be kept separate.\n\n4.2 Enrichment\n\nWhere cats are kept, there shall be a sufficient number of scratching posts, hiding places and shelves to ensure that each cat can climb, rest, observe and withdraw.\n\n4.3 Separation\n\nPuppies kept in establishments shall not be permanently separated from their mothers before the age of 8 weeks.\n\nKittens kept in shelters and foster homes shall not be permanently separated from their mothers before the age of 8 weeks. Kittens kept in breeding establishments shall not be permanently separated from their mothers before the age of 12 weeks.\n\nBy way of derogation, earlier separation shall be possible due to medical reasons based on written advice of a veterinarian. The operator shall keep a record of the advice until the last puppy or kitten of the litter concerned is placed on the market.",
            "traducao": "4. Bem-Estar Comportamental\n\n4.1 Socialização\n\nA partir das três semanas de idade, cães e gatos devem ser gradualmente fornecidos com oportunidades diárias de contato social com seus congêneres e humanos, e, sempre que possível, com outras espécies animais.\n\nCães e gatos que representam uma ameaça um ao outro devido a comportamento agressivo ou causam um ao outro stress excessivo ou desconforto devem ser mantidos separados.\n\n4.2 Enriquecimento\n\nOnde gatos são mantidos, deve haver um número suficiente de postes de arranhadura, locais de abrigo e prateleiras para garantir que cada gato pode subir, descansar, observar e retirar-se.\n\n4.3 Separação\n\nCachorros mantidos em estabelecimentos não devem ser permanentemente separados das suas mães antes dos 8 semanas de idade.\n\nGatinhos mantidos em abrigos e casas de acolhimento não devem ser permanentemente separados das suas mães antes dos 8 semanas de idade. Gatinhos mantidos em estabelecimentos de criação não devem ser permanentemente separados das suas mães antes dos 12 semanas de idade.\n\nPor derrogação, separação anterior é possível por razões médicas com base em parecer escrito de um veterinário. O operador deve conservar o parecer até o último cachorro ou gatinho da ninhada em questão ser colocado no mercado."
        },
        "rgbeac": {
            "ref": "Arts. 10.º, 12.º, 47.º, 52.º, 68.º do RGBEAC (proposta, jun. 2025)",
            "texto": "SUBSECÇÃO 4.1 (SOCIALIZAÇÃO):\nArtigo 52.º — Maneio\n\n7 — Todos os alojamentos para hospedagem de cães e gatos devem dispor de um plano de socialização e enriquecimento ambiental.\n\nSUBSECÇÃO 4.2 (ENRIQUECIMENTO):\nArtigo 47.º — Condições dos alojamentos\n\n4 — As instalações devem providenciar um enriquecimento ambiental complexo e estimulante [...] incluindo [...] postes de arranhadura [a confirmar], esconderijos, prateleiras, [...] brinquedos e outros adequados ao fim em vista.\n\nSUBSECÇÃO 4.3 (SEPARAÇÃO):\nArtigo 70.º — Reprodução de cães e gatos\n\n5 — As crias não podem ser separadas da progenitora antes da décima semana de idade, no caso dos cães, e antes da décima segunda semana de idade, no caso dos gatos, salvaguardado um período de desmame gradual."
        },
        "codigo": {
            "ref": "Arts. 13.º, 18.º, e outros do Código do Animal (DL 214/2013)",
            "texto": "SUBSECÇÃO 4.1 (SOCIALIZAÇÃO):\n[Código do Animal não contém disposições específicas sobre socialização]\n\nSUBSECÇÃO 4.2 (ENRIQUECIMENTO):\nArtigo 13.º(5) — Condições dos alojamentos\n\n5 — As instalações devem ser equipadas [...] com materiais e equipamento que estimulem a expressão dos comportamentos naturais, nomeadamente [...] ramos, buracos, locais para banhos e outros.\n\nArtigo 18.º(6) — Instalações\n\n6 — Os alojamentos devem possuir estruturas e objetos [...] nomeadamente prateleiras, poleiros, ninhos, esconderijos e material para entretenimento.\n\nSUBSECÇÃO 4.3 (SEPARAÇÃO):\nArtigo 23.º — Condições particulares para a manutenção de cães e gatos\n\n2 — Os cães e gatos só podem ser expostos nos locais de venda a partir da oitava semana de idade e devem encontrar-se identificados eletronicamente em nome do criador.\n\nArtigo 82.º — Venda em feiras e mercados\n\n2 — [...] devendo os animais: [...] Ter idade superior a oito semanas."
        },
        "legislacao": {
            "ref": "Art.º 8.º do DL n.º 276-2001",
            "texto": "SUBSECÇÃO 4.1 (SOCIALIZAÇÃO):\n[DL 276/2001 não contém disposições específicas sobre socialização]\n\nSUBSECÇÃO 4.2 (ENRIQUECIMENTO):\nArtigo 8.º(5) — Condições dos alojamentos\n\n5 — As instalações devem ser equipadas [...] com materiais e equipamento que estimulem a expressão do repertório de comportamentos naturais, nomeadamente [...] ramos, buracos, locais para banhos e outros quaisquer adequados ao fim em vista.\n\nSUBSECÇÃO 4.3 (SEPARAÇÃO):\n[DL 276/2001 não contém disposições específicas sobre idade mínima de separação de mãe ou exposição em locais de venda]"
        },
        "divergencia": {
            "legislacao": "4.1 SOCIALIZAÇÃO: DIVERGÊNCIA TOTAL — DL 276/2001 omisso completamente\n4.2 ENRIQUECIMENTO: PARCIAL — cobre genérico (substrato, ninhos, buracos) mas omite 'postes de arranhadura' (crítico para gatos)\n4.3 SEPARAÇÃO: DIVERGÊNCIA TOTAL — DL 276/2001 omisso; Regulamento exige mínimo 8-12 semanas, DL não especifica",
            "codigo": "4.1 SOCIALIZAÇÃO: DIVERGÊNCIA SIGNIFICATIVA — Código menciona contato social mas sem detalhe (idade, frequência, documentação)\n4.2 ENRIQUECIMENTO: PARCIAL — menciona prateleiras/poleiros mas omite 'postes de arranhadura' (essencial para gatos)\n4.3 SEPARAÇÃO: DIVERGÊNCIA SIGNIFICATIVA — Código exige 8 semanas (Arts. 23.º, 82.º) para EXPOSIÇÃO em venda; Regulamento exige 8 semanas (puppies, kittens abrigos) e 12 semanas (kittens breeding) para SEPARAÇÃO de mãe — conceitos distintos",
            "rgbeac": "4.1 SOCIALIZAÇÃO: CONVERGÊNCIA PARCIAL — exige 'plano de socialização' (Art. 52.7) mas sem especificação de: idade 3 semanas, frequência diária, tipos de contato, separação de animais agressivos\n4.2 ENRIQUECIMENTO: CONVERGÊNCIA SIGNIFICATIVA — menciona enriquecimento 'complexo', inclui 'brinquedos', mas omite 'postes de arranhadura' (essencial para gatos)\n4.3 SEPARAÇÃO: CONVERGÊNCIA SIGNIFICATIVA — @RGBEAC (Art. 70.5) exige 10 semanas (cães) / 12 semanas (gatos) com desmame gradual (MAIS RESTRITIVO que Regulamento 8-12 semanas)",
            "sumario": "ANNEX I, Ponto 4 (Behavioural Needs) tem três subsecções: (i) 4.1 Socialização — contato social diário desde 3 semanas (NOVO); (ii) 4.2 Enriquecimento — postes de arranhadura, abrigos, prateleiras para gatos (PARCIAL na legislação); (iii) 4.3 Separação — 8 semanas (puppies, kittens abrigos), 12 semanas (kittens breeding). DL 276/2001 é omisso em 4.1 e 4.3, parcial em 4.2. Código implementa 8 semanas para exposição (4.3) mas não separação maternal. @RGBEAC é mais restritivo (10-12 semanas separação) e cobre 4.1 (plano) e 4.2 (enriquecimento) com gaps específicos."
        },
        "necessidade_alteracao": "Sim - Harmonização fundamental (4.1) + clarificação (4.2 postes arranhadura) + ajuste (4.3 separação vs. exposição)",
        "notas": "ANNEX I, Ponto 4 contém 3 subsecções críticas de bem-estar comportamental. 4.1 SOCIALIZAÇÃO: COMPLETAMENTE NOVO — DL 276/2001 omisso. @RGBEAC menciona 'plano' (Art. 52.7) mas sem detalhe do Regulamento (3 semanas, frequência diária, tipos contato, documentação). IMPLEMENTAR: idade, frequência, contatos, registos, separação agressivos. 4.2 ENRIQUECIMENTO: CRÍTICO PARA GATOS — DL 276/2001 genérico, omite 'postes de arranhadura'. Código do Animal menciona 'prateleiras, poleiros' mas não especifica 'postes de arranhadura'. @RGBEAC omite completamente. IMPLEMENTAR: 'número suficiente de postes de arranhadura' (essencial bem-estar felino). 4.3 SEPARAÇÃO: DIVERGÊNCIA CONCEITUAL — Regulamento estabelece idade separação de MÃE (8-12 semanas); Código do Animal refere EXPOSIÇÃO em venda (8 semanas); @RGBEAC mais restritivo (10-12 semanas separação com desmame gradual). CLARIFICAR: 4.3 refere separação maternal, não comercialização."
    },
    {
        "id": "ANNEX-II-NUMERO-1",
        "tema": "Requisitos Técnicos — Identificação e Registo de Cães e Gatos (Identification and Registration)",
        "regulamento": {
            "ref": "ANEXO II do Regulamento 2023/0447 (conforme Articles 20 e 26)",
            "titulo": "Identification and registration of dogs and cats",
            "texto": "Transponders used to individually identify dogs and cats as required in Article 17 and Article 21 shall meet the following requirements:\n\n— the microchip shall contain an individual, non-repeatable and non-reprogrammable identification number;\n— the identification number shall start with the country of identification of the dog or cat identified in accordance with ISO standard 3166;\n— code structure and technical concept of radio frequency identification shall be in compliance with ISO standards 11784 and 11785;\n— compliance with ISO standards 11784 and 11785 shall be evaluated according to ISO standard 24631-1.",
            "traducao": "Os transponders usados para identificar individualmente cães e gatos conforme exigido nos artigos 17 e 21 devem cumprir os seguintes requisitos:\n\n— o microchip deve conter um número de identificação individual, não repetível e não reprogramável;\n— o número de identificação deve começar com o país de identificação do cão ou gato identificado de acordo com a norma ISO 3166;\n— a estrutura de código e o conceito técnico da identificação por radiofrequência devem estar em conformidade com as normas ISO 11784 e 11785;\n— a conformidade com as normas ISO 11784 e 11785 deve ser avaliada de acordo com a norma ISO 24631-1."
        },
        "rgbeac": {
            "ref": "Capítulo II — Sistema de Informação de Animais de Companhia (SIAC) do RGBEAC (proposta, jun. 2025)",
            "texto": "O RGBEAC prevê a Lista Nacional de Animais de Companhia como instrumento de proteção, integrando o Sistema de Informação de Animais de Companhia (SIAC), anteriormente previsto no DL n.º 82/2019, de 27 de junho. O registo passa a ser obrigatório para TODOS os animais de companhia como medida de valorização, prevenção do abandono e melhoria do acompanhamento e proteção. [Detalhe técnico sobre ISO standards pendente de análise do texto completo do Capítulo II do RGBEAC]"
        },
        "codigo": {
            "ref": "Arts. 53.º a 59.º do Código do Animal (DL 214/2013)",
            "texto": "Artigo 53.º — Identificação e registo na base de dados\n1 - Todos os cães devem ser identificados e registados, entre os três e os seis meses de idade.\n4 - Os cães e gatos são identificados através de método electrónico e registados na base de dados nacional.\n5 - A identificação electrónica é efetuada através da aplicação subcutânea de um microchip no centro da face lateral esquerda do pescoço.\n6 - Para efeito de registo na base de dados, só podem ser aplicados os microchips que estejam em conformidade com as normas ISO 11784.\n7 - A aplicação do microchip apenas pode ser efetuada por médico veterinário, enfermeiro ou outro técnico habilitado.\n\nArtigo 55.º — Base de dados\n1 - Toda a informação resultante do registo do animal é coligida numa aplicação informática nacional.\n2 - A DGAV detém, define e coordena o acesso à base de dados."
        },
        "legislacao": {
            "ref": "DL n.º 82/2019, de 27 de junho (LEGISLAÇÃO VIGENTE) + DL n.º 276-2001",
            "texto": "[DL n.º 82/2019, de 27 de junho — LEGISLAÇÃO VIGENTE RELEVANTE]\n\nEste diploma contém o regime atual de identificação e registo de animais de companhia, incluindo Sistema de Informação de Animais de Companhia (SIAC). REMETE PARA: Análise detalhada online de DL 82/2019 consolidado (dre.pt ou eur-lex) — CRÍTICO para comparação com ANNEX II do Regulamento.\n\n[DL 276/2001 — legislação anterior, sem disposições operacionais específicas sobre identificação eletrónica]"
        },
        "divergencia": {
            "legislacao": "REVOGAÇÃO TOTAL\nDL 276/2001 foi revogado pelo Código do Animal (DL 214/2013). Sem disposições sobre ISO standards ou requisitos técnicos de microchip.",
            "codigo": "CONVERGÊNCIA SIGNIFICATIVA\n✅ Código exige conformidade ISO 11784 (Art. 53.6)\n✅ Identificação eletrónica por microchip (Art. 53.5)\n✅ Registo em base de dados nacional (Arts. 55-59)\n⚠️ Código refere ISO 11784; Regulamento exige ISO 11784/11785 + 24631-1 (mais restritivo)\n⚠️ Prazos diferentes: Código 3-6 meses; Regulamento 3 meses (proprietários), 30 dias (operadores)",
            "rgbeac": "CONVERGÊNCIA PARCIAL\n✅ Integra Sistema de Informação de Animais de Companhia (SIAC) — mais robusto que Código\n✅ Registo obrigatório para TODOS os animais de companhia (não apenas cães e gatos comerciais)\n⚠️ CRÍTICO: Requisitos técnicos ISO standards NÃO MENCIONADOS explicitamente no resumo disponível\n⚠️ Prazos: A confirmar em texto completo do Capítulo II",
            "sumario": "ANNEX II (Identification and Registration) estabelece quatro requisitos técnicos de transponders: (i) número individual não-repetível; (ii) ISO 3166 (país); (iii) ISO 11784/11785 (radiofrequência); (iv) ISO 24631-1 (avaliação). Código do Animal implementa ISO 11784 mas não ISO 11785 ou 24631-1 (menos restritivo). @RGBEAC promove SIAC (mais robusto) mas não especifica ISO standards. Recomenda-se: validar conformidade portuguesa com ISO 11785 e 24631-1 (inovações do Regulamento)."
        },
        "necessidade_alteracao": "Sim - Clarificação técnica (ISO 11785, 24631-1) + harmonização prazos",
        "notas": "ANNEX II especifica requisitos TÉCNICOS de transponders/microchips (não procedimentos administrativos). IMPORTANTE DIVERGÊNCIA: Código do Animal exige ISO 11784; Regulamento exige ISO 11784/11785 + 24631-1. ISO 11785: estabelece padrão de radiofrequência para microchips. ISO 24631-1: metodologia de avaliação de conformidade. Ambas são NOVIDADES do Regulamento não presentes na legislação portuguesa anterior. @RGBEAC menciona SIAC (Sistema de Informação) mas não detalha requisitos técnicos ISO. Recomenda-se: (i) validar conformidade microchips portugueses com ISO 11785 e 24631-1; (ii) atualizar requisitos de aprovação/certificação de equipamentos; (iii) harmonizar prazos de identificação (3 meses vs. 30 dias por tipo operador)."
    },
    {
        "id": "ANNEX-III",
        "tema": "Recolha de Dados sobre Bem-Estar Animal — Relatório Trienal à Comissão UE",
        "regulamento": {
            "ref": "ANEXO III do Regulamento 2023/0447 (conforme Article 24)",
            "titulo": "Collection of data",
            "texto": "1. Number of dogs and cats registered per year, as referred to in Article 17 and Article 21(4).\n\n1a. Number of breeding establishments, selling establishments, shelters, and foster homes registered per year in accordance with Article 7.\n\n2. Number of breeding establishments approved per year, as referred to in Article 7a.\n\n2a. Number of breeding establishments whose approval has been suspended or withdrawn per year.",
            "traducao": "1. Número de cães e gatos registados por ano, conforme referido nos artigos 17.º e 21.º, n.º 4.\n\n1a. Número de estabelecimentos de criação, estabelecimentos de venda, abrigos e famílias de acolhimento registados por ano, em conformidade com o artigo 7.º.\n\n2. Número de estabelecimentos de criação aprovados por ano, conforme referido no artigo 7a.º.\n\n2a. Número de estabelecimentos de criação cuja aprovação foi suspensa ou revogada por ano."
        },
        "rgbeac": {
            "ref": "RGBEAC (Regime Geral do Bem-Estar dos Animais de Companhia, proposta jun. 2025) — Arts. 20.º, 46.º",
            "texto": "Artigo 20.º — Sistema de Informação de Animais de Companhia (SIAC)\n\n1 - O SIAC reúne informação relativa à: a) Rastreabilidade de dispositivos de identificação; b) Identificação dos animais de companhia; c) Sua titularidade ou detenção; d) Informação de saúde pública, saúde animal e bem-estar animal.\n\n2 – A DGAV é a entidade responsável pelo SIAC, assegurando seu funcionamento e tratamento seguro de informação.\n\n3 – A DGAV pode atribuir gestão do SIAC a outras entidades, mediante protocolo e parecer da Comissão Nacional de Proteção de Dados.\n\n4 - Normas e procedimentos constam de Manual de Procedimentos SIAC, aprovado pelo diretor-geral de alimentação e veterinária.\n\nArtigo 46.º — Relatório Nacional Anual\n\n1 - Centros de bem-estar animal e alojamentos com fins de promoção remetem à DGAV no primeiro mês de cada ano, relatórios de gestão do ano anterior.\n\n2 – A DGAV consolida informação a nível nacional sobre bem-estar animal em cada ano, incluindo: a) Acompanhamento da política internacional; b) Prestação de apoios públicos; c) Atividade do SIAC; d) Resultados de planos de controlo; e) Processos contraordenacionais; f) Coimas e sanções; g) Atividade de centros de bem-estar animal; h) Programas de interesse nacional."
        },
        "codigo": {
            "ref": "Código do Animal (DL 214/2013) — Arts. 55.º, 56.º, 141.º-145.º",
            "texto": "Artigo 55.º — Base de dados\n\n1 - Toda informação de registo coligida numa aplicação informática nacional.\n\n2 - A DGAV detém, define e coordena acesso à BD, podendo autorizar gestão em outras entidades mediante protocolos e parecer da Comissão Nacional de Proteção de Dados.\n\n3 - Só entidades autorizadas pela DGAV têm acesso.\n\nArtigo 56.º — Classificação dos animais\n\nCategorias para registo na BD: a) Cão (designado A); b) Cão Potencialmente Perigoso (designado G); c) Cão Perigoso (designado H); d) Gato (designado I).\n\nArtigos 141.º-145.º — Monitorização e Relatórios\n\nCódigo do Animal estabelece estrutura de: a) Monitorização de bem-estar animal; b) Recolha de dados de infrações; c) Aplicação de sanções; d) Atividades de centros de recolha; e) Estatísticas de animais processados; f) Sistema de contraordenações."
        },
        "legislacao": {
            "ref": "DL 82/2019 (Arts. 2.º-7.º), Lei 27/2016 (Arts. 1.º-9.º)",
            "texto": "[dim]Decreto-Lei n.º 82/2019 de 27 de junho — Bem-Estar de Animais de Companhia\n\nArtigo 2.º — Âmbito de aplicação — Aplicável a cães, gatos e furões detidos para fins não comerciais.\n\nArtigo 3.º — Obrigações de registo — Proprietários de cães e gatos devem proceder ao registo no SIAC com identificação eletrónica obrigatória.\n\nArtigo 4.º — Identificação eletrónica — Via aplicação subcutânea de microchip (ISO 11784 e 11785).\n\nArtigo 5.º — Procedimentos de registo — Procedimentos de registo, atualização e consulta do SIAC estabelecidos em regulação complementar.\n\nArtigo 6.º — Controlos de acesso — Segurança de dados e autorização de acesso ao SIAC sob responsabilidade DGAV.\n\nArtigo 7.º — Rastreabilidade — Sistema garante rastreabilidade de dispositivos de identificação e proprietários.\n\n[dim]Lei n.º 27/2016 de 23 de agosto — Rede de Centros de Recolha e Proibição do Abate\n\nArtigo 1.º — Objeto — Aprova medidas para criação de rede de centros de recolha oficial de animais errantes.\n\nArtigo 2.º — Definições — Centros de recolha, abrigos, famílias de acolhimento — conceitos e categorias.\n\nArtigos 3.º-9.º — Estrutura e Funcionamento — Procedimentos de registo, aprovação, funcionamento e supervisão."
        },
        "divergencia": {
            "legislacao": "ANÁLISE POR PONTO:\n\nPONTO 1 (Cães/gatos registados): ✅ CONFORMIDADE COMPLETA\nSIAC operacional desde 2019 (DL 82/2019) — registo obrigatório e eletrónico. Dados disponíveis anualmente.\n\nPONTO 1a (Estabelecimentos registados): ⚠️ CONFORMIDADE PARCIAL\nLei 27/2016 e DL 82/2019 cobrem criação, venda, abrigos. LACUNA: conceito \"famílias de acolhimento\" não explicitamente codificado em legislação vigente.\n\nPONTO 2 (Estabelecimentos aprovados): ❌ LACUNA CRÍTICA\nLegislação vigente menciona REGISTO de estabelecimentos de criação. FALTA: sistema específico de \"aprovação\" formal (vs. apenas registo). Portarias específicas de aprovação existem para outros setores (científico) — NÃO claramente para criadores de animais de companhia.\n\nPONTO 2a (Suspensão/revogação): ❌ LACUNA CRÍTICA\nDL 82/2019 e Lei 27/2016 NÃO mencionam procedimentos de suspensão ou revogação de aprovação de estabelecimentos.",
            "codigo": "ANÁLISE POR PONTO:\n\nPONTO 1: ✅ CONFORMIDADE COMPLETA\nArts. 55.º-56.º estabelecem base de dados nacional com classificação detalhada de animais. Sistema estruturado.\n\nPONTO 1a: ❌ NÃO IMPLEMENTADO\nCódigo do Animal não estabelece sistema específico de registo de estabelecimentos de criação, venda, abrigos, famílias de acolhimento.\n\nPONTO 2: ❌ NÃO IMPLEMENTADO\nCódigo do Animal não estabelece sistema de aprovação de criadores.\n\nPONTO 2a: ❌ NÃO IMPLEMENTADO\nCódigo do Animal não menciona suspensão/revogação de aprovação.",
            "rgbeac": "ANÁLISE POR PONTO:\n\nPONTO 1: ✅ CONFORMIDADE COMPLETA\nArt. 20.º (SIAC) — registo obrigatório e operacional. Relatório nacional anual possível (Art. 46.º).\n\nPONTO 1a: ⚠️ CONFORMIDADE PARCIAL\nArt. 20.º implicitamente cobre estabelecimentos através de rastreabilidade de dispositivos. REQUER complementação: codificação explícita de registo de \"famílias de acolhimento\".\n\nPONTO 2: ❌ NÃO IMPLEMENTADO\nRGBEAC não menciona sistema de aprovação de criadores.\n\nPONTO 2a: ❌ NÃO IMPLEMENTADO\nRGBEAC não menciona suspensão/revogação de aprovação.",
            "sumario": "CONCLUSÃO GERAL:\n\nLegislação portuguesa implementa PARCIALMENTE os 4 pontos de Annex III (dados para relatório trienal à Comissão UE):\n\n✅ PONTO 1 — COBERTURA COMPLETA: SIAC operacional desde 2019. Dados de cães/gatos registados disponíveis anualmente.\n\n⚠️ PONTO 1a — COBERTURA PARCIAL: Sistema de registo de estabelecimentos existe (Lei 27/2016); conceito \"famílias de acolhimento\" requer codificação explícita.\n\n❌ PONTO 2 — LACUNA ESTRUTURAL: Falta sistema formal de \"aprovação\" de criadores. Existe apenas registo administrativo. Necessário: distinguir registo (mera inscrição) de aprovação (avaliação de conformidade com requisitos técnicos).\n\n❌ PONTO 2a — LACUNA ESTRUTURAL: Falta procedimentos formais de suspensão/revogação de aprovação. Necessário: mecanismo administrativo com critérios de suspensão e revogação.\n\nRECOMENDAÇÕES:\n1. Complementar Lei 27/2016 com definição e registo de \"famílias de acolhimento\" (ponto 1a).\n2. Criar sistema de aprovação prévia de estabelecimentos de criação, com avaliação de conformidade (ponto 2).\n3. Estabelecer procedimentos formais de suspensão/revogação com critérios administrativos (ponto 2a).\n4. Integrar dados coletados em relatório trienal da Comissão UE (conforme Art. 20 do Regulamento)."
        },
        "necessidade_alteracao": "Sim — Estabelecer sistema de aprovação, suspensão e revogação de estabelecimentos de criação + codificar famílias de acolhimento",
        "notas": "ANNEX III define dados OBRIGATÓRIOS para relatório trienal à Comissão UE (Art. 20 do Regulamento). SIAC português fornece dados para ponto 1 (cães/gatos registados) — sistema operacional desde 2019. Pontos 2 e 2a requerem regulação complementar urgente: aprovação formal de criadores (vs. apenas registo) + procedimentos de suspensão/revogação. Ponto 1a (estabelecimentos) — codificação explícita de \"famílias de acolhimento\" como categoria administrativa. Recomenda-se: Lei complementar ou Portaria reguladora para implementação completa dos 4 pontos antes do próximo período de reporte (trienal)."
    }
]


# ---------------------------------------------------------------------------
# CORES (sistema cromático por diploma)
# ---------------------------------------------------------------------------

COR = {
    "regulamento_header": "8AAFCF",   # azul pastel médio (menos intenso)
    "regulamento_body":   "E8F2F8",   # azul pastel muito claro
    "regulamento_trad":   "F4F9FD",   # azul pastel extremamente claro
    "rgbeac_header":      "7FAA8C",   # verde pastel médio (menos intenso)
    "rgbeac_body":        "E8F3E6",   # verde pastel muito claro
    "codigo_header":      "B8956A",   # castanho pastel médio (menos intenso)
    "codigo_body":        "FFF5ED",   # castanho pastel muito claro
    "legislacao_header":  "75A6AE",   # teal pastel médio (menos intenso)
    "legislacao_body":    "F0FBFD",   # teal pastel muito claro
    "divergencia_header": "A689C6",   # roxo pastel médio (menos intenso)
    "divergencia_body":   "F5EEF9",   # roxo pastel muito claro
    "notas_header":       "909090",   # cinza pastel médio (menos intenso)
    "notas_body":         "FAFAFA",   # cinza pastel muito claro
    "tema_header":        "6A7A8A",   # cinza-azul pastel (menos intenso)
    "alternado_a":        "FBFBFB",
    "alternado_b":        "F7F9FE",
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font_white_bold(size=10):
    return Font(color="FFFFFF", bold=True, size=size, name="Calibri")

def font_normal(size=10, bold=False, color="000000"):
    return Font(size=size, bold=bold, color=color, name="Calibri")

def border_thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def wrap_align(horizontal="left", vertical="top"):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=True)

# ---------------------------------------------------------------------------
# EXCEL
# ---------------------------------------------------------------------------

def criar_excel(path):
    wb = openpyxl.Workbook()

    # --- Folha 1: Vista de Reunião (dados completos) ---
    ws = wb.active
    ws.title = "Vista de Reunião"
    ws.sheet_view.showGridLines = False

    # Cabeçalhos
    headers = [
        ("Tema", 18),
        ("Art. @regulamento", 22),
        ("Texto @regulamento (EN)", 55),
        ("Tradução @regulamento (PT)", 55),
        ("Art. @rgbeac", 22),
        ("Texto @rgbeac", 55),
        ("Art. @codigo", 22),
        ("Texto @codigo", 55),
        ("Art. @legislacao", 22),
        ("Texto @legislacao (vigente)", 55),
        ("Div. vs @legislacao", 38),
        ("Div. vs @codigo", 38),
        ("Div. vs @rgbeac", 38),
        ("Sumário / Proposta", 38),
        ("Necessidade de Alteração", 14),
        ("Notas de Reunião", 40),
    ]

    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = fill(COR["tema_header"])
        cell.font = font_white_bold(11)
        cell.alignment = wrap_align("center", "center")
        cell.border = border_thin()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    # Dados
    bg_toggle = [COR["alternado_a"], COR["alternado_b"]]
    for row_idx, art in enumerate(ARTIGOS, start=2):
        bg = bg_toggle[(row_idx) % 2]
        valores = [
            art["tema"],
            art["regulamento"]["ref"],
            art["regulamento"]["texto"],
            art["regulamento"]["traducao"],
            art["rgbeac"]["ref"],
            art["rgbeac"]["texto"],
            art["codigo"]["ref"],
            art["codigo"]["texto"],
            art["legislacao"]["ref"],
            art["legislacao"]["texto"],
            art["divergencia"].get("legislacao", ""),
            art["divergencia"].get("codigo", ""),
            art["divergencia"].get("rgbeac", ""),
            art["divergencia"].get("sumario", ""),
            art["necessidade_alteracao"],
            art["notas"],
        ]
        for col_idx, valor in enumerate(valores, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.alignment = wrap_align()
            cell.border = border_thin()
            # Cor por coluna
            if col_idx in (1,):
                cell.fill = fill(COR["tema_header"])
                cell.font = font_white_bold()
                cell.alignment = wrap_align("center", "center")
            elif col_idx in (2, 3, 4):
                cell.fill = fill(COR["regulamento_body"] if col_idx != 4 else COR["regulamento_trad"])
                cell.font = font_normal()
            elif col_idx in (5, 6):
                cell.fill = fill(COR["rgbeac_body"])
                cell.font = font_normal()
            elif col_idx in (7, 8):
                cell.fill = fill(COR["codigo_body"])
                cell.font = font_normal()
            elif col_idx in (9, 10):
                cell.fill = fill(COR["legislacao_body"])
                cell.font = font_normal()
            elif col_idx == 11:
                cell.fill = fill(COR["divergencia_body"])
                cell.font = font_normal(bold=True)
            elif col_idx == 12:
                cell.fill = fill(COR["divergencia_body"])
                cell.font = Font(bold=True, color="8B0000", size=10, name="Calibri")
                cell.alignment = wrap_align("center", "center")
            elif col_idx == 13:
                cell.fill = fill(COR["notas_body"])
                cell.font = font_normal(color="4A4A4A")
        ws.row_dimensions[row_idx].height = 120

    # Freeze panes
    ws.freeze_panes = "B2"

    # --- Folha 2: Legenda / Sistema de Cores ---
    ws2 = wb.create_sheet("Legenda")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 50

    legenda_items = [
        ("DIPLOMA", "COR", True),
        ("@regulamento (texto EN)", COR["regulamento_body"], False),
        ("@regulamento (tradução PT)", COR["regulamento_trad"], False),
        ("@rgbeac", COR["rgbeac_body"], False),
        ("@codigo", COR["codigo_body"], False),
        ("@legislacao (legislação vigente)", COR["legislacao_body"], False),
        ("Divergência", COR["divergencia_body"], False),
        ("Notas de Reunião", COR["notas_body"], False),
    ]
    for i, (label, cor_ou_header, is_header) in enumerate(legenda_items, start=1):
        a = ws2.cell(row=i, column=1, value=label)
        b = ws2.cell(row=i, column=2, value="" if is_header else f"Cor de fundo: #{cor_ou_header}")
        if is_header:
            for c in (a, b):
                c.fill = fill(COR["tema_header"])
                c.font = font_white_bold(12)
                c.alignment = wrap_align("center", "center")
        else:
            a.fill = fill(cor_ou_header)
            b.fill = fill(cor_ou_header)
            a.font = font_normal(bold=True)
            b.font = font_normal()
        for c in (a, b):
            c.border = border_thin()
        ws2.row_dimensions[i].height = 22

    wb.save(path)
    print(f"Excel guardado: {path}")


# ---------------------------------------------------------------------------
# HTML
# ---------------------------------------------------------------------------

def criar_html(path, artigos):
    """Gera visualizador HTML interativo para reunião."""

    # Serializar artigos para JSON (injetado no HTML)
    dados_json = json.dumps(artigos, ensure_ascii=False, indent=2)

    html = f"""<!DOCTYPE html>
<html lang="pt">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Comparativo Artigo a Artigo — Regulamento 2023/0447</title>
<style>
  :root {{
    --reg:    #8AAFCF;
    --reg-bg: #F0F7FC;
    --reg-tr: #F7FAFD;
    --rgb:    #7FAA8C;
    --rgb-bg: #F0F7F4;
    --cod:    #B8956A;
    --cod-bg: #FFF9F2;
    --leg:    #75A6AE;
    --leg-bg: #F5FCFD;
    --div:    #A689C6;
    --div-bg: #F9F6FD;
    --nota:   #909090;
    --nota-bg:#FAFAFA;
    --dark:   #2C3E50;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI', Calibri, sans-serif; background: #F8F9FB; color: #333; }}

  /* HEADER */
  header {{
    background: var(--dark);
    color: #fff;
    padding: 18px 32px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    position: sticky; top: 0; z-index: 100;
    border-bottom: 1px solid rgba(0,0,0,0.08);
  }}
  header h1 {{ font-size: 1.1rem; font-weight: 700; letter-spacing: .5px; }}
  header span {{ font-size: .85rem; opacity: .7; }}

  /* LAYOUT */
  .layout {{ display: flex; min-height: calc(100vh - 60px); }}

  /* SIDEBAR */
  nav {{
    width: 220px; min-width: 220px;
    background: var(--dark);
    padding: 24px 0;
    position: sticky; top: 60px;
    height: calc(100vh - 60px);
    overflow-y: auto;
  }}
  nav h2 {{ color: #fff; font-size: .7rem; text-transform: uppercase;
            letter-spacing: 1.5px; padding: 0 20px 12px; font-weight: 700; }}
  nav button {{
    display: block; width: 100%;
    background: rgba(255,255,255,.04); border: none;
    color: #fff; text-align: left;
    padding: 14px 20px; cursor: pointer;
    font-size: .95rem; line-height: 1.45;
    border-left: 3px solid transparent;
    transition: all .15s;
    font-weight: 500;
  }}
  nav button:hover {{ background: rgba(255,255,255,.12); color: #fff; }}
  nav button.active {{
    background: rgba(255,255,255,.15);
    color: #fff; font-weight: 700;
    border-left-color: #8AAFCF;
  }}
  nav button small {{ display: block; font-size: .74rem; opacity: .85; margin-top: 2px; color: #e0e0e0; }}

  /* MAIN */
  main {{ flex: 1; padding: 28px 32px; overflow-x: auto; }}

  /* BADGE DE ARTIGO */
  .art-badge {{
    display: inline-block;
    background: var(--dark); color: #fff;
    border-radius: 6px; padding: 4px 14px;
    font-size: .8rem; font-weight: 700;
    margin-bottom: 18px; letter-spacing: .5px;
  }}
  .tema-title {{
    font-size: 1.4rem; font-weight: 700;
    color: var(--dark); margin-bottom: 22px;
    padding-bottom: 8px;
    border-bottom: 1px solid rgba(138,175,207,0.25);
  }}

  /* FORMATAÇÃO DE TEXTO LEGAL */
  .leg-block + .leg-block {{ margin-top: 7px; }}
  .leg-p {{ margin: 0; line-height: 1.65; text-align: justify; }}
  .leg-alinea {{
    margin: 2px 0; line-height: 1.6;
    padding-left: 1.8em; text-indent: -1.8em;
  }}
  .leg-sub {{
    margin: 2px 0; line-height: 1.55;
    padding-left: 3.2em; text-indent: -1.0em;
  }}
  /* Texto secundário (cinza) */
  .leg-block.leg-dim p {{ color: #b0b0b0; }}
  /* Cabeçalho de artigo dentro de coluna (múltiplos artigos por diploma) */
  .leg-art-header {{
    font-weight: 700; font-size: .8rem; margin: 10px 0 3px;
    padding-top: 8px; border-top: 1px solid rgba(0,0,0,0.06);
    display: block;
  }}
  .leg-block:first-child > .leg-art-header:first-child {{
    margin-top: 0; padding-top: 0; border-top: none;
  }}

  /* DIVERGÊNCIA ESTRUTURADA */
  .div-section {{ margin-bottom: 10px; }}
  .div-section:last-child {{ margin-bottom: 0; }}
  .div-tag {{
    display: inline-block;
    border-radius: 3px; padding: 1px 8px;
    font-size: .72rem; font-weight: 700;
    text-transform: uppercase; letter-spacing: .5px;
    margin-right: 6px; vertical-align: middle;
    white-space: nowrap;
  }}
  .div-tag.tag-leg  {{ background: var(--leg); color: #fff; }}
  .div-tag.tag-cod  {{ background: var(--cod); color: #fff; }}
  .div-tag.tag-rgb  {{ background: var(--rgb); color: #fff; }}
  .div-tag.tag-sum  {{ background: var(--div); color: #fff; }}
  .div-text {{ display: inline; font-size: .87rem; line-height: 1.6; }}

  /* DIVERGÊNCIA EM GRID (3 colunas + sumário) */
  .divergencia-grid {{
    display: grid;
    grid-template-columns: 1fr 1fr 1fr;
    gap: 12px;
    margin-bottom: 12px;
  }}
  @media (max-width: 1100px) {{ .divergencia-grid {{ grid-template-columns: 1fr 1fr; }} }}
  @media (max-width: 700px)  {{ .divergencia-grid {{ grid-template-columns: 1fr; }} }}

  .div-col {{
    border-radius: 6px;
    overflow: hidden;
    border: 1px solid rgba(0, 0, 0, 0.08);
    display: flex;
    flex-direction: column;
  }}
  .div-col.leg .div-col-header {{ background: var(--leg); color: #fff; }}
  .div-col.leg .div-col-body   {{ background: var(--leg-bg); }}
  .div-col.cod .div-col-header {{ background: var(--cod); color: #fff; }}
  .div-col.cod .div-col-body   {{ background: var(--cod-bg); }}
  .div-col.rgb .div-col-header {{ background: var(--rgb); color: #fff; }}
  .div-col.rgb .div-col-body   {{ background: var(--rgb-bg); }}

  .div-col-header {{
    padding: 10px 12px;
    font-size: .8rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: .4px;
  }}
  .div-col-body {{
    padding: 12px;
    flex: 1;
    font-size: .87rem;
    line-height: 1.6;
    overflow-y: auto;
    color: #222;
  }}

  .div-sumario {{
    border-radius: 6px;
    overflow: hidden;
    border: 1px solid rgba(0, 0, 0, 0.08);
    background: var(--div-bg);
  }}
  .div-sumario-header {{
    padding: 10px 12px;
    background: var(--div);
    color: #fff;
    font-size: .8rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: .4px;
  }}
  .div-sumario-body {{
    padding: 12px;
    font-size: .87rem;
    line-height: 1.6;
    color: #222;
  }}

  /* GRID 3 COLUNAS (diplomas nacionais) */
  .grid {{ display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 16px; margin-bottom: 20px; }}
  @media (max-width: 1100px) {{ .grid {{ grid-template-columns: 1fr 1fr; }} }}
  @media (max-width: 700px)  {{ .grid {{ grid-template-columns: 1fr; }} }}

  /* CARD */
  .card {{ border-radius: 6px; overflow: hidden; border: 1px solid rgba(0,0,0,0.06); box-shadow: none; }}
  .card-header {{
    padding: 10px 16px; font-size: .78rem;
    font-weight: 700; text-transform: uppercase; letter-spacing: .8px;
    color: #fff; display: flex; align-items: center; gap: 10px;
  }}
  .card-header-ref {{
    font-size: .7rem; font-weight: 400; opacity: .7;
    text-transform: none; letter-spacing: 0;
  }}
  .card-body {{ padding: 14px 16px; font-size: .88rem; line-height: 1.65; }}
  .card-ref {{ font-size: .75rem; font-weight: 700; margin-bottom: 10px; opacity: .75; }}

  .card.reg .card-header {{ background: var(--reg); }}
  .card.reg .card-body   {{ background: var(--reg-bg); }}
  .card.reg-tr .card-header {{ background: #7FAED4; }}
  .card.reg-tr .card-body   {{ background: var(--reg-tr); font-style: italic; }}
  .card.rgb .card-header {{ background: var(--rgb); }}
  .card.rgb .card-body   {{ background: var(--rgb-bg); }}
  .card.cod .card-header {{ background: var(--cod); }}
  .card.cod .card-body   {{ background: var(--cod-bg); }}
  .card.leg .card-header {{ background: var(--leg); }}
  .card.leg .card-body   {{ background: var(--leg-bg); }}

  /* DIVERGÊNCIA */
  .div-box {{
    background: var(--div-bg);
    border-left: 5px solid var(--div);
    border-radius: 0 8px 8px 0;
    padding: 14px 18px;
    margin-bottom: 20px;
    font-size: .88rem; line-height: 1.6;
  }}
  .div-box strong {{ color: var(--div); display: block; margin-bottom: 6px;
                     font-size: .78rem; text-transform: uppercase; letter-spacing: .8px; }}
  .badge-alt {{
    display: inline-block;
    background: #8B0000; color: #fff;
    border-radius: 4px; padding: 2px 10px;
    font-size: .75rem; font-weight: 700; margin-left: 10px;
  }}

  /* NOTAS */
  .notas-box {{ margin-top: 4px; }}
  .notas-box label {{
    display: block; font-size: .78rem; font-weight: 700;
    text-transform: uppercase; letter-spacing: .8px;
    color: var(--nota); margin-bottom: 6px;
  }}
  .notas-box textarea {{
    width: 100%; min-height: 90px;
    border: 2px solid #ccc; border-radius: 6px;
    padding: 10px; font-size: .88rem; font-family: inherit;
    resize: vertical; background: var(--nota-bg);
    transition: border-color .2s;
  }}
  .notas-box textarea:focus {{ outline: none; border-color: var(--div); }}

  /* NAVEGAÇÃO INFERIOR */
  .nav-btns {{
    display: flex; gap: 12px; margin-top: 24px; justify-content: flex-end;
  }}
  .btn {{
    padding: 9px 22px; border: none; border-radius: 6px;
    cursor: pointer; font-size: .9rem; font-weight: 700;
    transition: opacity .15s;
  }}
  .btn:hover {{ opacity: .82; }}
  .btn-prev {{ background: #ddd; color: #333; }}
  .btn-next {{ background: var(--reg); color: #fff; }}
  .btn-export {{ background: var(--rgb); color: #fff; }}

  /* EXPORTAÇÃO */
  #export-msg {{
    display: none; background: #D5E8D4; border: 1px solid var(--rgb);
    border-radius: 6px; padding: 10px 16px; margin-top: 12px;
    font-size: .85rem; color: var(--rgb);
  }}
  pre {{ white-space: pre-wrap; word-break: break-word; }}
  mark {{ background: #FFE066; color: #000; border-radius: 2px; padding: 0 2px; }}

  /* BARRA DE PESQUISA */
  .search-wrap {{ position: relative; }}
  .search-wrap input {{
    padding: 7px 32px 7px 12px;
    border-radius: 6px; border: none;
    font-size: .88rem; width: 240px;
    background: rgba(255,255,255,.14); color: #fff;
    outline: none; transition: background .2s;
  }}
  .search-wrap input::placeholder {{ color: rgba(255,255,255,.4); }}
  .search-wrap input:focus {{ background: rgba(255,255,255,.24); }}
  .search-wrap .clear-btn {{
    position: absolute; right: 8px; top: 50%; transform: translateY(-50%);
    background: none; border: none; color: rgba(255,255,255,.55);
    cursor: pointer; font-size: 1rem; line-height: 1; padding: 0;
    display: none;
  }}
  .search-count {{ color: #90b0c8; font-size: .73rem; padding: 2px 20px 8px; }}
</style>
</head>
<body>

<header>
  <h1>Comparativo Artigo a Artigo — Regulamento 2023/0447 (Cães e Gatos)</h1>
  <div style="display:flex;align-items:center;gap:16px;">
    <div class="search-wrap">
      <input id="search-input" type="search"
             placeholder="🔍 Pesquisar palavra-chave…"
             oninput="pesquisar(this.value)"
             autocomplete="off">
      <button class="clear-btn" id="clear-btn"
              onclick="limparPesquisa()" title="Limpar pesquisa">✕</button>
    </div>
    <span id="progresso"></span>
  </div>
</header>

<div class="layout">
  <nav id="sidebar">
    <h2>Artigos</h2>
  </nav>
  <main id="main-content"></main>
</div>

<script>
const ARTIGOS = {dados_json};

let atual = 0;
let searchTerm = '';

/* ---- PESQUISA ---- */
function pesquisar(q) {{
  searchTerm = q.trim().toLowerCase();
  const clearBtn = document.getElementById('clear-btn');
  if (clearBtn) clearBtn.style.display = searchTerm ? 'block' : 'none';
  if (searchTerm) {{
    const primeiroIdx = ARTIGOS.findIndex(a => artMatch(a, searchTerm));
    if (primeiroIdx >= 0 && !artMatch(ARTIGOS[atual], searchTerm)) {{
      atual = primeiroIdx;
    }}
  }}
  render();
}}

function limparPesquisa() {{
  searchTerm = '';
  const inp = document.getElementById('search-input');
  if (inp) inp.value = '';
  const clearBtn = document.getElementById('clear-btn');
  if (clearBtn) clearBtn.style.display = 'none';
  render();
}}

function artMatch(art, q) {{
  const div = art.divergencia || {{}};
  const campos = [
    art.id, art.tema,
    art.regulamento.ref, art.regulamento.titulo,
    art.regulamento.texto, art.regulamento.traducao,
    art.rgbeac.ref, art.rgbeac.texto,
    art.codigo.ref, art.codigo.texto,
    art.legislacao.ref, art.legislacao.texto,
    div.legislacao, div.codigo, div.rgbeac, div.sumario
  ];
  return campos.some(c => (c || '').toLowerCase().includes(q));
}}

function getSnippet(art, q) {{
  const div = art.divergencia || {{}};
  const campos = [
    art.regulamento.traducao, art.regulamento.texto,
    art.rgbeac.texto, art.codigo.texto,
    art.legislacao.texto,
    div.legislacao, div.codigo, div.rgbeac, div.sumario
  ];
  for (const c of campos) {{
    const idx = (c || '').toLowerCase().indexOf(q);
    if (idx >= 0) {{
      const s = Math.max(0, idx - 28);
      const e = Math.min(c.length, idx + q.length + 40);
      return (s > 0 ? '…' : '') +
             c.slice(s, e).replace(/\\n/g, ' ') +
             (e < c.length ? '…' : '');
    }}
  }}
  return '';
}}

function highlight(str) {{
  if (!searchTerm || !str) return nl2br(str || '');
  const re = new RegExp(searchTerm.replace(/[.*+?^${{}}()|[\\]\\\\]/g, '\\\\$&'), 'gi');
  return nl2br(str.replace(re, m => `<mark>${{m}}</mark>`));
}}

/* Formata texto legal com deteção automática de alíneas e subalíneas */
function hl(s) {{
  if (!searchTerm || !s) return s || '';
  const re = new RegExp(searchTerm.replace(/[.*+?^${{}}()|[\\]\\\\]/g, '\\\\$&'), 'gi');
  return s.replace(re, m => `<mark>${{m}}</mark>`);
}}

function formatarTexto(str) {{
  if (!str) return '';
  const blocos = str.split(/\\n\\n+/);
  return blocos.map(bloco => {{
    const isDim = bloco.startsWith('[dim]');
    const blockText = isDim ? bloco.slice(5) : bloco;
    const dimCls = isDim ? ' leg-dim' : '';
    const linhas = blockText.split('\\n').filter(l => l.trim() !== '');
    if (!linhas.length) return '';
    const ps = linhas.map(linha => {{
      const t = linha.trim();
      if (/^Artigo\\s+\\d+/i.test(t)) {{
        return `<p class="leg-art-header">${{hl(t)}}</p>`;
      }}
      if (/^[—–]\\s/.test(t) || t === '—' || t === '–') {{
        return `<p class="leg-sub">${{hl(t)}}</p>`;
      }}
      if (/^[a-z]\\)/.test(t) || /^\\([a-z-]+\\)/.test(t)) {{
        return `<p class="leg-alinea">${{hl(t)}}</p>`;
      }}
      return `<p class="leg-p">${{hl(t)}}</p>`;
    }}).join('');
    return `<div class="leg-block${{dimCls}}">${{ps}}</div>`;
  }}).join('');
}}

function renderDiv(div) {{
  if (!div) return '';

  // Grid 3 colunas: rgbeac, codigo, legislacao (ordem da legislação equivalente)
  const gridHTML = `
    <div class="divergencia-grid">
      <div class="div-col rgb">
        <div class="div-col-header">@rgbeac</div>
        <div class="div-col-body">${{hl(div.rgbeac || '')}}</div>
      </div>
      <div class="div-col cod">
        <div class="div-col-header">@codigo</div>
        <div class="div-col-body">${{hl(div.codigo || '')}}</div>
      </div>
      <div class="div-col leg">
        <div class="div-col-header">@legislacao</div>
        <div class="div-col-body">${{hl(div.legislacao || '')}}</div>
      </div>
    </div>
  `;

  // Sumário como linha comum abaixo
  const sumarioHTML = `
    <div class="div-sumario">
      <div class="div-sumario-header">Sumário / Proposta</div>
      <div class="div-sumario-body">${{hl(div.sumario || '')}}</div>
    </div>
  `;

  return gridHTML + sumarioHTML;
}}

/* ---- SIDEBAR ---- */
function renderSidebar() {{
  const nav = document.getElementById('sidebar');
  nav.innerHTML = '<h2>Artigos</h2>';
  let nResultados = 0;
  ARTIGOS.forEach((art, i) => {{
    const mostrar = !searchTerm || artMatch(art, searchTerm);
    if (!mostrar) return;
    nResultados++;
    const btn = document.createElement('button');
    const snippet = searchTerm ? getSnippet(art, searchTerm) : '';
    btn.innerHTML =
      `<b>${{art.id}}</b><small>${{art.tema}}</small>` +
      (snippet ? `<em style="font-size:.7rem;opacity:.65;display:block;margin-top:3px;font-style:normal;">${{
        snippet.replace(new RegExp(searchTerm.replace(/[.*+?^${{}}()|[\\]\\\\]/g,'\\\\$&'),'gi'),
                        m => `<mark>${{m}}</mark>`)
      }}</em>` : '');
    btn.className = i === atual ? 'active' : '';
    btn.onclick = () => {{ atual = i; render(); }};
    nav.appendChild(btn);
  }});
  if (searchTerm) {{
    const ct = document.createElement('div');
    ct.className = 'search-count';
    ct.textContent = nResultados > 0
      ? `${{nResultados}} resultado(s) para "${{searchTerm}}"`
      : 'Sem resultados.';
    nav.insertBefore(ct, nav.children[1]);
  }}
}}

/* ---- UTILIDADES ---- */
function nl2br(str) {{
  return (str || '').replace(/\\n/g, '<br>');
}}

function render() {{
  const art = ARTIGOS[atual];
  renderSidebar();
  document.getElementById('progresso').textContent =
    `${{atual + 1}} / ${{ARTIGOS.length}}`;

  document.getElementById('main-content').innerHTML = `
    <div class="art-badge">${{art.id}}</div>
    <div class="tema-title">${{art.tema}}</div>

    <div class="card reg" style="margin-bottom:14px;">
      <div class="card-header">
        @regulamento — ${{art.regulamento.titulo}}
        <span class="card-header-ref">${{art.regulamento.ref}} · Texto original EN</span>
      </div>
      <div class="card-body">
        ${{formatarTexto(art.regulamento.texto)}}
      </div>
    </div>

    <div class="card reg-tr" style="margin-bottom:20px;">
      <div class="card-header">
        @regulamento — Tradução PT-PT
        <span class="card-header-ref">${{art.regulamento.ref}}</span>
      </div>
      <div class="card-body">
        ${{formatarTexto(art.regulamento.traducao)}}
      </div>
    </div>

    <div class="grid">
      <div class="card rgb">
        <div class="card-header">@rgbeac (proposta jun. 2025)</div>
        <div class="card-body">
          <div class="card-ref">${{art.rgbeac.ref}}</div>
          ${{formatarTexto(art.rgbeac.texto)}}
        </div>
      </div>
      <div class="card cod">
        <div class="card-header">@codigo (DL n.º 214/2013)</div>
        <div class="card-body">
          <div class="card-ref">${{art.codigo.ref}}</div>
          ${{formatarTexto(art.codigo.texto)}}
        </div>
      </div>
      <div class="card leg">
        <div class="card-header">@legislacao (legislação vigente)</div>
        <div class="card-body">
          <div class="card-ref">${{art.legislacao.ref}}</div>
          ${{formatarTexto(art.legislacao.texto)}}
        </div>
      </div>
    </div>

    <div class="div-box">
      <strong>Divergência face ao Regulamento
        <span class="badge-alt">Necessidade de alteração: ${{art.necessidade_alteracao}}</span>
      </strong>
      ${{renderDiv(art.divergencia)}}
    </div>

    <div class="notas-box">
      <label>Notas de Reunião</label>
      <textarea id="notas-input" placeholder="Escreve aqui as decisões ou observações da reunião..."
        oninput="ARTIGOS[${{atual}}].notas = this.value">${{art.notas}}</textarea>
    </div>

    <div class="nav-btns">
      ${{atual > 0 ? '<button class="btn btn-prev" onclick="navegar(-1)">← Anterior</button>' : ''}}
      <button class="btn btn-export" onclick="exportarNotas()">Exportar Notas (CSV)</button>
      ${{atual < ARTIGOS.length - 1 ? '<button class="btn btn-next" onclick="navegar(1)">Próximo →</button>' : ''}}
    </div>
    <div id="export-msg"></div>
  `;
}}

function navegar(dir) {{
  atual = Math.max(0, Math.min(ARTIGOS.length - 1, atual + dir));
  render();
  window.scrollTo({{top: 0, behavior: 'smooth'}});
}}

function exportarNotas() {{
  const linhas = [['ID', 'Tema', 'Art. Regulamento', 'Art. RGBEAC', 'Art. Código',
                   'Art. Legislação Vigente', 'Div. @legislacao', 'Div. @codigo',
                   'Div. @rgbeac', 'Sumário', 'Necessidade Alteração', 'Notas de Reunião']];
  ARTIGOS.forEach(a => {{
    const d = a.divergencia || {{}};
    linhas.push([
      a.id, a.tema,
      a.regulamento.ref, a.rgbeac.ref, a.codigo.ref,
      a.legislacao.ref,
      d.legislacao || '', d.codigo || '', d.rgbeac || '', d.sumario || '',
      a.necessidade_alteracao,
      a.notas
    ]);
  }});
  const csv = linhas.map(r => r.map(c => `"${{String(c).replace(/"/g,'""')}}"`).join(',')).join('\\n');
  const blob = new Blob([csv], {{type: 'text/csv;charset=utf-8;'}});
  const url = URL.createObjectURL(blob);
  const a = document.createElement('a');
  a.href = url; a.download = 'notas_reuniao_regulamento.csv';
  document.body.appendChild(a); a.click(); document.body.removeChild(a);
  const msg = document.getElementById('export-msg');
  if (msg) {{ msg.style.display = 'block'; msg.textContent = 'CSV exportado com as notas da reunião.'; }}
}}

// Iniciar
render();
</script>
</body>
</html>
"""
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"HTML guardado: {path}")


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    base = os.path.dirname(os.path.abspath(__file__))
    criar_excel(os.path.join(base, "comparativo_reuniao_exemplo.xlsx"))
    criar_html(os.path.join(base, "comparativo_reuniao_exemplo.html"), ARTIGOS)
    print("Concluído.")
