#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para gerar relatório de legislação vigente em Excel
Formato estruturado para análise comparativa com Regulamento 2023/0447
"""

import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Dados da legislação
legislacao = [
    {
        'referencia': 'Lei n.º 92/95',
        'titulo': 'Protecção aos Animais',
        'data': '12 de Setembro de 1995',
        'tema': 'Proteção Geral',
        'subtema': 'Violência e Proteção Básica',
        'status': 'Vigente',
        'descricao_breve': 'Proíbe violência injustificada contra animais; obriga assistência a animais doentes/feridos',
        'artigos_principais': 'Art. 1-10 (violência proibida)',
        'diplomas_relacionados': 'Lei 8/2017, Lei 69/2014',
        'impacto_regulamento': 'Complementa disposições gerais de bem-estar do Regulamento 2023/0447'
    },
    {
        'referencia': 'Lei n.º 8/2017',
        'titulo': 'Estatuto Jurídico dos Animais',
        'data': '3 de Março de 2017',
        'tema': 'Direitos e Estatuto Jurídico',
        'subtema': 'Reconhecimento como Seres Sensíveis',
        'status': 'Vigente',
        'descricao_breve': 'Reconhece animais como seres vivos sensíveis; altera Código Civil, CPC e CP; criminaliza maus tratos',
        'artigos_principais': 'Alterações ao CC, CPC, CP',
        'diplomas_relacionados': 'Lei 92/95, Lei 69/2014, DL 276/2001',
        'impacto_regulamento': 'Harmoniza com princípios europeus; base para direitos animais em proprietários'
    },
    {
        'referencia': 'DL n.º 276/2001',
        'titulo': 'Proteção de Animais de Companhia',
        'data': '17 de Outubro de 2001',
        'tema': 'Bem-Estar de Animais de Companhia',
        'subtema': 'Regime Legal Geral',
        'status': 'Vigente',
        'descricao_breve': 'Implementa Convenção Europeia para Proteção Animais Companhia; regime especial para animais potencialmente perigosos',
        'artigos_principais': 'Art. 1-20+ (alterações sucessivas)',
        'diplomas_relacionados': 'Lei 95/2017, DL 314/2003, DL 82/2019',
        'impacto_regulamento': 'Base legal de referência para harmonização com Regulamento 2023/0447'
    },
    {
        'referencia': 'DL n.º 314/2003',
        'titulo': 'Raiva e Zoonoses / PNLVERAZ',
        'data': '17 de Dezembro de 2003',
        'tema': 'Raiva e Zoonoses',
        'subtema': 'Programa Nacional de Vigilância',
        'status': 'Vigente',
        'descricao_breve': 'Aprova PNLVERAZ; regula posse/detenção de animais suscetíveis à raiva; limita 3 cães/4 gatos por fogo urbano',
        'artigos_principais': 'Art. 1-50+ (PNLVERAZ)',
        'diplomas_relacionados': 'Portaria 264/2013, Lei 27/2016',
        'impacto_regulamento': 'Suporta saúde pública e bem-estar; alinhado com disposições sanitárias do Regulamento'
    },
    {
        'referencia': 'Lei n.º 27/2016',
        'titulo': 'Rede de Centros de Recolha e Proibição do Abate',
        'data': '23 de Agosto de 2016',
        'tema': 'Animais Errantes',
        'subtema': 'Centros de Recolha Oficial',
        'status': 'Vigente',
        'descricao_breve': 'Aprova medidas para rede de centros recolha oficial; proíbe abate como controlo populacional; promove programas CED',
        'artigos_principais': 'Art. 1-10+ (medidas e rede)',
        'diplomas_relacionados': 'Portaria 146/2017, DL 82/2019',
        'impacto_regulamento': 'Alinha com recomendações europeias sobre bem-estar de animais errantes'
    },
    {
        'referencia': 'Lei n.º 95/2017',
        'titulo': 'Comércio de Animais de Companhia',
        'data': '23 de Agosto de 2017',
        'tema': 'Comércio e Reprodução',
        'subtema': 'Regulação de Venda e Reprodução',
        'status': 'Vigente',
        'descricao_breve': 'Regulamenta compra/venda em estabelecimentos e online; restringe transferência livre; combate abandono',
        'artigos_principais': 'Alteração ao DL 276/2001 (art. 6.ª)',
        'diplomas_relacionados': 'Portaria 67/2018, DL 82/2019',
        'impacto_regulamento': 'Complementa Regulamento 2023/0447 em aspetos de responsabilidade do vendedor'
    },
    {
        'referencia': 'DL n.º 82/2019',
        'titulo': 'Identificação e Registo de Animais (SIAC)',
        'data': '27 de Junho de 2019',
        'tema': 'Identificação e Rastreabilidade',
        'subtema': 'Sistema SIAC',
        'status': 'Vigente',
        'descricao_breve': 'Estabelece regras de identificação; cria SIAC; obrigatoriedade de microchip para cães, gatos e furões',
        'artigos_principais': 'Art. 1-50+ (SIAC)',
        'diplomas_relacionados': 'Lei 95/2017, Portaria 67/2018',
        'impacto_regulamento': 'Totalmente harmónico com Regulamento 2023/0447 em matéria de identificação eletrónica'
    },
    {
        'referencia': 'DL n.º 74/2007',
        'titulo': 'Cães de Assistência',
        'data': '27 de Março de 2007',
        'tema': 'Cães de Assistência',
        'subtema': 'Direito de Acesso e Identificação',
        'status': 'Vigente',
        'descricao_breve': 'Direito de acesso para pessoas com deficiência com cães de assistência; identificação e seguro obrigatórios',
        'artigos_principais': 'Art. 1-20+ (direitos e requisitos)',
        'diplomas_relacionados': 'DL 276/2001, Lei 8/2017',
        'impacto_regulamento': 'Complementar ao Regulamento 2023/0447; aspecto de bem-estar social'
    },
    {
        'referencia': 'DL n.º 202/2004',
        'titulo': 'Lei de Bases Gerais da Caça',
        'data': '18 de Agosto de 2004',
        'tema': 'Caça',
        'subtema': 'Conservação e Bem-estar Cinegético',
        'status': 'Vigente',
        'descricao_breve': 'Regulamenta conservação e exploração sustentável de recursos cinegéticos; bem-estar de cães de caça',
        'artigos_principais': 'Art. 1-100+ (caça)',
        'diplomas_relacionados': 'Portaria 148/2016',
        'impacto_regulamento': 'Fora escopo direto do Regulamento 2023/0447; aspecto de bem-estar específico'
    },
    {
        'referencia': 'DL n.º 255/2009',
        'titulo': 'Animais em Circos e Espetáculos',
        'data': '24 de Setembro de 2009',
        'tema': 'Circos e Espetáculos',
        'subtema': 'Identificação e Proteção',
        'status': 'Vigente',
        'descricao_breve': 'Implementa Reg. (CE) 1739/2005; identificação, registo e proteção de animais em circos e eventos',
        'artigos_principais': 'Art. 1-30+ (circos)',
        'diplomas_relacionados': 'Lei 20/2019',
        'impacto_regulamento': 'Complementar ao Regulamento 2023/0447; aspecto de bem-estar em confinamento'
    },
    {
        'referencia': 'DL n.º 265/2007',
        'titulo': 'Transporte de Animais - Autorização Transportadores',
        'data': '24 de Julho de 2007',
        'tema': 'Transportes',
        'subtema': 'Registação e Autorização',
        'status': 'Vigente',
        'descricao_breve': 'Registação e autorização de transportadores; regulamenta bem-estar em transporte (Reg. CE 1/2005)',
        'artigos_principais': 'Art. 1-20+ (transportadores)',
        'diplomas_relacionados': 'Portaria 968/2009',
        'impacto_regulamento': 'Alinhado com disposições de transporte do Regulamento 2023/0447'
    },
    {
        'referencia': 'Lei n.º 69/2014',
        'titulo': 'Criminalização de Maus Tratos e Abandono',
        'data': '29 de Agosto de 2014',
        'tema': 'Sanções Criminais',
        'subtema': 'Maus Tratos e Abandono',
        'status': 'Vigente',
        'descricao_breve': '33.ª alteração ao CP; criminaliza maus tratos (art. 387) e abandono (art. 388); sanções penais e multas',
        'artigos_principais': 'CP Art. 387-388 (maus tratos, abandono)',
        'diplomas_relacionados': 'Lei 8/2017, Lei 92/95',
        'impacto_regulamento': 'Suporta aplicação coerciva do Regulamento 2023/0447; questão constitucional em debate'
    },
    {
        'referencia': 'Lei n.º 20/2019',
        'titulo': 'Proteção Reforçada de Animais em Circos',
        'data': '22 de Fevereiro de 2019',
        'tema': 'Circos e Espetáculos',
        'subtema': 'Proibição de Animais Selvagens',
        'status': 'Vigente',
        'descricao_breve': 'Reforça proteção; proíbe animais selvagens em circos (período transitório até 2025); registação obrigatória',
        'artigos_principais': 'Art. 1-10+ (proibições)',
        'diplomas_relacionados': 'DL 255/2009',
        'impacto_regulamento': 'Alinha com princípios europeus; aspecto de bem-estar avançado'
    },
    {
        'referencia': 'Dec. Regulamentar n.º 3/2021',
        'titulo': 'Provedor do Animal',
        'data': '25 de Junho de 2021',
        'tema': 'Administração e Supervisão',
        'subtema': 'Órgão Independente',
        'status': 'Vigente',
        'descricao_breve': 'Institui Provedor do Animal com autonomia administrativa; defende bem-estar; monitora conformidade',
        'artigos_principais': 'Art. 1-50+ (Provedor)',
        'diplomas_relacionados': 'Lei 8/2017',
        'impacto_regulamento': 'Aspecto de governança; mecanismo de supervisão de aplicação do Regulamento 2023/0447'
    },
    {
        'referencia': 'Portaria n.º 146/2017',
        'titulo': 'Regulamentação de Centros de Recolha',
        'data': '26 de Abril de 2017',
        'tema': 'Animais Errantes',
        'subtema': 'Normas Técnicas e Operação',
        'status': 'Vigente',
        'descricao_breve': 'Regulamenta Lei 27/2016; normas técnicas para centros; destino de animais; regras CED',
        'artigos_principais': 'Art. 1-30+ (normas técnicas)',
        'diplomas_relacionados': 'Lei 27/2016, DL 82/2019',
        'impacto_regulamento': 'Operacionaliza bem-estar de animais errantes conforme Regulamento 2023/0447'
    },
    {
        'referencia': 'Portaria n.º 148/2016',
        'titulo': 'Registo de Matilhas de Caça Maior',
        'data': '2016',
        'tema': 'Caça',
        'subtema': 'Rastreabilidade de Cães de Caça',
        'status': 'Vigente',
        'descricao_breve': 'Obrigatoriedade de registo de cães em matilhas; identificação de proprietários e matilheiros',
        'artigos_principais': 'Art. 1-20+ (registo e obrigações)',
        'diplomas_relacionados': 'DL 202/2004',
        'impacto_regulamento': 'Aspecto de rastreabilidade conforme Regulamento 2023/0447'
    },
    {
        'referencia': 'Portaria n.º 264/2013',
        'titulo': 'Programa Nacional PNLVERAZ',
        'data': '16 de Agosto de 2013',
        'tema': 'Raiva e Zoonoses',
        'subtema': 'Normas Técnicas de Execução',
        'status': 'Vigente',
        'descricao_breve': 'Aprova normas técnicas do PNLVERAZ; vacinação antirrábica obrigatória (cães >3 meses); campanhas',
        'artigos_principais': 'Art. 1-40+ (vacinação e vigilância)',
        'diplomas_relacionados': 'DL 314/2003',
        'impacto_regulamento': 'Suporta saúde pública e bem-estar conforme Regulamento 2023/0447'
    },
    {
        'referencia': 'Portaria n.º 67/2018',
        'titulo': 'Regulação de Criadouros Comerciais',
        'data': '8 de Março de 2018',
        'tema': 'Comércio e Reprodução',
        'subtema': 'Normas Técnicas de Criadores',
        'status': 'Vigente',
        'descricao_breve': 'Implementa Lei 95/2017; regras para compra/venda; normas de reprodução comercial; registação',
        'artigos_principais': 'Art. 1-25+ (criadouros)',
        'diplomas_relacionados': 'Lei 95/2017, DL 82/2019',
        'impacto_regulamento': 'Complementa Regulamento 2023/0447 em aspetos de responsabilidade do criador'
    },
    {
        'referencia': 'Portaria n.º 968/2009',
        'titulo': 'Transporte em Meios de Transporte Público',
        'data': '26 de Agosto de 2009',
        'tema': 'Transportes',
        'subtema': 'Meios Públicos Rodoviários, Ferroviários, Fluviais',
        'status': 'Vigente',
        'descricao_breve': 'Regulamenta deslocações em meios públicos; exigências de contentor; animais acompanhados; restrições',
        'artigos_principais': 'Art. 1-20+ (condições transporte)',
        'diplomas_relacionados': 'DL 265/2007',
        'impacto_regulamento': 'Alinhado com bem-estar em transporte conforme Regulamento 2023/0447'
    }
]

# Criar workbook
wb = Workbook()
ws = wb.active
ws.title = "Legislação Vigente"

# Estilos
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Definir colunas e larguras
colunas = [
    ('Referência', 15),
    ('Título Completo', 35),
    ('Data', 20),
    ('Tema Principal', 20),
    ('Sub-tema', 20),
    ('Status', 10),
    ('Descrição Breve', 40),
    ('Artigos Principais', 25),
    ('Diplomas Relacionados', 25),
    ('Impacto Regulamento 2023/0447', 35)
]

# Cabeçalho
for col_num, (col_name, col_width) in enumerate(colunas, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = col_name
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border
    ws.column_dimensions[get_column_letter(col_num)].width = col_width

# Dados
for row_num, diploma in enumerate(legislacao, 2):
    ws.cell(row=row_num, column=1).value = diploma['referencia']
    ws.cell(row=row_num, column=2).value = diploma['titulo']
    ws.cell(row=row_num, column=3).value = diploma['data']
    ws.cell(row=row_num, column=4).value = diploma['tema']
    ws.cell(row=row_num, column=5).value = diploma['subtema']
    ws.cell(row=row_num, column=6).value = diploma['status']
    ws.cell(row=row_num, column=7).value = diploma['descricao_breve']
    ws.cell(row=row_num, column=8).value = diploma['artigos_principais']
    ws.cell(row=row_num, column=9).value = diploma['diplomas_relacionados']
    ws.cell(row=row_num, column=10).value = diploma['impacto_regulamento']
    
    # Aplicar estilos a todas as células
    for col_num in range(1, len(colunas) + 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

# Congelar cabeçalho
ws.freeze_panes = "A2"

# Salvar
wb.save('/home/user/Legislacao/legislacao_vigente_animais_completa.xlsx')
print("✓ Ficheiro Excel gerado: legislacao_vigente_animais_completa.xlsx")
print(f"✓ Total de diplomas catalogados: {len(legislacao)}")
print("✓ Estrutura: Referência | Título | Data | Tema | Sub-tema | Status | Descrição | Artigos | Relacionados | Impacto Regulamento")

